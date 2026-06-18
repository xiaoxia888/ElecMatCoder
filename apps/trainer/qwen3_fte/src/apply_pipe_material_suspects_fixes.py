#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Apply reviewed MATERIAL fixes from the suspect analysis workbook back to the pipe draft dataset."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[4]
BASE = ROOT / "apps/trainer/qwen3_fte/output/pipe_project_sampling_full"
DATASET = BASE / "直管训练草稿_语义补强版.json"
REVIEW_XLSX = BASE / "直管材质可疑样本分析_改进版.xlsx"
CHANGE_XLSX = BASE / "直管材质批量修正明细.xlsx"


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def as_list(value: Any) -> list[Any]:
    if value in (None, ""):
        return []
    return value if isinstance(value, list) else [value]


def split_csv(value: str) -> list[str]:
    raw = text(value)
    if not raw:
        return []
    out: list[str] = []
    seen: set[str] = set()
    for part in raw.split(","):
        part = part.strip()
        if not part or part in seen:
            continue
        seen.add(part)
        out.append(part)
    return out


def material_items(item: dict[str, Any]) -> list[dict[str, Any]]:
    mats = item.get("output", {}).get("MATERIAL", []) or []
    out: list[dict[str, Any]] = []
    for m in mats:
        if isinstance(m, dict):
            out.append(m)
        elif text(m):
            out.append(
                {
                    "ROLE": "MAIN",
                    "VALUE": text(m),
                    "COATING": {"INNER": [], "OUTER": []},
                    "SPECIAL_REQ": [],
                }
            )
    return out


def normalize_material_container(item: dict[str, Any]) -> list[dict[str, Any]]:
    mats = material_items(item)
    item.setdefault("output", {})["MATERIAL"] = mats
    return mats


def load_review_rows() -> list[dict[str, Any]]:
    wb = load_workbook(REVIEW_XLSX, read_only=True)
    ws = wb["可疑材质"]
    rows: list[dict[str, Any]] = []
    headers = [text(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[i]: values[i] for i in range(len(headers))}
        rows.append(row)
    return rows


def save_change_report(changes: list[list[Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "修正明细"
    headers = [
        "记录序号",
        "材质项序号",
        "问题类型",
        "原VALUE",
        "新VALUE",
        "原SPECIAL_REQ",
        "新SPECIAL_REQ",
        "原文描述",
    ]
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    for row in changes:
        ws.append(row)
    widths = {"A": 10, "B": 10, "C": 28, "D": 20, "E": 24, "F": 18, "G": 18, "H": 120}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    wb.save(CHANGE_XLSX)


def main() -> None:
    data = json.loads(DATASET.read_text(encoding="utf-8"))
    review_rows = load_review_rows()
    changes: list[list[Any]] = []
    applied = 0
    skipped = 0

    for row in review_rows:
        rec_idx = int(row["记录序号"]) - 1
        mat_idx = int(row["材质项序号"]) - 1
        suggested_value = text(row.get("建议VALUE"))
        suggested_special = split_csv(text(row.get("建议SPECIAL_REQ")))
        issue = text(row.get("问题类型"))

        if rec_idx < 0 or rec_idx >= len(data):
            skipped += 1
            continue
        mats = normalize_material_container(data[rec_idx])
        if mat_idx < 0 or mat_idx >= len(mats):
            skipped += 1
            continue
        mat = mats[mat_idx]
        old_value = text(mat.get("VALUE"))
        old_special = [text(x) for x in as_list(mat.get("SPECIAL_REQ")) if text(x)]

        if not suggested_value:
            skipped += 1
            continue

        changed = False
        if old_value != suggested_value:
            mat["VALUE"] = suggested_value
            changed = True

        if suggested_special != old_special:
            mat["SPECIAL_REQ"] = suggested_special
            changed = True

        if changed:
            applied += 1
            changes.append(
                [
                    rec_idx + 1,
                    mat_idx + 1,
                    issue,
                    old_value,
                    suggested_value,
                    ",".join(old_special),
                    ",".join(suggested_special),
                    text(data[rec_idx].get("input")),
                ]
            )

    DATASET.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    save_change_report(changes)
    print(f"dataset={DATASET}")
    print(f"review={REVIEW_XLSX}")
    print(f"change_report={CHANGE_XLSX}")
    print(f"review_rows={len(review_rows)}")
    print(f"applied={applied}")
    print(f"skipped={skipped}")


if __name__ == "__main__":
    main()
