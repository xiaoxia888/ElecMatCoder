#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import math
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


DEFAULT_INPUT = Path("/Users/guoxi/Desktop/数据集.xlsx")
DEFAULT_OUTPUT_DIR = Path("apps/trainer/qwen3_fte/output/elbow_tee_project_sampling")

FIELD_NAMES = [
    "材料描述",
    "项目名称",
    "分类",
    "国标美标标记",
    "编码",
    "材料名称代码",
    "公称直径代码",
    "壁厚等级代码",
    "压力等级代码",
    "材质代码",
    "标准号代码",
    "材质分类",
]

TEE_CODES = {
    "T", "TS", "TW", "TN", "TFN", "TFT", "TP", "TSF", "TJ", "TPL",
    "RT", "RTS", "RTW", "RTN", "RTFN", "RTP", "RTT", "RTR", "RTE", "RTSF", "RTJ",
    "45T", "45RT", "45LT", "45LTS", "45LTW", "45RLT", "45RLTS", "45RLTW",
    "YLT", "IT", "FT", "FRT", "CRTW",
}

STD_RE = re.compile(r"\b(?:GB/T|HG/T|SH/T|NB/T|SY/T|ASME|ASTM|MSS|DIN|EN|ISO|JIS|API|AISI)\s*[-/]?\s*[A-Z0-9.()/-]*", re.IGNORECASE)
PRESSURE_RE = re.compile(r"\b(?:CL|CLASS|PN)\s*\.?\s*\d+(?:/\s*PN\d+)?\b|\b\d+#\b|\b\d+LB\b", re.IGNORECASE)
THK_RE = re.compile(r"\b(?:SCH\.?\s*\w+|S-\d+\w*|\d+(?:\.\d+)?\s*MM|THK\s*=\s*\d+(?:\.\d+)?\s*MM|XXS|XS|STD)\b", re.IGNORECASE)
SIZE_RE = re.compile(
    r"\b(?:DN|NPS|NB)\s*\d+\b|"
    r"[Φφ]\s*\d+(?:\.\d+)?\s*[xX×*]\s*\d+(?:\.\d+)?(?:\s*[/xX×*]\s*[Φφ]?\s*\d+(?:\.\d+)?)?|"
    r"\b\d+(?:\.\d+)?\s*[xX×*]\s*\d+(?:\.\d+)?(?:\s*[xX×*]\s*\d+(?:\.\d+)?)?\b|"
    r"\b\d+(?:\.\d+)?\s*\"\b",
    re.IGNORECASE,
)
MATERIAL_TOKEN_RE = re.compile(
    r"\b(?:A105|A106|A234|A403|A420|A815|A182|A516|A694|LF2|WPB|WPL6|WP\d+|"
    r"Q235B|Q245R|Q345R|CF415K?|L245N?|L290N|X65|"
    r"06Cr\d+|022Cr\d+|S\d{5}|SF\d+|TP\d+|16MnD?|20#?|20G|304L?|316L?|321|2205|2507|"
    r"PTFE|RPTFE|FRP/PVC|FRP/CPVC|FRP)\b",
    re.IGNORECASE,
)
NUMBER_RE = re.compile(r"\d+(?:\.\d+)?")


@dataclass
class RowItem:
    row_no: int
    values: dict[str, Any]
    type_code: str
    category: str
    pattern_signature: str
    size_combo: str
    thk_combo: str
    pressure: str
    material: str
    standard: str


def text(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def uniq_keep_order(values: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        if not value or value in seen:
            continue
        seen.add(value)
        out.append(value)
    return out


def normalize_description(desc: str) -> str:
    s = unicodedata.normalize("NFKC", desc or "").upper()
    s = s.replace("\n", " ").replace("\r", " ")
    replacements = [
        (STD_RE, "<STD>"),
        (PRESSURE_RE, "<PRESSURE>"),
        (THK_RE, "<THK>"),
        (SIZE_RE, "<SIZE>"),
        (MATERIAL_TOKEN_RE, "<MAT>"),
        (NUMBER_RE, "<N>"),
    ]
    for pattern, token in replacements:
        s = pattern.sub(token, s)
    return re.sub(r"\s+", " ", s).strip(" ,;")


def infer_type_code(values: dict[str, Any]) -> str:
    direct = text(values.get("材料名称代码"))
    if direct:
        return direct
    code = text(values.get("编码"))
    match = re.match(r"[A-Z0-9.]+", code.upper())
    return match.group(0) if match else "<EMPTY>"


def is_elbow_code(code: str) -> bool:
    c = code.upper()
    if re.fullmatch(r"(?:F|LF)?\d+(?:\.\d+)?(?:EL|ES)[A-Z0-9.]*", c):
        return True
    return False


def classify_code(code: str) -> str:
    c = code.upper()
    if c in TEE_CODES:
        return "三通"
    if is_elbow_code(c):
        return "弯头"
    return ""


def read_rows(path: Path) -> tuple[list[RowItem], Counter[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    row_iter = ws.iter_rows(values_only=True)
    header = [text(x) for x in next(row_iter)]
    idx = {name: i for i, name in enumerate(header)}
    missing = [name for name in FIELD_NAMES if name not in idx]
    if missing:
        raise KeyError(f"Excel 缺少列: {missing}")

    out: list[RowItem] = []
    skipped: Counter[str] = Counter()
    for row_no, row in enumerate(row_iter, start=2):
        values = {name: row[idx[name]] for name in FIELD_NAMES}
        if text(values["分类"]) != "管件":
            continue
        type_code = infer_type_code(values)
        category = classify_code(type_code)
        if not category:
            skipped[type_code] += 1
            continue
        desc = text(values["材料描述"])
        out.append(
            RowItem(
                row_no=row_no,
                values=values,
                type_code=type_code,
                category=category,
                pattern_signature=normalize_description(desc),
                size_combo=text(values["公称直径代码"]),
                thk_combo=text(values["壁厚等级代码"]),
                pressure=text(values["压力等级代码"]),
                material=text(values["材质代码"]),
                standard=text(values["标准号代码"]),
            )
        )
    return out, skipped


def rarity_bonus(value: Any, counter: Counter[Any]) -> float:
    if not value:
        return 0.0
    freq = counter.get(value, 0)
    return 0.0 if freq <= 0 else 1.0 / math.sqrt(freq)


def select_group_rows(items: list[RowItem], limit: int) -> list[tuple[RowItem, str]]:
    pressure_counter = Counter(x.pressure for x in items if x.pressure)
    material_counter = Counter(x.material for x in items if x.material)
    standard_counter = Counter(x.standard for x in items if x.standard)
    size_thk_counter = Counter((x.size_combo, x.thk_combo) for x in items if x.size_combo or x.thk_combo)

    covered_pressures: set[str] = set()
    covered_materials: set[str] = set()
    covered_standards: set[str] = set()
    covered_size_thk: set[tuple[str, str]] = set()
    selected: list[tuple[RowItem, str]] = []
    remaining = items[:]

    while remaining and len(selected) < limit:
        best_item: RowItem | None = None
        best_score = -1.0
        best_reason = ""
        for item in remaining:
            score = 0.0
            reasons: list[str] = ["同原文模板补充覆盖"]
            if item.pressure and item.pressure not in covered_pressures:
                score += 15
                reasons.append(f"新压力={item.pressure}")
            if item.material and item.material not in covered_materials:
                score += 14
                reasons.append(f"新材质={item.material}")
            if item.standard and item.standard not in covered_standards:
                score += 14
                reasons.append(f"新规范={item.standard}")
            size_thk = (item.size_combo, item.thk_combo)
            if size_thk not in covered_size_thk and (item.size_combo or item.thk_combo):
                score += 12
                reasons.append(f"新尺寸壁厚组合={item.size_combo or '-'}|{item.thk_combo or '-'}")
            score += 3 * rarity_bonus(item.pressure, pressure_counter)
            score += 3 * rarity_bonus(item.material, material_counter)
            score += 3 * rarity_bonus(item.standard, standard_counter)
            score += 2 * rarity_bonus(size_thk, size_thk_counter)
            if score > best_score:
                best_item = item
                best_score = score
                best_reason = "；".join(reasons)
        assert best_item is not None
        selected.append((best_item, best_reason))
        remaining.remove(best_item)
        if best_item.pressure:
            covered_pressures.add(best_item.pressure)
        if best_item.material:
            covered_materials.add(best_item.material)
        if best_item.standard:
            covered_standards.add(best_item.standard)
        if best_item.size_combo or best_item.thk_combo:
            covered_size_thk.add((best_item.size_combo, best_item.thk_combo))
    return selected


def infer_manu(type_code: str, desc: str) -> list[str]:
    desc_u = desc.upper()
    if re.search(r"SEAMLESS\s+OR\s+WELDED|SMLS\s+OR\s+WELDED|无缝\s*或\s*焊接|无缝\s*/\s*焊接", desc_u):
        return ["WELDED"]

    values: list[str] = []
    for token in ["SAWL", "SAWH", "ERW", "EFW", "HFW", "LSAW"]:
        if token in desc_u:
            values.append(token)
    if "SMLS" in desc_u or "SEAMLESS" in desc_u or "无缝" in desc:
        values.insert(0, "SMLS")
    if "SMLS" not in values and any(token in desc_u or token in desc for token in ["WELDED", "WELD ", "有缝", "焊接", "焊制"]):
        values.append("WELDED")

    c = type_code.upper()
    if c.endswith("W") and not values:
        values.append("WELDED")
    return uniq_keep_order(values)


def infer_conn(type_code: str, desc: str) -> list[str]:
    desc_u = desc.upper()
    c = type_code.upper()
    values: list[str] = []
    if re.search(r"(?<![A-Z0-9])SW(?![A-Z0-9])", desc_u) or "承插焊" in desc or "插焊" in desc:
        values.append("SW")
    if re.search(r"(?<![A-Z0-9])SF(?![A-Z0-9])", desc_u):
        values.append("SF")
    if "TH(NPT)" in desc_u or (re.search(r"(?<![A-Z0-9])NPT(?![A-Z0-9])", desc_u) and not any(x in desc_u for x in ["FNPT", "MNPT", "NPTF"])):
        values.append("NPT")
    if "THD" in desc_u or "THREADED" in desc_u or "螺纹" in desc:
        values.append("THD")
    if c.endswith("S") and not c.endswith("ES") and not c.endswith("ELS") and "SW" in desc_u:
        values.append("SW")
    return uniq_keep_order(values)


def infer_ends(type_code: str, desc: str) -> list[str]:
    desc_u = desc.upper()
    values: list[str] = []
    for token in ["FNPT", "MNPT", "NPTF", "MTE", "TSE", "FTE"]:
        if token in desc_u:
            values.append(token)
    c = type_code.upper()
    if c.endswith("FN"):
        values.append("FNPT")
    if c.endswith("FT") or c.endswith("TE"):
        values.append("FTE")
    return uniq_keep_order(values)


def infer_elbow_type(type_code: str, desc: str) -> dict[str, Any]:
    c = type_code.upper()

    body = "弯头"
    if "夹套" in desc or c.endswith("J"):
        body = "夹套弯头"
    elif c.startswith("F") or c.startswith("LF"):
        body = "法兰弯头"
    elif "异径弯头" in desc or "ELR" in c:
        body = "异径弯头"

    return {
        "BODY": body,
        "MANU": infer_manu(c, desc),
        "CONN": infer_conn(c, desc),
        "ENDS": infer_ends(c, desc),
    }


def infer_tee_body(type_code: str, desc: str) -> str:
    c = type_code.upper()
    desc_u = desc.upper()

    if c == "IT" or "INSTRUMENT TEE" in desc_u:
        return "仪表三通"
    if "夹套" in desc or c in {"TJ", "RTJ"}:
        return "夹套异径三通" if c.startswith("RT") or "异径" in desc or "REDUC" in desc_u else "夹套等径三通"
    if c in {"FRT", "FT", "TPL"}:
        return "法兰异径三通" if c == "FRT" or "异径" in desc or "REDUC" in desc_u else "法兰等径三通"
    if c == "RTP":
        return "异径三通(纵向剖分成对包装)"
    if c == "TP":
        return "等径三通(纵向剖分成对包装)"
    if c in {"45RLT", "45RLTS", "45RLTW", "45RT"}:
        return "异径斜三通"
    if c in {"45LT", "45LTS", "45LTW", "45T", "YLT"} or "LATERAL" in desc_u or "斜三通" in desc:
        return "斜三通"
    if c.startswith("RT") or c in {"RTE", "CRTW", "RTSF", "RTT", "RTR"}:
        return "异径三通"
    if c.startswith("T") or c in {"TSF"}:
        return "等径三通"
    if "异径" in desc or "REDUC" in desc_u or "RED TEE" in desc_u:
        return "异径三通"
    if "等径" in desc or "STRAIGHT TEE" in desc_u or "EQUAL TEE" in desc_u:
        return "等径三通"
    return "三通"


def infer_tee_type(type_code: str, desc: str) -> dict[str, Any]:
    c = type_code.upper()
    body = infer_tee_body(c, desc)
    return {
        "BODY": body,
        "MANU": infer_manu(c, desc),
        "CONN": infer_conn(c, desc),
        "ENDS": infer_ends(c, desc),
    }


def build_type(item: RowItem) -> dict[str, Any]:
    desc = text(item.values["材料描述"])
    if item.category == "弯头":
        return infer_elbow_type(item.type_code, desc)
    return infer_tee_type(item.type_code, desc)


def conflict_reason(item: RowItem, type_obj: dict[str, Any]) -> str:
    desc = text(item.values["材料描述"])
    desc_u = desc.upper()
    reasons: list[str] = []
    if item.category == "弯头":
        if any(x in desc_u for x in ["TEE", "LATERAL", "OLET"]) or any(x in desc for x in ["三通", "支管台", "管接台"]):
            reasons.append("弯头编码中出现三通/支管台关键词")
    if item.category == "三通":
        if any(x in desc_u for x in ["ELBOW", "BEND", "OLET"]) or any(x in desc for x in ["弯头", "弯管", "支管台", "管接台"]):
            reasons.append("三通编码中出现弯头/支管台关键词")
    manu = type_obj.get("MANU", [])
    if "SMLS" in manu and any(x in manu for x in ["EFW", "ERW", "SAWL", "SAWH", "WELDED"]):
        reasons.append("MANU 同时包含无缝和焊接制造方式")
    return "；".join(reasons)


def write_sampling_workbook(path: Path, rows: list[dict[str, Any]], pattern_rows: list[dict[str, Any]], type_rows: list[dict[str, Any]], conflict_rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "模板抽样结果"
    headers = ["模型类别", "类型编码", "模板桶内总数", "抽样原因", *FIELD_NAMES, "Excel行号", "原文模板签名"]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    ws2 = wb.create_sheet("模板分布")
    p_headers = ["模型类别", "类型编码", "模板总数", "原文模板签名", "模板条数", "抽样条数", "典型描述1", "典型描述2", "典型描述3"]
    ws2.append(p_headers)
    for row in pattern_rows:
        ws2.append([row.get(h, "") for h in p_headers])

    ws3 = wb.create_sheet("类型分布")
    t_headers = ["模型类别", "类型编码", "总条数", "模板总数", "抽样条数", "典型描述1", "典型描述2", "典型描述3"]
    ws3.append(t_headers)
    for row in type_rows:
        ws3.append([row.get(h, "") for h in t_headers])

    ws4 = wb.create_sheet("疑似冲突")
    c_headers = ["模型类别", "Excel行号", "类型编码", "材料描述", "TYPE", "冲突原因"]
    ws4.append(c_headers)
    for row in conflict_rows:
        ws4.append([row.get(h, "") for h in c_headers])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            width = min(max(len(text(c.value)) for c in col[:30]) + 2, 100)
            sheet.column_dimensions[col[0].column_letter].width = max(12, width)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def dump_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def build_outputs(items: list[RowItem], category: str, per_pattern: int) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    category_items = [x for x in items if x.category == category]
    type_buckets: dict[str, list[RowItem]] = defaultdict(list)
    for item in category_items:
        type_buckets[item.type_code].append(item)

    selected_rows: list[dict[str, Any]] = []
    pattern_rows: list[dict[str, Any]] = []
    type_rows: list[dict[str, Any]] = []
    dataset: list[dict[str, Any]] = []
    conflict_rows: list[dict[str, Any]] = []
    seen_inputs: set[str] = set()

    summary: dict[str, Any] = {"input_rows": len(category_items), "types": {}}
    for type_code in sorted(type_buckets, key=lambda x: (-len(type_buckets[x]), x)):
        group = type_buckets[type_code]
        pattern_buckets: dict[str, list[RowItem]] = defaultdict(list)
        for item in group:
            pattern_buckets[item.pattern_signature].append(item)
        selected_count = 0
        pattern_json: list[dict[str, Any]] = []
        for signature in sorted(pattern_buckets, key=lambda x: (-len(pattern_buckets[x]), x)):
            pattern_group = pattern_buckets[signature]
            selected = select_group_rows(pattern_group, per_pattern)
            selected_count += len(selected)
            examples = [text(x.values["材料描述"]) for x in pattern_group[:3]]
            for item, reason in selected:
                desc = text(item.values["材料描述"])
                selected_rows.append({
                    "模型类别": category,
                    "类型编码": item.type_code,
                    "模板桶内总数": len(pattern_group),
                    "抽样原因": reason,
                    **{name: text(item.values.get(name)) for name in FIELD_NAMES},
                    "Excel行号": item.row_no,
                    "原文模板签名": item.pattern_signature,
                })
                if desc not in seen_inputs:
                    seen_inputs.add(desc)
                    type_obj = build_type(item)
                    dataset.append({"input": desc, "output": {"TYPE": type_obj}})
                    reason_text = conflict_reason(item, type_obj)
                    if reason_text:
                        conflict_rows.append({
                            "模型类别": category,
                            "Excel行号": item.row_no,
                            "类型编码": item.type_code,
                            "材料描述": desc,
                            "TYPE": json.dumps(type_obj, ensure_ascii=False),
                            "冲突原因": reason_text,
                        })
            pattern_rows.append({
                "模型类别": category,
                "类型编码": type_code,
                "模板总数": len(pattern_buckets),
                "原文模板签名": signature,
                "模板条数": len(pattern_group),
                "抽样条数": len(selected),
                "典型描述1": examples[0] if len(examples) > 0 else "",
                "典型描述2": examples[1] if len(examples) > 1 else "",
                "典型描述3": examples[2] if len(examples) > 2 else "",
            })
            pattern_json.append({"pattern_signature": signature, "total_rows": len(pattern_group), "selected_rows": len(selected), "examples": examples})
        examples = [text(x.values["材料描述"]) for x in group[:3]]
        type_rows.append({
            "模型类别": category,
            "类型编码": type_code,
            "总条数": len(group),
            "模板总数": len(pattern_buckets),
            "抽样条数": selected_count,
            "典型描述1": examples[0] if len(examples) > 0 else "",
            "典型描述2": examples[1] if len(examples) > 1 else "",
            "典型描述3": examples[2] if len(examples) > 2 else "",
        })
        summary["types"][type_code] = {
            "total_rows": len(group),
            "pattern_count": len(pattern_buckets),
            "selected_rows": selected_count,
            "examples": examples,
            "patterns": pattern_json,
        }
    return selected_rows, pattern_rows, type_rows, dataset, conflict_rows, summary


def main() -> None:
    parser = argparse.ArgumentParser(description="从项目 Excel 抽取弯头和三通 TYPE 训练草稿")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--per-pattern", type=int, default=5)
    args = parser.parse_args()

    items, skipped = read_rows(args.input)
    args.output_dir.mkdir(parents=True, exist_ok=True)

    for category in ["弯头", "三通"]:
        selected_rows, pattern_rows, type_rows, dataset, conflict_rows, summary = build_outputs(items, category, args.per_pattern)
        write_sampling_workbook(args.output_dir / f"{category}覆盖抽样.xlsx", selected_rows, pattern_rows, type_rows, conflict_rows)
        dump_json(args.output_dir / f"{category}TYPE草稿.json", dataset)
        dump_json(args.output_dir / f"{category}覆盖抽样汇总.json", summary)
        print(f"{category}: candidates={summary['input_rows']}, selected_rows={len(selected_rows)}, dataset={len(dataset)}, conflicts={len(conflict_rows)}")
        for row in type_rows[:30]:
            print(f"  {row['类型编码']}: total={row['总条数']}, patterns={row['模板总数']}, selected={row['抽样条数']}")

    dump_json(args.output_dir / "非弯头三通管件编码汇总.json", skipped.most_common())
    print(f"输出目录: {args.output_dir}")


if __name__ == "__main__":
    main()
