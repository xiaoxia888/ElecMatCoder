#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Cover pipe draft MATERIAL with stage1 output and upgrade to structured MATERIAL schema."""

from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[4]
BASE = ROOT / "apps/trainer/qwen3_fte/output/pipe_project_sampling_full"
DRAFT_PATH = BASE / "直管训练草稿.json"
STAGE_PATH = BASE / "stage1_dataset_2026-04-23.json"
OUT_XLSX = BASE / "直管MATERIAL覆盖与结构升级.xlsx"


def text(v: Any) -> str:
    return "" if v is None else str(v).strip()


def as_list(v: Any) -> list[Any]:
    if v in (None, ""):
        return []
    return v if isinstance(v, list) else [v]


def material_items(output: dict[str, Any]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for item in as_list(output.get("MATERIAL")):
        if isinstance(item, dict):
            out.append({
                "ROLE": text(item.get("ROLE")) or "MAIN",
                "VALUE": text(item.get("VALUE")),
                "SPECIAL_REQ": [text(x) for x in as_list(item.get("SPECIAL_REQ")) if text(x)],
            })
        elif text(item):
            out.append({"ROLE": "MAIN", "VALUE": text(item), "SPECIAL_REQ": []})
    return out


def base_item(
    *,
    role: str = "MAIN",
    value: str = "",
    lining: list[str] | None = None,
    inner: list[str] | None = None,
    outer: list[str] | None = None,
    special: list[str] | None = None,
) -> dict[str, Any]:
    return {
        "ROLE": role or "MAIN",
        "VALUE": value,
        "LINING": lining or [],
        "COATING": {"INNER": inner or [], "OUTER": outer or []},
        "SPECIAL_REQ": special or [],
    }


def skeletonize(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        base_item(
            role=text(item.get("ROLE")) or "MAIN",
            value=text(item.get("VALUE")),
            special=[text(x) for x in as_list(item.get("SPECIAL_REQ")) if text(x)],
        )
        for item in items
    ]


def clean_grade_pair(value: str, input_text: str) -> str:
    s = value or input_text
    u = s.upper().replace(" ", "")
    if (
        "304/304L" in u
        or "TP304/304L" in u
        or "TP304/TP304L" in u
        or "GR.304/304L" in u
        or "GR304/304L" in u
    ):
        return "304/304L"
    if (
        "316/316L" in u
        or "TP316/316L" in u
        or "TP316/TP316L" in u
        or "GR.316/316L" in u
        or "GR316/316L" in u
        or "F316/F316L" in u
    ):
        return "316/316L"
    return value


def upgrade_material(input_text: str, stage_items: list[dict[str, Any]]) -> tuple[str, list[dict[str, Any]], str]:
    if not stage_items:
        return "空材质", [], "模型 MATERIAL 为空"

    value = text(stage_items[0].get("VALUE"))
    special = [text(x) for x in as_list(stage_items[0].get("SPECIAL_REQ")) if text(x)]
    value_compact = re.sub(r"[\s\-_/]+", "", value.upper())
    combined = input_text + " " + json.dumps(stage_items, ensure_ascii=False)

    # 法兰管多角色：管体和法兰材质分开。
    if re.search(r"FLANGED\s*PIPE", input_text, re.I) and re.search(r"A105", input_text, re.I):
        main_value = "20" if re.search(r"\b20\b", input_text) else value
        main_lining = ["GLASS LINED"] if re.search(r"GLASS\s*LINED", input_text, re.I) else []
        flange_outer = ["PLASTIC DIPPING"] if re.search(r"PLASTIC\s*DIPPING", input_text, re.I) else []
        return "多角色材质", [
            base_item(role="MAIN", value=main_value, lining=main_lining),
            base_item(role="FLANGE", value="A105", outer=flange_outer),
        ], "FLANGED PIPE 中管体材质和法兰材质分开"

    # 双牌号/双认证：不拆为两个 item，只补新骨架。
    if re.search(r"(304/304L|316/316L|TP304/304L|TP304/TP304L|TP316/316L|TP316/TP316L|GR\.?304/304L|GR\.?316/316L|F316/F316L)", combined, re.I):
        return "双牌号不拆", [base_item(value=clean_grade_pair(value, input_text), special=special)], "双牌号保留在同一个 VALUE"

    # 内外涂层。
    m = re.search(r"\b(Q235B|20)\s*外加强级?PE\s*内EP", input_text, re.I)
    if m:
        return "内外涂层", [base_item(value=m.group(1), inner=["EP"], outer=["加强级PE"], special=special)], "外加强级PE/内EP 拆入 COATING"

    # 内衬类。
    m = re.fullmatch(r"\s*(20|304|CS|Q235B)\s*[/+]\s*(PTFE|PE)\s*", value, re.I)
    if m:
        return "内衬", [base_item(value=m.group(1), lining=[m.group(2).upper()], special=special)], "主材/内衬材质拆分"

    m = re.fullmatch(r"\s*(PTFE)\s*/\s*(CS)\s*", value, re.I)
    if m:
        return "内衬", [base_item(value="CS", lining=["PTFE"], special=special)], "PTFE/CS 按 CS 主材 + PTFE 内衬"

    m = re.fullmatch(r"(20|304)PTFE", value_compact, re.I)
    if m:
        return "内衬", [base_item(value=m.group(1), lining=["PTFE"], special=special)], "压缩 PTFE 写法拆分"

    if re.search(r"20\s+GLASS\s*LINED", value, re.I) or re.search(r"20\s+GLASS\s*LINED", input_text, re.I):
        return "内衬", [base_item(value="20", lining=["GLASS LINED"], special=special)], "20 GLASS LINED 拆分"

    # 用户已确认：搪玻璃主体材质就是搪玻璃。
    if re.search(r"搪玻璃", value) or re.search(r"搪玻璃", input_text):
        return "搪玻璃主体", [base_item(value="搪玻璃", special=special)], "搪玻璃作为材质主体"

    # 外涂层/外防腐。
    m = re.fullmatch(r"(20|Q235B|L245)3PE", value_compact, re.I)
    if m:
        return "外涂层", [base_item(value=m.group(1), outer=["3PE"], special=special)], "压缩 3PE 写法拆分"

    if value_compact == "203PE":
        return "外涂层", [base_item(value="20", outer=["3PE"], special=special)], "203PE 拆为 20 + 3PE"

    if re.search(r"钢管自带3PE加强级外防腐|3PE加强级外防腐", input_text, re.I) and value:
        return "外涂层", [base_item(value=value, outer=["3PE"], special=special)], "原文有 3PE 外防腐"

    # 特殊要求保留 SPECIAL_REQ，只补完整骨架。
    if special:
        return "特殊要求", skeletonize(stage_items), "保留模型 SPECIAL_REQ"

    return "普通材质", skeletonize(stage_items), "仅补全新 MATERIAL 骨架"


def build_stage_index(stage: list[dict[str, Any]]) -> dict[tuple[str, int], dict[str, Any]]:
    seen = defaultdict(int)
    index: dict[tuple[str, int], dict[str, Any]] = {}
    for item in stage:
        inp = text(item.get("input"))
        occ = seen[inp]
        seen[inp] += 1
        index[(inp, occ)] = item
    return index


def append_sheet(wb: Workbook, name: str, rows: list[list[Any]], headers: list[str]):
    ws = wb.create_sheet(name)
    ws.append(headers)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for row in rows:
        ws.append(row)
    widths = [10, 18, 90, 55, 80, 50]
    for idx, width in enumerate(widths[: len(headers)], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"


def main():
    draft = json.loads(DRAFT_PATH.read_text(encoding="utf-8"))
    stage = json.loads(STAGE_PATH.read_text(encoding="utf-8"))
    stage_index = build_stage_index(stage)

    seen = defaultdict(int)
    rows_by_category: dict[str, list[list[Any]]] = defaultdict(list)
    unmatched: list[list[Any]] = []
    category_counter: Counter[str] = Counter()
    changed_count = 0

    for idx, item in enumerate(draft):
        inp = text(item.get("input"))
        occ = seen[inp]
        seen[inp] += 1
        stage_item = stage_index.get((inp, occ))
        old_material = (item.get("output") or {}).get("MATERIAL", [])
        if stage_item is None:
            unmatched.append([idx, inp, json.dumps(old_material, ensure_ascii=False), "", "", "未匹配到模型结果"])
            continue

        stage_material = material_items(stage_item.get("output") or {})
        category, new_material, reason = upgrade_material(inp, stage_material)
        item.setdefault("output", {})["MATERIAL"] = new_material
        changed_count += int(json.dumps(old_material, ensure_ascii=False, sort_keys=True) != json.dumps(new_material, ensure_ascii=False, sort_keys=True))
        category_counter[category] += 1
        rows_by_category[category].append([
            idx,
            category,
            inp,
            json.dumps(old_material, ensure_ascii=False),
            json.dumps(new_material, ensure_ascii=False),
            reason,
        ])

    DRAFT_PATH.write_text(json.dumps(draft, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    wb = Workbook()
    wb.remove(wb.active)
    headers = ["索引", "类别", "材料描述", "覆盖前草稿MATERIAL", "覆盖后新MATERIAL", "原因"]
    order = ["普通材质", "双牌号不拆", "内衬", "外涂层", "内外涂层", "多角色材质", "搪玻璃主体", "特殊要求", "空材质"]
    append_sheet(wb, "汇总", [[k, category_counter.get(k, 0)] for k in order] + [["未匹配", len(unmatched)], ["实际变更", changed_count]], ["类别", "数量"])
    for name in order:
        append_sheet(wb, name, rows_by_category.get(name, []), headers)
    append_sheet(wb, "未匹配", unmatched, headers)
    wb.save(OUT_XLSX)

    print("draft", len(draft), "stage", len(stage))
    print("changed", changed_count)
    print("unmatched", len(unmatched))
    for name in order:
        print(name, category_counter.get(name, 0))
    print(OUT_XLSX)


if __name__ == "__main__":
    main()
