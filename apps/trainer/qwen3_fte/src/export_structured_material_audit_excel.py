#!/usr/bin/env python3
"""Export current v3 train/val material labels as an audit workbook."""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[4]
DATA_DIR = (
    ROOT
    / "apps/trainer/qwen3_fte/output/按8类拆分数据集"
    / "材质规范/结构化原始牌号/重新划分_v3"
)
DEFAULT_TRAIN = DATA_DIR / "材质规范_结构化原始牌号_train.json"
DEFAULT_VAL = DATA_DIR / "材质规范_结构化原始牌号_val.json"
DEFAULT_OUTPUT = (
    ROOT
    / "outputs/material_audit_20260730"
    / "材质规范_train_val标注统计.xlsx"
)

HEADER_FILL = PatternFill("solid", fgColor="FF0F766E")
SECTION_FILL = PatternFill("solid", fgColor="FFCCFBF1")
SUBTLE_FILL = PatternFill("solid", fgColor="FFF1F5F9")
HIGH_FILL = PatternFill("solid", fgColor="FFFEE2E2")
MEDIUM_FILL = PatternFill("solid", fgColor="FFFEF3C7")
LOW_FILL = PatternFill("solid", fgColor="FFE0F2FE")
HEADER_FONT = Font(name="Arial", color="FFFFFFFF", bold=True, size=10)
TITLE_FONT = Font(name="Arial", color="FF0F172A", bold=True, size=18)
BODY_FONT = Font(name="Arial", color="FF1E293B", size=10)
THIN_GRAY = Side(style="thin", color="FFCBD5E1")
BOTTOM_BORDER = Border(bottom=THIN_GRAY)
VALID_PARTS = {"BODY", "LINING", "INNER_PIPE", "OUTER_PIPE", "FLANGE"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--train", type=Path, default=DEFAULT_TRAIN)
    parser.add_argument("--val", type=Path, default=DEFAULT_VAL)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


def load_rows(path: Path, split: str) -> list[dict[str, Any]]:
    with path.open(encoding="utf-8") as handle:
        rows = json.load(handle)
    if not isinstance(rows, list):
        raise ValueError(f"{path} 顶层必须是数组")
    for index, row in enumerate(rows):
        row["_split"] = split
        row["_index"] = index
    return rows


def product_standard_values(output: dict[str, Any]) -> list[str]:
    values: list[str] = []
    for item in output.get("STANDARD", []):
        if isinstance(item, dict):
            values.extend(str(value) for value in item.values() if value)
    return list(dict.fromkeys(values))


def item_signature(item: dict[str, Any]) -> tuple[Any, ...]:
    return (
        item.get("PART", ""),
        item.get("VALUE", ""),
        tuple(item.get("SPECIAL_REQ", [])),
    )


def audit_record(row: dict[str, Any]) -> list[tuple[str, str, str]]:
    output = row.get("output", {})
    material = output.get("MATERIAL", [])
    standards = output.get("STANDARD", [])
    issues: list[tuple[str, str, str]] = []

    if set(output) != {"MATERIAL", "STANDARD"}:
        issues.append(
            (
                "高",
                "输出字段异常",
                f"实际字段={','.join(sorted(output)) or '<空>'}",
            )
        )
    if not isinstance(material, list) or not material:
        return issues + [("高", "无材质项", "MATERIAL为空或不是数组")]
    if not isinstance(standards, list):
        issues.append(("高", "STANDARD结构异常", "STANDARD不是数组"))
        standards = []

    signatures = [
        item_signature(item) for item in material if isinstance(item, dict)
    ]
    if len(signatures) != len(set(signatures)):
        issues.append(("高", "重复材质项", "同一条记录存在完全重复的材质项"))

    part_counts = Counter(
        str(item.get("PART", ""))
        for item in material
        if isinstance(item, dict)
    )
    repeated_parts = [part for part, count in part_counts.items() if part and count > 1]
    if repeated_parts:
        issues.append(("中", "同一PART多材质项", "、".join(repeated_parts)))

    for item_index, item in enumerate(material, 1):
        prefix = f"第{item_index}项"
        if not isinstance(item, dict):
            issues.append(("高", "材质项结构异常", f"{prefix}: 不是对象"))
            continue
        if set(item) != {"PART", "VALUE", "SPECIAL_REQ"}:
            issues.append(
                (
                    "高",
                    "材质项字段异常",
                    f"{prefix}: {','.join(sorted(item)) or '<空>'}",
                )
            )
        part = str(item.get("PART", ""))
        value = str(item.get("VALUE", ""))
        special = item.get("SPECIAL_REQ", [])
        if part not in VALID_PARTS:
            issues.append(("高", "PART异常", f"{prefix}: {part or '<空>'}"))
        if not value.strip():
            issues.append(("高", "VALUE为空", prefix))
        if len(value.strip()) == 1 and re.fullmatch(r"[A-Za-z]", value.strip()):
            issues.append(("高", "孤立单字符VALUE", f"{prefix}: VALUE={value}"))
        if not isinstance(special, list):
            issues.append(("高", "SPECIAL_REQ结构异常", f"{prefix}: 不是数组"))
        elif len(special) != len(set(map(str, special))):
            issues.append(("中", "SPECIAL_REQ重复", prefix))

    for standard_index, item in enumerate(standards, 1):
        if (
            not isinstance(item, dict)
            or set(item) != {"BODY"}
            or not str(item.get("BODY", "")).strip()
        ):
            issues.append(
                (
                    "高",
                    "STANDARD项异常",
                    f"第{standard_index}项: {item!r}",
                )
            )

    standard_values = [
        str(item.get("BODY", ""))
        for item in standards
        if isinstance(item, dict) and item.get("BODY")
    ]
    if len(standard_values) != len(set(standard_values)):
        issues.append(("中", "STANDARD重复", "顶层STANDARD存在重复值"))

    return issues


def join_items(material: list[dict[str, Any]], key: str) -> str:
    values: list[str] = []
    for item in material:
        if not isinstance(item, dict):
            values.append(str(item))
            continue
        value = item.get(key, "")
        if isinstance(value, list):
            value = ",".join(str(part) for part in value)
        values.append(str(value or ""))
    return " | ".join(values)


def style_header(sheet: Any, row: int, start_col: int, end_col: int) -> None:
    for cells in sheet.iter_cols(
        min_col=start_col,
        max_col=end_col,
        min_row=row,
        max_row=row,
    ):
        cell = cells[0]
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[row].height = 26


def style_data_sheet(
    sheet: Any,
    widths: Iterable[float],
    freeze: str = "A2",
    wrap_columns: Iterable[int] = (),
) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = freeze
    sheet.auto_filter.ref = sheet.dimensions
    style_header(sheet, 1, 1, sheet.max_column)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = BOTTOM_BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=cell.column in wrap_columns,
            )


def add_title(sheet: Any, title: str, subtitle: str) -> None:
    sheet.merge_cells("A1:H1")
    sheet["A1"] = title
    sheet["A1"].font = TITLE_FONT
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 34
    sheet.merge_cells("A2:H2")
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(name="Arial", color="FF475569", size=10)
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 32


def append_percentage_rows(
    sheet: Any,
    rows: Iterable[list[Any]],
    count_column: int,
    percentage_column: int,
    total: int,
) -> None:
    for values in rows:
        sheet.append(values)
    for row_index in range(2, sheet.max_row + 1):
        count = sheet.cell(row_index, count_column).value or 0
        sheet.cell(row_index, percentage_column, count / total if total else 0)
        sheet.cell(row_index, percentage_column).number_format = "0.00%"


def create_workbook(
    rows: list[dict[str, Any]],
    train_path: Path,
    val_path: Path,
) -> Workbook:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "概览"
    record_sheet = workbook.create_sheet("逐条标注")
    item_sheet = workbook.create_sheet("材质项明细")
    value_sheet = workbook.create_sheet("VALUE统计")
    combo_sheet = workbook.create_sheet("标签组合统计")
    standard_sheet = workbook.create_sheet("产品规范统计")
    part_sheet = workbook.create_sheet("PART统计")
    special_sheet = workbook.create_sheet("SPECIAL_REQ统计")
    issue_sheet = workbook.create_sheet("结构异常")

    value_counts: Counter[str] = Counter()
    value_split: dict[str, Counter[str]] = defaultdict(Counter)
    value_parts: dict[str, set[str]] = defaultdict(set)
    value_standards: dict[str, set[str]] = defaultdict(set)
    standard_counts: Counter[str] = Counter()
    standard_split: dict[str, Counter[str]] = defaultdict(Counter)
    standard_values: dict[str, set[str]] = defaultdict(set)
    combo_counts: Counter[tuple[str, str, str]] = Counter()
    combo_split: dict[tuple[str, str, str], Counter[str]] = defaultdict(Counter)
    part_counts: Counter[str] = Counter()
    part_split: dict[str, Counter[str]] = defaultdict(Counter)
    special_counts: Counter[str] = Counter()
    special_split: dict[str, Counter[str]] = defaultdict(Counter)
    issue_counts: Counter[str] = Counter()
    record_rows: list[list[Any]] = []
    item_rows: list[list[Any]] = []
    issue_rows: list[list[Any]] = []

    for row in rows:
        split = row["_split"]
        source_index = row["_index"]
        text = str(row.get("input", ""))
        output = row.get("output", {})
        material = output.get("MATERIAL", [])
        standards = product_standard_values(output)
        standard_text = " | ".join(standards)
        issues = audit_record(row)
        issue_types = " | ".join(dict.fromkeys(issue[1] for issue in issues))

        record_rows.append(
            [
                split,
                source_index,
                text,
                len(material),
                join_items(material, "PART"),
                join_items(material, "VALUE"),
                join_items(material, "SPECIAL_REQ"),
                standard_text,
                issue_types,
            ]
        )

        record_values: set[str] = set()
        for item_index, item in enumerate(material, 1):
            if not isinstance(item, dict):
                continue
            part = str(item.get("PART", ""))
            value = str(item.get("VALUE", ""))
            special_values = [str(value) for value in item.get("SPECIAL_REQ", [])]
            special_text = ",".join(special_values)
            item_rows.append(
                [
                    split,
                    source_index,
                    text,
                    item_index,
                    part,
                    value,
                    special_text,
                    standard_text,
                ]
            )
            value_counts[value] += 1
            value_split[value][split] += 1
            value_parts[value].add(part)
            value_standards[value].update(standards)
            part_counts[part] += 1
            part_split[part][split] += 1
            combo = (part, value, special_text)
            combo_counts[combo] += 1
            combo_split[combo][split] += 1
            record_values.add(value)
            for special in special_values:
                special_counts[special] += 1
                special_split[special][split] += 1

        for standard in standards:
            standard_counts[standard] += 1
            standard_split[standard][split] += 1
            standard_values[standard].update(record_values)

        for priority, issue_type, detail in issues:
            issue_counts[issue_type] += 1
            issue_rows.append(
                [
                    priority,
                    issue_type,
                    split,
                    source_index,
                    text,
                    join_items(material, "PART"),
                    join_items(material, "VALUE"),
                    join_items(material, "SPECIAL_REQ"),
                    standard_text,
                    detail,
                ]
            )

    record_sheet.append(
        [
            "数据集",
            "原序号",
            "原始描述",
            "材质项数",
            "PART",
            "VALUE",
            "SPECIAL_REQ",
            "产品规范STANDARD",
            "审计提示",
        ]
    )
    for values in record_rows:
        record_sheet.append(values)
    style_data_sheet(
        record_sheet,
        [10, 11, 75, 10, 24, 42, 28, 42, 28],
        freeze="C2",
        wrap_columns={3, 5, 6, 7, 8, 9},
    )

    item_sheet.append(
        [
            "数据集",
            "原序号",
            "原始描述",
            "材质项序号",
            "PART",
            "VALUE",
            "SPECIAL_REQ",
            "产品规范STANDARD",
        ]
    )
    for values in item_rows:
        item_sheet.append(values)
    style_data_sheet(
        item_sheet,
        [10, 11, 75, 12, 16, 42, 28, 42],
        freeze="C2",
        wrap_columns={3, 6, 7, 8},
    )

    value_sheet.append(
        ["VALUE", "材质项数", "占比", "train", "val", "关联PART", "关联产品规范"]
    )
    append_percentage_rows(
        value_sheet,
        (
            [
                value or "(空)",
                count,
                None,
                value_split[value]["train"],
                value_split[value]["val"],
                " | ".join(sorted(value_parts[value])),
                " | ".join(sorted(value_standards[value])),
            ]
            for value, count in value_counts.most_common()
        ),
        2,
        3,
        sum(value_counts.values()),
    )
    style_data_sheet(
        value_sheet,
        [42, 14, 12, 12, 12, 28, 70],
        wrap_columns={1, 6, 7},
    )

    combo_sheet.append(
        [
            "PART",
            "VALUE",
            "SPECIAL_REQ",
            "出现次数",
            "占比",
            "train",
            "val",
        ]
    )
    append_percentage_rows(
        combo_sheet,
        (
            [
                combo[0],
                combo[1],
                combo[2],
                count,
                None,
                combo_split[combo]["train"],
                combo_split[combo]["val"],
            ]
            for combo, count in combo_counts.most_common()
        ),
        4,
        5,
        sum(combo_counts.values()),
    )
    style_data_sheet(
        combo_sheet,
        [18, 45, 30, 14, 12, 12, 12],
        wrap_columns={2, 3},
    )

    standard_sheet.append(
        ["产品规范STANDARD", "记录数", "占比", "train", "val", "VALUE种数", "VALUE示例"]
    )
    append_percentage_rows(
        standard_sheet,
        (
            [
                standard,
                count,
                None,
                standard_split[standard]["train"],
                standard_split[standard]["val"],
                len(standard_values[standard]),
                " | ".join(sorted(standard_values[standard])[:40]),
            ]
            for standard, count in standard_counts.most_common()
        ),
        2,
        3,
        sum(standard_counts.values()),
    )
    style_data_sheet(
        standard_sheet,
        [32, 14, 12, 12, 12, 14, 90],
        wrap_columns={1, 7},
    )

    part_sheet.append(["PART", "材质项数", "占比", "train", "val"])
    append_percentage_rows(
        part_sheet,
        (
            [
                part or "(空)",
                count,
                None,
                part_split[part]["train"],
                part_split[part]["val"],
            ]
            for part, count in part_counts.most_common()
        ),
        2,
        3,
        sum(part_counts.values()),
    )
    style_data_sheet(part_sheet, [24, 16, 12, 12, 12])

    special_sheet.append(["SPECIAL_REQ", "材质项数", "占比", "train", "val"])
    append_percentage_rows(
        special_sheet,
        (
            [
                special,
                count,
                None,
                special_split[special]["train"],
                special_split[special]["val"],
            ]
            for special, count in special_counts.most_common()
        ),
        2,
        3,
        sum(special_counts.values()),
    )
    style_data_sheet(special_sheet, [32, 16, 12, 12, 12])

    issue_sheet.append(
        [
            "优先级",
            "问题类型",
            "数据集",
            "原序号",
            "原始描述",
            "PART",
            "VALUE",
            "SPECIAL_REQ",
            "产品规范STANDARD",
            "问题说明",
        ]
    )
    for values in issue_rows:
        issue_sheet.append(values)
    style_data_sheet(
        issue_sheet,
        [10, 24, 10, 11, 75, 24, 42, 28, 42, 42],
        freeze="E2",
        wrap_columns={5, 6, 7, 8, 9, 10},
    )
    if issue_sheet.max_row >= 2:
        issue_range = f"A2:A{issue_sheet.max_row}"
        issue_sheet.conditional_formatting.add(
            issue_range,
            FormulaRule(formula=["$A2=\"高\""], fill=HIGH_FILL),
        )
        issue_sheet.conditional_formatting.add(
            issue_range,
            FormulaRule(formula=["$A2=\"中\""], fill=MEDIUM_FILL),
        )
        issue_sheet.conditional_formatting.add(
            issue_range,
            FormulaRule(formula=["$A2=\"低\""], fill=LOW_FILL),
        )

    overview.sheet_view.showGridLines = False
    add_title(
        overview,
        "材质规范 v3 训练集标注统计",
        "统计范围为当前重新划分_v3的 train + val；用于核查 VALUE、PART、SPECIAL_REQ 与顶层产品规范。",
    )
    overview["A4"] = "核心统计"
    overview["A4"].fill = SECTION_FILL
    overview["A4"].font = Font(name="Arial", bold=True, color="FF115E59", size=12)
    overview.merge_cells("A4:B4")
    split_counts = Counter(row["_split"] for row in rows)
    overview_rows = [
        ("训练记录", split_counts["train"]),
        ("验证记录", split_counts["val"]),
        ("总记录数", len(rows)),
        ("材质项总数", len(item_rows)),
        ("唯一VALUE数", len(value_counts)),
        ("唯一产品规范数", len(standard_counts)),
        ("唯一标签组合数", len(combo_counts)),
        ("结构异常记录数", len({(row[2], row[3]) for row in issue_rows})),
        ("结构异常项数", len(issue_rows)),
    ]
    for offset, (label, value) in enumerate(overview_rows, 5):
        overview.cell(offset, 1, label)
        overview.cell(offset, 2, value)
        overview.cell(offset, 1).font = Font(
            name="Arial",
            bold=True,
            color="FF475569",
        )
        overview.cell(offset, 2).font = Font(
            name="Arial",
            bold=True,
            color="FF0F766E",
            size=13,
        )
        overview.cell(offset, 1).fill = SUBTLE_FILL
        overview.cell(offset, 2).fill = SUBTLE_FILL
        overview.cell(offset, 2).number_format = "#,##0"

    overview["D4"] = "异常类型汇总"
    overview["D4"].fill = SECTION_FILL
    overview["D4"].font = Font(name="Arial", bold=True, color="FF115E59", size=12)
    overview.merge_cells("D4:H4")
    if issue_counts:
        for offset, (issue_type, count) in enumerate(issue_counts.most_common(12), 5):
            overview.cell(offset, 4, f"{issue_type}：{count:,}")
            overview.merge_cells(
                start_row=offset,
                start_column=4,
                end_row=offset,
                end_column=8,
            )
            overview.cell(offset, 4).font = BODY_FONT
    else:
        overview["D5"] = "未发现结构异常"
        overview.merge_cells("D5:H5")
        overview["D5"].font = BODY_FONT

    source_row = max(18, 6 + len(issue_counts.most_common(12)))
    overview.cell(source_row, 1, "数据来源")
    overview.cell(source_row, 1).fill = SECTION_FILL
    overview.cell(source_row, 1).font = Font(
        name="Arial",
        bold=True,
        color="FF115E59",
    )
    overview.merge_cells(
        start_row=source_row,
        start_column=1,
        end_row=source_row,
        end_column=8,
    )
    overview.cell(source_row + 1, 1, f"train: {train_path.relative_to(ROOT)}")
    overview.cell(source_row + 2, 1, f"val: {val_path.relative_to(ROOT)}")
    overview.merge_cells(
        start_row=source_row + 1,
        start_column=1,
        end_row=source_row + 1,
        end_column=8,
    )
    overview.merge_cells(
        start_row=source_row + 2,
        start_column=1,
        end_row=source_row + 2,
        end_column=8,
    )
    overview.cell(source_row + 1, 1).font = Font(
        name="Arial",
        color="FF475569",
        size=9,
    )
    overview.cell(source_row + 2, 1).font = Font(
        name="Arial",
        color="FF475569",
        size=9,
    )

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "高频 VALUE（前10）"
    chart.y_axis.title = "VALUE"
    chart.x_axis.title = "材质项数"
    chart.height = 7.5
    chart.width = 14
    chart.add_data(
        Reference(
            value_sheet,
            min_col=2,
            min_row=1,
            max_row=min(11, value_sheet.max_row),
        ),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(
            value_sheet,
            min_col=1,
            min_row=2,
            max_row=min(11, value_sheet.max_row),
        )
    )
    overview.add_chart(chart, "J4")

    for column, width in {
        "A": 22,
        "B": 18,
        "C": 3,
        "D": 28,
        "E": 14,
        "F": 14,
        "G": 14,
        "H": 14,
        "I": 3,
        "J": 16,
        "K": 16,
        "L": 16,
        "M": 16,
        "N": 16,
        "O": 16,
        "P": 16,
    }.items():
        overview.column_dimensions[column].width = width

    return workbook


def main() -> int:
    args = parse_args()
    rows = load_rows(args.train, "train") + load_rows(args.val, "val")
    workbook = create_workbook(rows, args.train, args.val)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(
        json.dumps(
            {
                "output": str(args.output),
                "train_rows": sum(row["_split"] == "train" for row in rows),
                "val_rows": sum(row["_split"] == "val" for row in rows),
                "sheets": workbook.sheetnames,
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
