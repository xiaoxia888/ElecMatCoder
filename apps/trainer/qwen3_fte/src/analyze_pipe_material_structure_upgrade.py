#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Analyze pipe MATERIAL rows that should be upgraded to structured material schema."""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[4]
BASE = ROOT / "apps/trainer/qwen3_fte/output/pipe_project_sampling_full"
STAGE_PATH = BASE / "stage1_dataset_2026-04-23.json"
OUT_XLSX = BASE / "直管MATERIAL结构升级候选.xlsx"


def text(v: Any) -> str:
    return "" if v is None else str(v).strip()


def as_list(v: Any) -> list[Any]:
    if v in (None, ""):
        return []
    return v if isinstance(v, list) else [v]


def mat_items(output: dict[str, Any]) -> list[dict[str, Any]]:
    out = []
    for m in as_list(output.get("MATERIAL")):
        if isinstance(m, dict):
            out.append(m)
        elif text(m):
            out.append({"ROLE": "MAIN", "VALUE": text(m), "SPECIAL_REQ": []})
    return out


def base_item(role="MAIN", value="", lining=None, inner=None, outer=None, special=None):
    return {
        "ROLE": role,
        "VALUE": value,
        "LINING": lining or [],
        "COATING": {"INNER": inner or [], "OUTER": outer or []},
        "SPECIAL_REQ": special or [],
    }


def normalize_special(s: str) -> str:
    u = s.upper().strip()
    if u in {"GALV", "GALVANIZED", "ZN", "ZINC", "镀锌"}:
        return "Zn"
    if u in {"NACE", "ANTI-H2S", "H2S", "抗硫"}:
        return s
    return s


def propose(input_text: str, mats: list[dict[str, Any]]) -> tuple[str, list[dict[str, Any]] | None, str]:
    combined = input_text + " " + json.dumps(mats, ensure_ascii=False)
    val = text(mats[0].get("VALUE")) if mats else ""
    special = [normalize_special(text(x)) for x in as_list(mats[0].get("SPECIAL_REQ")) if text(x)] if mats else []
    val_upper = val.upper().replace(" ", "")

    # 双牌号/双认证：不拆，只补完整新骨架。
    if re.search(r"(304/304L|316/316L|TP304/304L|TP316/316L|GR\.304/304L|GR\.316/316L)", combined, re.I):
        clean = val
        clean = re.sub(r"ASTM\s*A\d+\s*(GR\.?|GRADE\s*)?", "", clean, flags=re.I).strip()
        clean = re.sub(r"\bTP", "", clean, flags=re.I).strip()
        clean = clean.replace(" ", "")
        if "304/304L" in clean.upper(): clean = "304/304L"
        elif "316/316L" in clean.upper(): clean = "316/316L"
        return "双牌号不拆", [base_item(value=clean or val, special=special)], "双牌号/双认证保留在 VALUE"

    # 法兰管多角色：管体 + 法兰。
    if re.search(r"FLANGED\s*PIPE", input_text, re.I) and re.search(r"A105", input_text, re.I):
        main_value = "20" if re.search(r"\b20\b", input_text) else val
        main_lining = ["GLASS LINED"] if re.search(r"GLASS\s*LINED", input_text, re.I) else []
        outer = ["PLASTIC DIPPING"] if re.search(r"PLASTIC\s*DIPPING", input_text, re.I) else []
        return "多角色材质", [
            base_item(role="MAIN", value=main_value, lining=main_lining),
            base_item(role="FLANGE", value="A105", outer=outer),
        ], "法兰管包含管体材质和法兰材质"

    # 内外涂层。
    m = re.search(r"\b(Q235B|20)\s*外加强级?PE\s*内EP", input_text, re.I)
    if m:
        return "内外涂层", [base_item(value=m.group(1), inner=["EP"], outer=["加强级PE"])], "外加强级PE/内EP 拆入 COATING"

    # PTFE / PE lining.
    m = re.search(r"^(20|304|CS|Q235B)\s*/\s*(PTFE|PE)$", val, re.I)
    if m:
        return "内衬", [base_item(value=m.group(1), lining=[m.group(2).upper()], special=special)], "主材/内衬材质"
    m = re.search(r"^(PTFE)\s*/\s*(CS)$", val, re.I)
    if m:
        return "内衬", [base_item(value="CS", lining=["PTFE"], special=special)], "PTFE/CS 按 CS 主材 + PTFE 内衬"
    m = re.search(r"^(20|304)(PTFE)$", val_upper, re.I)
    if m:
        return "内衬", [base_item(value=m.group(1), lining=["PTFE"], special=special)], "压缩写法拆分"
    if re.search(r"20\s+GLASS\s*LINED", val, re.I) or re.search(r"20\s+GLASS\s*LINED", input_text, re.I):
        return "内衬", [base_item(value="20", lining=["GLASS LINED"], special=special)], "20 GLASS LINED 拆分"
    if re.search(r"搪玻璃", val) or re.search(r"搪玻璃", input_text):
        return "需人工确认", [base_item(value=val if val and val != "搪玻璃" else "", lining=["搪玻璃"], special=special)], "只有搪玻璃时主材不明确"

    # 3PE/outside coating compressed values.
    m = re.search(r"^(20|Q235B|L245)3PE$", val_upper, re.I)
    if m:
        return "外涂层", [base_item(value=m.group(1), outer=["3PE"], special=special)], "压缩 3PE 拆分"
    if val_upper == "203PE":
        return "外涂层", [base_item(value="20", outer=["3PE"], special=special)], "203PE 拆为 20 + 3PE"
    if re.search(r"钢管自带3PE加强级外防腐|3PE加强级外防腐", input_text, re.I):
        # 模型值常常已只剩主材，此时补 OUTER。
        if val:
            return "外涂层", [base_item(value=val, outer=["3PE"], special=special)], "原文有3PE外防腐"

    # SPECIAL_REQ-only: keep in special, complete skeleton.
    if special:
        return "特殊要求", [base_item(value=val, special=special)], "特殊要求保留 SPECIAL_REQ"

    return "", None, ""


def append_sheet(wb: Workbook, name: str, rows: list[list[Any]], headers: list[str]):
    ws = wb.create_sheet(name)
    ws.append(headers)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = fill
    for row in rows:
        ws.append(row)
    widths = [10, 18, 90, 55, 90, 45]
    for idx, width in enumerate(widths[: len(headers)], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"


def main():
    data = json.loads(STAGE_PATH.read_text(encoding="utf-8"))
    rows_by_category: dict[str, list[list[Any]]] = {}
    for idx, item in enumerate(data):
        mats = mat_items(item.get("output") or {})
        if not mats:
            continue
        category, new_mats, reason = propose(text(item.get("input")), mats)
        if not category:
            continue
        rows_by_category.setdefault(category, []).append([
            idx,
            category,
            text(item.get("input")),
            json.dumps(mats, ensure_ascii=False),
            json.dumps(new_mats, ensure_ascii=False),
            reason,
        ])

    wb = Workbook()
    wb.remove(wb.active)
    headers = ["索引", "类别", "材料描述", "当前模型MATERIAL", "建议新MATERIAL", "原因"]
    for name in ["内衬", "外涂层", "内外涂层", "多角色材质", "双牌号不拆", "特殊要求", "需人工确认"]:
        append_sheet(wb, name, rows_by_category.get(name, []), headers)
    summary = [[name, len(rows_by_category.get(name, []))] for name in ["内衬", "外涂层", "内外涂层", "多角色材质", "双牌号不拆", "特殊要求", "需人工确认"]]
    append_sheet(wb, "汇总", summary, ["类别", "数量"])
    wb.save(OUT_XLSX)
    print(OUT_XLSX)
    for row in summary:
        print(row[0], row[1])


if __name__ == "__main__":
    main()
