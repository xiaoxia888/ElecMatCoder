#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from collections import OrderedDict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[4]
DEFAULT_INPUT = PROJECT_ROOT / "apps" / "trainer" / "qwen3_fte" / "output" / "flange_project_sampling" / "法兰覆盖抽样.xlsx"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "apps" / "trainer" / "qwen3_fte" / "output" / "flange_project_sampling"


SHEET_NAME = "模板抽样结果"
INPUT_COLUMNS = [
    "材料描述",
    "项目名称",
    "分类",
    "国标美标标记",
    "编码",
    "材料名称代码",
    "公称直径代码",
    "壁厚等级代码",
    "压力等级代码",
    "材质代码",
    "标准号代码",
    "材质分类",
]


def text(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def load_rows(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    rows = ws.iter_rows(values_only=True)
    header = [text(x) for x in next(rows)]
    idx = {name: i for i, name in enumerate(header)}
    missing = [name for name in INPUT_COLUMNS if name not in idx]
    if missing:
        raise KeyError(f"缺少列: {missing}")

    out: list[dict[str, str]] = []
    for row in rows:
        desc = text(row[idx["材料描述"]])
        if not desc:
            continue
        item = {name: text(row[idx[name]]) for name in INPUT_COLUMNS}
        item["Excel行号"] = text(row[idx["Excel行号"]]) if "Excel行号" in idx else ""
        item["抽样原因"] = text(row[idx["抽样原因"]]) if "抽样原因" in idx else ""
        out.append(item)
    return out


def uniq_keep_order(values: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        if not value:
            continue
        if value in seen:
            continue
        seen.add(value)
        out.append(value)
    return out


def infer_body(type_code: str, desc: str) -> tuple[str, list[str]]:
    desc_u = desc.upper()
    compact = re.sub(r"[^A-Z0-9]", "", desc_u)
    notes: list[str] = []

    if any(
        token in desc_u
        for token in [
            "BLANK+SPACER",
            "BLANK & SPACER",
            "BLANK&SPACER",
            "BLANKANDSPACER",
            "SPACER & BLANKS",
            "SPACER&BLANKS",
            "SPACERRING",
            "PADDLEBLANK&SPACER",
            "PADDLE BLANK & SPACER",
            "PADDLE BLANK+SPACER",
        ]
    ) or "插板" in desc or ("盲板" in desc and "垫环" in desc):
        return "插板", notes
    if any(token in desc_u for token in ["PADDLE BLIND", "SPECTACLE BLIND", "SPECTACLE BLANK", "FIGURE-8", "FIGURE8"]) or any(token in desc for token in ["8字盲板", "八字盲板"]):
        return "8字盲板", notes
    if type_code.startswith("8B"):
        notes.append("BODY 按 8B* 类型编码推断为 8字盲板")
        return "8字盲板", notes
    has_jacket = any(token in desc_u for token in ["JACKET", "JACKETED FLANGE"]) or "夹套法兰" in desc or type_code in {"FJ", "BFJ", "JSOF"} or type_code.startswith("J")
    if has_jacket and (type_code == "BFJ" or "盲法兰" in desc):
        return "夹套盲法兰", notes
    if has_jacket and ("JWN" in desc_u or "WN" in desc_u or "带颈对焊" in desc):
        return "带颈对焊夹套法兰", notes
    if has_jacket and ("JSO" in desc_u or "SO" in desc_u or "带颈平焊" in desc or "平焊" in desc):
        return "带颈平焊夹套法兰", notes
    if has_jacket:
        return "夹套法兰", notes
    if any(token in desc_u for token in ["ORIFICE FLANGE"]) or type_code == "ORF":
        return "孔板对焊法兰", notes
    if any(token in desc_u for token in ["LAP JOINT FLANGE", "FLANGE LAPPED"]) or any(token in desc for token in ["松套法兰", "活套法兰"]) or type_code.startswith("LF"):
        return "松套法兰", notes
    if any(token in desc_u for token in ["PAD FLANGE", "PADFLANGE"]) or "板式平焊法兰" in desc or type_code == "FPL":
        return "板式平焊法兰", notes
    if (
        any(token in desc_u for token in ["THREADED FLANGE", "THREADFLANGE", "SCREWED FLANGE"])
        or "螺纹法兰" in desc
        or type_code in {"FTH", "FN", "FFN"}
        or any(token in desc_u for token in ["FNPT", "MNPT", "NPTF"])
        or re.search(r"(?<![A-Z0-9])(?:NPT|THD)(?![A-Z0-9])", desc_u)
        or "TH(NPT)" in desc_u
    ):
        return "螺纹法兰", notes
    if (
        any(token in desc_u for token in ["SOCKET WELD FLANGE", "SOCKET FLANGE"])
        or "承插焊法兰" in desc
        or type_code in {"FS", "FSRJ", "FRJS", "FSLF", "FSLM"}
        or re.search(r"(?<![A-Z0-9])SW(?![A-Z0-9])", desc_u)
        or re.search(r"FLANGE.*SW|SW.*FLANGE", desc_u)
    ):
        return "承插焊法兰", notes
    if "法兰盖" in desc:
        return "法兰盖", notes
    if any(token in desc_u for token in ["BLIND FLANGE", "BLDFLANGE", " BLIND "]) or "盲法兰" in desc or type_code in {"BF", "BFRJ", "BFRF", "BFN"}:
        return "盲法兰", notes
    if type_code in {"BFA", "BFFF", "BFS"}:
        notes.append("BODY 按 B* 法兰盖类型编码推断为 法兰盖")
        return "法兰盖", notes
    if any(token in desc_u for token in ["SO FLANGE", "SLIP ON FLANGE"]) or any(token in desc for token in ["带颈平焊法兰", "平焊法兰"]) or type_code == "FSO":
        return "带颈平焊法兰", notes
    if (
        any(
            token in desc_u
            for token in [
                "WELD NECK FLANGE",
                "WELDNECK FLANGE",
                "WELDING NECK FLANGE",
                "WELDINGNECKFLANGE",
                "WELDINGNECK",
                "FLANGE WELDING NECK",
                "FLANGEWELDINGNECK",
                "WN-FLANGE",
                "WN FLANGE",
                "WNFLANGE",
                "FLANGE WN",
            ]
        )
        or "带颈对焊法兰" in desc
        or "带颈对焊钢制管法兰" in desc
        or re.search(r"(?<![A-Z0-9])WN(?![A-Z0-9])|WN\d|\dWN|WNRF|WNRJ|WNFF|WNMF|WNLM|WNLF", desc_u)
        or re.search(r"(?:RF|RJ|FF|MFM|MF|LM|LF)WN|WN(?:RF|RJ|FF|MFM|MF|LM|LF|ASME|HG|NB|GB|SH|CL|PN|DN|SCH|S\d|$)", compact)
    ):
        return "带颈对焊法兰", notes
    if "对焊法兰" in desc:
        return "带颈对焊法兰", notes
    if "法兰" in desc or "FLANGE" in desc_u:
        notes.append("BODY 仅按原文法兰关键词兜底推断")
        return "法兰", notes
    notes.append("BODY 未命中明确规则，兜底为 法兰")
    return "法兰", notes


def is_non_flange_main_body(desc: str) -> bool:
    desc_u = desc.upper()
    non_flange_patterns = [
        r"FLANGED\s*\d+\s*[°º]?\s*ELBOW",
        r"FLANGED\d+[°º]?ELBOW",
        r"FLANGED\s*ELBOW",
    ]
    return any(re.search(pattern, desc_u) for pattern in non_flange_patterns)


def infer_conn(body: str, desc: str) -> list[str]:
    values: list[str] = []
    desc_u = desc.upper()
    has_thread_end = any(token in desc_u for token in ["FNPT", "MNPT", "NPTF"])

    if " SW " in f" {desc_u} " or ";SW" in desc_u or ",SW" in desc_u or "SOCKET WELD" in desc_u:
        values.append("SW")
    if "TH(NPT)" in desc_u:
        values.append("NPT")
    elif "NPT" in desc_u and not has_thread_end:
        values.append("NPT")
    elif "THD" in desc_u or "THREADED" in desc_u:
        values.append("THD")
    if body == "承插焊法兰" and "SW" not in values:
        values.append("SW")
    if body == "螺纹法兰" and not has_thread_end and not any(v in values for v in ["NPT", "THD"]):
        values.append("THD")
    return uniq_keep_order(values)


def infer_seal(desc: str) -> list[str]:
    desc_u = desc.upper()
    compact = re.sub(r"[^A-Z0-9]", "", desc_u)
    values: list[str] = []
    if "凹凸面" in desc:
        values.append("MFM")
    if "DOUBLE-LIP" in desc_u or "DOUBLE LIP" in desc_u or "双唇" in desc:
        values.append("Double-lip")
    if "WAF/RF" in desc_u:
        values.append("WAF/RF")
    elif "WAF/FF" in desc_u:
        values.append("WAF/FF")
    elif "WAF/RJ" in desc_u:
        values.append("WAF/RJ")
    if "MFM" in desc_u:
        values.append("MFM")
    if re.search(r"(?<![A-Z0-9])MF(?![A-Z0-9])", desc_u):
        values.append("MFM")
    if re.search(r"(?<![A-Z0-9])LM(?![A-Z0-9])", desc_u):
        values.append("LM")
    if re.search(r"(?<![A-Z0-9])LF(?![A-Z0-9])", desc_u):
        values.append("LF")
    if "RTJ" in desc_u:
        values.append("RTJ")
    elif re.search(r"(?<![A-Z0-9])RJ(?![A-Z0-9])", desc_u):
        values.append("RJ")
    elif re.search(r"RJ(?=(?:WN|SO|SW|ASME|HG|NB|GB|SH|CL|PN|DN|SCH|S\d|$))", compact):
        values.append("RJ")
    if re.search(r"(?<![A-Z0-9])FF(?![A-Z0-9])", desc_u):
        values.append("FF")
    if re.search(r"(?<![A-Z0-9])RF(?![A-Z0-9])", desc_u):
        values.append("RF")
    elif re.search(r"RF(?=(?:WN|SO|SW|ASME|HG|NB|GB|SH|CL|PN|DN|SCH|S\d|\d|$))", compact):
        values.append("RF")
    return uniq_keep_order(values)


def infer_ends(desc: str) -> list[str]:
    desc_u = desc.upper()
    values: list[str] = []
    for token in ["FNPT", "MNPT", "MTE", "TSE", "FTE", "NPTF"]:
        if token in desc_u:
            values.append(token)
    return uniq_keep_order(values)


def build_size(size_code: str, desc: str) -> dict[str, list[str]]:
    dn_values: list[str] = []
    code = size_code.upper()
    for match in re.finditer(r"\d+(?:\.\d+)?", code):
        dn_values.append(f"DN{match.group(0).rstrip('0').rstrip('.') if '.' in match.group(0) else match.group(0)}")

    inch_values: list[str] = []
    for match in re.finditer(r'(\d+(?:\s+\d+/\d+)?|\d+/\d+)\s*"{1,2}', desc):
        inch_values.append(match.group(1).replace(" ", "") + '"')

    return {
        "DN": uniq_keep_order(dn_values),
        "OD": [],
        "INCH": uniq_keep_order(inch_values),
        "LENGTH": [],
    }


def build_thickness(thk_code: str, desc: str) -> dict[str, list[str]]:
    mm_values: list[str] = []
    schedule_values: list[str] = []
    series_values: list[str] = []

    code = thk_code.upper().replace(" ", "")
    if code:
        if code in {"XS", "XXS", "STD"}:
            series_values.append(code)
        elif code.endswith("MM"):
            mm_values.append(code[:-2])
        else:
            normalized = code.replace("SCH", "")
            normalized = normalized.lstrip("S-")
            if normalized in {"XS", "XXS", "STD"}:
                series_values.append(normalized)
            elif re.fullmatch(r"\d+(?:\.\d+)?S?", normalized):
                schedule_values.append(f"SCH{normalized}")
            else:
                # 例如 S40S / S160 / XXS x XXS
                parts = re.split(r"[xX×]", code)
                for part in parts:
                    part = part.strip()
                    if not part:
                        continue
                    if part in {"XS", "XXS", "STD"}:
                        series_values.append(part)
                    elif part.endswith("MM"):
                        mm_values.append(part[:-2])
                    else:
                        part_norm = part.replace("SCH", "").lstrip("S-")
                        if part_norm in {"XS", "XXS", "STD"}:
                            series_values.append(part_norm)
                        elif re.fullmatch(r"\d+(?:\.\d+)?S?", part_norm):
                            schedule_values.append(f"SCH{part_norm}")

    if not schedule_values:
        for match in re.finditer(r"\bSCH\.?\s*([0-9]+S?|XXS|XS|STD)\b", desc, re.IGNORECASE):
            token = match.group(1).upper()
            if token in {"XS", "XXS", "STD"}:
                series_values.append(token)
            else:
                schedule_values.append(f"SCH{token}")

    return {
        "MM": uniq_keep_order(mm_values),
        "SCHEDULE": uniq_keep_order(schedule_values),
        "SERIES": uniq_keep_order(series_values),
        "BWG": [],
        "INCH": [],
    }


def build_pressure(pressure_code: str) -> str:
    code = pressure_code.upper().replace(" ", "")
    if not code:
        return ""
    if code.startswith("C") and code[1:].replace(".", "").isdigit():
        return f"CL{code[1:]}"
    return code


def compact_desc(desc: str) -> str:
    return re.sub(r"\s+", "", desc.upper())


def normalize_exec_standard(raw: str) -> str:
    token = raw.upper().replace(" ", "").replace(".", "")
    mapping = [
        (r"^(?:ASTM|ASTMA|SA|A)A?182M?$", "ASTM A182"),
        (r"^(?:ASTM|ASTMA|SA|A)A?105N$", "ASTM A105N"),
        (r"^(?:ASTM|ASTMA|SA|A)A?105M?$", "ASTM A105"),
        (r"^(?:ASTM|ASTMA|SA|A)A?350M?$", "ASTM A350"),
        (r"^(?:ASTM|ASTMA|SA|A)A?216$", "ASTM A216"),
        (r"^(?:ASTM|ASTMA|SA|A)A?240$", "ASTM A240"),
        (r"^(?:ASTM|ASTMA|SA|A)A?694$", "ASTM A694"),
        (r"^(?:ASTM|ASTMA|SA|A)A?351$", "ASTM A351"),
        (r"^(?:NB/T|NBT)47010$", "NB/T47010"),
        (r"^(?:NB/T|NBT)47009$", "NB/T47009"),
        (r"^(?:NB/T|NBT)47008$", "NB/T47008"),
        (r"^(?:GB/T|GBT)12228$", "GB/T12228"),
        (r"^(?:GB/T|GBT)91242$", "GB/T9124.2"),
        (r"^(?:GB/T|GBT)91241$", "GB/T9124.1"),
    ]
    for pattern, normalized in mapping:
        if re.fullmatch(pattern, token):
            return normalized
    return raw.strip()


def normalize_special_req(value: str) -> str:
    token = value.upper().replace(" ", "")
    if not token:
        return ""
    if "NACE" in token:
        return "NACE"
    if any(x in token for x in ["GALV", "镀锌", "ZN", "GALVANIZED"]):
        return "GALV"
    if "ANTI-H2S" in token:
        return "ANTI-H2S"
    if token == "CE":
        return "CE"
    return value.strip()


def normalize_material_grade(raw: str) -> str:
    value = text(raw).strip().strip(",;")
    if not value:
        return ""
    value = value.replace("_", " ")
    value = re.sub(r"\s+", " ", value)
    value = re.sub(r"(?i)^GRADE\s*", "", value)
    value = re.sub(r"(?i)^GR\.?\s*", "", value)
    value = re.sub(r"(?i)^GRADE\.?\s*", "", value)
    value = re.sub(r"\s+", "", value)
    value = value.strip("-/,;")
    if not value:
        return ""
    if re.fullmatch(r"[A-Z]$", value):
        return f"Gr.{value}"
    if re.fullmatch(r"[A-Z]\.?[0-9A-Z.]*", value) and value == "B":
        return "Gr.B"
    return value


def material_family(value: str) -> str:
    token = normalize_material_grade(value).upper()
    token = token.replace("GR.", "")
    token = re.sub(r"[^A-Z0-9#]+", "", token)
    return token


MATERIAL_DESC_PATTERNS: list[tuple[str, str]] = [
    (
        "ASTM A182",
        r"(?:ASTMA182M?|SA182|A182M?)(?:GR(?:ADE)?\.?)?([A-Z0-9#.\-()/≤]+?)(?=(?:CL\d|PN\d|DN\d|RF|FF|RJ|RTJ|MFM|WAF|ASMEB|HG/T|HGT|NB/T|NBT|GB/T|GBT|SH/T|SHT|EN\d|SCH|SERIAL|SERIES|$))",
    ),
    (
        "ASTM A105N",
        r"(?:ASTMA105N|A105N)(?:GR(?:ADE)?\.?)?([A-Z0-9#.\-()/≤]+?)(?=(?:CL\d|PN\d|DN\d|RF|FF|RJ|RTJ|ASMEB|HG/T|HGT|NB/T|NBT|GB/T|GBT|SH/T|SHT|EN\d|SCH|SERIAL|SERIES|$))",
    ),
    (
        "ASTM A105",
        r"(?:ASTMA105M?|A105M?)(?:GR(?:ADE)?\.?)?([A-Z0-9#.\-()/≤]+?)(?=(?:CL\d|PN\d|DN\d|RF|FF|RJ|RTJ|ASMEB|HG/T|HGT|NB/T|NBT|GB/T|GBT|SH/T|SHT|EN\d|SCH|SERIAL|SERIES|$))",
    ),
    (
        "ASTM A350",
        r"(?:ASTMA350M?|A350M?)(?:GR(?:ADE)?\.?)?([A-Z0-9#.\-()/≤]+?)(?=(?:CL\d|PN\d|DN\d|RF|FF|RJ|RTJ|ASMEB|HG/T|HGT|NB/T|NBT|GB/T|GBT|SH/T|SHT|EN\d|SCH|SERIAL|SERIES|$))",
    ),
    (
        "ASTM A216",
        r"(?:ASTMA216|A216)(?:GR(?:ADE)?\.?)?([A-Z0-9#.\-()/≤]+?)(?=(?:CL\d|PN\d|DN\d|RF|FF|RJ|RTJ|ASMEB|HG/T|HGT|NB/T|NBT|GB/T|GBT|SH/T|SHT|EN\d|SCH|SERIAL|SERIES|$))",
    ),
    (
        "NB/T47010",
        r"(?:NBT47010)[-/]?([A-Z0-9#.\-()/≤]+?)(?=(?:CL\d|PN\d|DN\d|RF|FF|RJ|RTJ|ASMEB|HG/T|HGT|NB/T|NBT|GB/T|GBT|SH/T|SHT|EN\d|SCH|SERIAL|SERIES|$))",
    ),
    (
        "NB/T47009",
        r"(?:NBT47009)[-/]?([A-Z0-9#.\-()/≤]+?)(?=(?:CL\d|PN\d|DN\d|RF|FF|RJ|RTJ|ASMEB|HG/T|HGT|NB/T|NBT|GB/T|GBT|SH/T|SHT|EN\d|SCH|SERIAL|SERIES|$))",
    ),
    (
        "NB/T47008",
        r"(?:NBT47008)[-/]?([A-Z0-9#.\-()/≤]+?)(?=(?:CL\d|PN\d|DN\d|RF|FF|RJ|RTJ|ASMEB|HG/T|HGT|NB/T|NBT|GB/T|GBT|SH/T|SHT|EN\d|SCH|SERIAL|SERIES|$))",
    ),
    (
        "GB/T12228",
        r"(?:GBT12228)[-/]?([A-Z0-9#.\-()/≤]+?)(?=(?:CL\d|PN\d|DN\d|RF|FF|RJ|RTJ|ASMEB|HG/T|HGT|NB/T|NBT|GB/T|GBT|SH/T|SHT|EN\d|SCH|SERIAL|SERIES|$))",
    ),
]


def extract_material_candidates(desc: str) -> list[dict[str, Any]]:
    compact = compact_desc(desc)
    found: list[tuple[int, dict[str, Any]]] = []
    for exec_standard, pattern in MATERIAL_DESC_PATTERNS:
        for match in re.finditer(pattern, compact, re.IGNORECASE):
            grade = normalize_material_grade(match.group(1))
            if not grade:
                continue
            found.append(
                (
                    match.start(),
                    {
                        "EXEC_STANDARD": exec_standard,
                        "GRADE": grade,
                        "SPECIAL_REQ": [],
                    },
                )
            )
    found.sort(key=lambda x: x[0])
    out: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for _, item in found:
        key = (item["EXEC_STANDARD"], item["GRADE"])
        if key in seen:
            continue
        seen.add(key)
        out.append(item)
    return out


def infer_exec_standard_from_desc(part: str, desc: str) -> str:
    family = material_family(part)
    compact = compact_desc(desc)
    if not family:
        return ""
    checks = [
        ("ASTM A182", [r"ASTMA182", r"SA182", r"A182"]),
        ("ASTM A105N", [r"ASTMA105N", r"A105N"]),
        ("ASTM A105", [r"ASTMA105", r"A105"]),
        ("ASTM A350", [r"ASTMA350", r"A350"]),
        ("ASTM A216", [r"ASTMA216", r"A216"]),
        ("NB/T47010", [r"NBT47010"]),
        ("NB/T47009", [r"NBT47009"]),
        ("NB/T47008", [r"NBT47008"]),
        ("GB/T12228", [r"GBT12228"]),
    ]
    for normalized, patterns in checks:
        for pattern in patterns:
            if re.search(pattern, compact):
                if normalized == "ASTM A182" and re.match(r"^(F|TP|WP|304|304L|316|316L|321|347|310|2205|2507)", family):
                    return normalized
                if normalized == "ASTM A105N" and family.startswith("A105N"):
                    return normalized
                if normalized == "ASTM A105" and family.startswith("A105"):
                    return normalized
                if normalized == "ASTM A350" and family.startswith(("LF", "A350")):
                    return normalized
                if normalized == "ASTM A216" and family.startswith(("WCB", "WC6", "CF")):
                    return normalized
                if normalized.startswith("NB/T") and family:
                    return normalized
                if normalized == "GB/T12228" and family:
                    return normalized
    return ""


def split_material_parts(material_code: str, desc: str) -> tuple[str, list[str]]:
    raw = text(material_code).strip()
    if not raw:
        return "", []
    parts = [text(x) for x in re.split(r"\s*/\s*", raw) if text(x)]
    if not parts:
        return "", []

    composite_markers = ["PTFE", "RPTFE", "衬胶", "LINED", "CLAD", "覆层", "衬里", "PFA", "FEP"]
    joined = raw.upper() + " " + desc.upper()
    if len(parts) > 1 and any(marker in joined for marker in composite_markers):
        return "composite", parts
    if len(parts) > 1:
        return "alternative", parts
    return "single", parts


def build_material_item_from_part(part: str, desc: str, candidates: list[dict[str, Any]]) -> dict[str, Any]:
    normalized_part = normalize_material_grade(part)
    family = material_family(normalized_part)

    for candidate in candidates:
        candidate_family = material_family(candidate["GRADE"])
        if family and candidate_family and (candidate_family.endswith(family) or family.endswith(candidate_family)):
            return {
                "EXEC_STANDARD": candidate["EXEC_STANDARD"],
                "GRADE": candidate["GRADE"],
                "SPECIAL_REQ": candidate["SPECIAL_REQ"],
            }

    exec_standard = infer_exec_standard_from_desc(normalized_part, desc)
    return {
        "EXEC_STANDARD": exec_standard,
        "GRADE": normalized_part,
        "SPECIAL_REQ": [],
    }


def compose_material_value(item: dict[str, Any]) -> str:
    exec_standard = text(item.get("EXEC_STANDARD"))
    grade = text(item.get("GRADE"))
    if exec_standard and grade:
        return f"{exec_standard} {grade}"
    if grade:
        return grade
    return exec_standard


def build_material(material_code: str, desc: str) -> list[dict[str, Any]]:
    relation, parts = split_material_parts(material_code, desc)
    if not parts:
        return []

    candidates = extract_material_candidates(desc)
    items: list[dict[str, Any]] = []
    for part in parts:
        item = build_material_item_from_part(part, desc, candidates)
        if item["GRADE"]:
            items.append(item)

    desc_u = desc.upper()
    extra_special_req: list[str] = []
    if "NACE" in desc_u:
        extra_special_req.append("NACE")
    if any(token in desc_u for token in ["GALV", "GALVANIZED", "镀锌"]):
        extra_special_req.append("GALV")
    if "ANTI-H2S" in desc_u:
        extra_special_req.append("ANTI-H2S")
    if re.search(r"(?<![A-Z0-9])CE(?![A-Z0-9])", desc_u):
        extra_special_req.append("CE")
    if extra_special_req:
        for item in items:
            item["SPECIAL_REQ"] = uniq_keep_order(item["SPECIAL_REQ"] + extra_special_req)

    if not items:
        return []

    primary = items[0]
    value = compose_material_value(primary)
    if not value:
        return []
    return [
        {
            "ROLE": "MAIN",
            "VALUE": value,
            "SPECIAL_REQ": primary["SPECIAL_REQ"],
        }
    ]


STANDARD_PATTERNS: list[tuple[str, str, str | None]] = [
    ("AB1648", "ASME B16.48", None),
    ("AB1647", "ASME B16.47", None),
    ("AB1636", "ASME B16.36", None),
    ("AB165", "ASME B16.5", None),
    ("AB318", "ASME B31.8", None),
    ("GBT91242", "GB/T9124.2", None),
    ("GBT91241", "GB/T9124.1", None),
    ("GBT12228", "GB/T12228", None),
    ("GBT21547", "GB/T21547", None),
    ("NBT47010", "NB/T47010", None),
    ("NBT47009", "NB/T47009", None),
    ("NBT47008", "NB/T47008", None),
    ("SHT3425", "SH/T3425", None),
    ("SHT3406", "SH/T3406", None),
    ("HGT21547", "HG/T21547", None),
    ("HGT20623B", "HG/T20623", "B"),
    ("HGT20623A", "HG/T20623", "A"),
    ("HGT20623", "HG/T20623", None),
    ("HGT20615B", "HG/T20615", "B"),
    ("HGT20615A", "HG/T20615", "A"),
    ("HGT20615", "HG/T20615", None),
    ("HGT20592B", "HG/T20592", "B"),
    ("HGT20592A", "HG/T20592", "A"),
    ("HGT20592", "HG/T20592", None),
    ("EN10921", "EN1092-1", None),
]


def build_standard(standard_code: str, desc: str) -> list[dict[str, str]]:
    code = standard_code.upper().replace(" ", "")
    found: list[tuple[int, dict[str, str]]] = []
    for token, body, grade in STANDARD_PATTERNS:
        start = 0
        while True:
            idx = code.find(token, start)
            if idx < 0:
                break
            found.append(
                (
                    idx,
                    {
                        "BODY": body,
                        "GRADE": grade or "",
                        "METHOD": "",
                        "APPENDIX": "",
                    },
                )
            )
            start = idx + len(token)

    for raw, body, grade in [
        (r"ASME\s*B16\.48", "ASME B16.48", ""),
        (r"ASME\s*B16\.47", "ASME B16.47", ""),
        (r"ASME\s*B16\.36", "ASME B16.36", ""),
        (r"ASME\s*B16\.5", "ASME B16.5", ""),
        (r"GB/T\s*12228", "GB/T12228", ""),
        (r"GB/T\s*9124\.1", "GB/T9124.1", ""),
        (r"GB/T\s*9124\.2", "GB/T9124.2", ""),
        (r"NB/T\s*47010", "NB/T47010", ""),
        (r"NB/T\s*47009", "NB/T47009", ""),
        (r"NB/T\s*47008", "NB/T47008", ""),
        (r"HG/T\s*20623(?:\s*[-(]?\s*B)?", "HG/T20623", "B"),
        (r"HG/T\s*20615(?:\s*[-(]?\s*([AB]))?", "HG/T20615", ""),
        (r"HG/T\s*20592(?:\s*[-(]?\s*([AB]))?", "HG/T20592", ""),
        (r"HG/T\s*21547", "HG/T21547", ""),
        (r"SH/T\s*3406", "SH/T3406", ""),
        (r"SH/T\s*3425", "SH/T3425", ""),
        (r"EN\s*1092-1", "EN1092-1", ""),
    ]:
        m = re.search(raw, desc, re.IGNORECASE)
        if m:
            resolved_grade = grade
            if body in {"HG/T20615", "HG/T20592"} and m.lastindex:
                resolved_grade = text(m.group(1)).upper()
            found.append((m.start(), {"BODY": body, "GRADE": resolved_grade, "METHOD": "", "APPENDIX": ""}))

    found.sort(key=lambda x: x[0])
    dedup: list[dict[str, str]] = []
    seen: set[tuple[str, str, str, str]] = set()
    for _, item in found:
        key = (item["BODY"], item["GRADE"], item["METHOD"], item["APPENDIX"])
        if key in seen:
            continue
        seen.add(key)
        dedup.append(item)

    preferred: OrderedDict[tuple[str, str, str], dict[str, str]] = OrderedDict()
    for item in dedup:
        coarse_key = (item["BODY"], item["METHOD"], item["APPENDIX"])
        existing = preferred.get(coarse_key)
        if existing is None:
            preferred[coarse_key] = item
            continue
        if existing["GRADE"] == "" and item["GRADE"] != "":
            preferred[coarse_key] = item

    return list(preferred.values())


def make_output(row: dict[str, str]) -> tuple[dict[str, Any], list[str]]:
    desc = row["材料描述"]
    type_code = row["材料名称代码"]

    body, notes = infer_body(type_code, desc)
    conn = infer_conn(body, desc)
    seal = infer_seal(desc)
    ends = infer_ends(desc)

    output = {
        "TYPE": {
            "BODY": body,
            "CONN": conn,
            "SEAL": seal,
            "ENDS": ends,
        },
        "SIZE": build_size(row["公称直径代码"], desc),
        "THICKNESS": build_thickness(row["壁厚等级代码"], desc),
        "PRESSURE": build_pressure(row["压力等级代码"]),
        "MATERIAL": build_material(row["材质代码"], desc),
        "STANDARD": build_standard(row["标准号代码"], desc),
    }

    if not seal:
        notes.append("SEAL 未自动识别")
    if body in {"法兰", "带颈对焊法兰"} and "BODY 按" in " ".join(notes):
        notes.append("BODY 依赖类型编码推断，建议复核")
    if not output["STANDARD"]:
        notes.append("STANDARD 未自动识别")

    return output, uniq_keep_order(notes)


def write_review_xlsx(path: Path, rows: list[dict[str, Any]], excluded_rows: list[dict[str, Any]] | None = None) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "需审核"
    headers = [
        "材料描述",
        "编码",
        "材料名称代码",
        "自动BODY",
        "自动CONN",
        "自动SEAL",
        "自动ENDS",
        "自动SIZE",
        "自动THICKNESS",
        "自动PRESSURE",
        "自动MATERIAL",
        "自动STANDARD",
        "审核备注",
        "Excel行号",
        "抽样原因",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    for col in ws.columns:
        width = min(max(len(text(cell.value)) for cell in col[:20]) + 2, 80)
        ws.column_dimensions[col[0].column_letter].width = max(12, width)

    if excluded_rows is not None:
        ws2 = wb.create_sheet("排除样本")
        exclude_headers = ["材料描述", "编码", "材料名称代码", "排除原因", "Excel行号", "抽样原因"]
        ws2.append(exclude_headers)
        for row in excluded_rows:
            ws2.append([row.get(h, "") for h in exclude_headers])
        for col in ws2.columns:
            width = min(max(len(text(cell.value)) for cell in col[:20]) + 2, 80)
            ws2.column_dimensions[col[0].column_letter].width = max(12, width)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="将法兰抽样 Excel 转换为训练草稿")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args()

    rows = load_rows(args.input)
    draft_records: list[dict[str, Any]] = []
    review_rows: list[dict[str, Any]] = []
    excluded_rows: list[dict[str, Any]] = []
    seen: "OrderedDict[tuple[str, str], dict[str, Any]]" = OrderedDict()

    for row in rows:
        if is_non_flange_main_body(row["材料描述"]):
            excluded_rows.append(
                {
                    "材料描述": row["材料描述"],
                    "编码": row["编码"],
                    "材料名称代码": row["材料名称代码"],
                    "排除原因": "主体不是法兰，疑似带法兰弯头/管件",
                    "Excel行号": row["Excel行号"],
                    "抽样原因": row["抽样原因"],
                }
            )
            continue

        output, notes = make_output(row)
        record = {"input": row["材料描述"], "output": output}
        key = (record["input"], json.dumps(record["output"], ensure_ascii=False, sort_keys=True))
        if key not in seen:
            seen[key] = record

        review_rows.append(
            {
                "材料描述": row["材料描述"],
                "编码": row["编码"],
                "材料名称代码": row["材料名称代码"],
                "自动BODY": output["TYPE"]["BODY"],
                "自动CONN": json.dumps(output["TYPE"]["CONN"], ensure_ascii=False),
                "自动SEAL": json.dumps(output["TYPE"]["SEAL"], ensure_ascii=False),
                "自动ENDS": json.dumps(output["TYPE"]["ENDS"], ensure_ascii=False),
                "自动SIZE": json.dumps(output["SIZE"], ensure_ascii=False),
                "自动THICKNESS": json.dumps(output["THICKNESS"], ensure_ascii=False),
                "自动PRESSURE": output["PRESSURE"],
                "自动MATERIAL": json.dumps(output["MATERIAL"], ensure_ascii=False),
                "自动STANDARD": json.dumps(output["STANDARD"], ensure_ascii=False),
                "审核备注": "；".join(notes),
                "Excel行号": row["Excel行号"],
                "抽样原因": row["抽样原因"],
            }
        )

    draft_records = list(seen.values())

    args.output_dir.mkdir(parents=True, exist_ok=True)
    draft_path = args.output_dir / "法兰训练草稿.json"
    review_path = args.output_dir / "法兰训练草稿_需审核.xlsx"
    draft_path.write_text(json.dumps(draft_records, ensure_ascii=False, indent=2), encoding="utf-8")
    write_review_xlsx(review_path, review_rows, excluded_rows)

    print(f"输入抽样行数: {len(rows)}")
    print(f"排除非纯法兰主体条数: {len(excluded_rows)}")
    print(f"去重后训练草稿条数: {len(draft_records)}")
    print(f"训练草稿: {draft_path}")
    print(f"审核表: {review_path}")


if __name__ == "__main__":
    main()
