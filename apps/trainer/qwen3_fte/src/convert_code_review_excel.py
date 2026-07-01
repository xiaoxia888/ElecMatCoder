#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
将编码结果核对 Excel 转换为审核/导入用的新表结构。

输入表头示例：
序号 原始描述 原始总编码 修正总编码 是否需审核 最低相似度 分类
TYPE_原始结果 TYPE_原始编码 TYPE_修正结果 TYPE_修正编码
SIZE_原始结果 SIZE_原始编码 SIZE_修正结果 SIZE_修正编码
THICKNESS_原始结果 THICKNESS_原始编码 THICKNESS_修正结果 THICKNESS_修正编码
PRESSURE_原始结果 PRESSURE_原始编码 PRESSURE_修正结果 PRESSURE_修正编码
MATERIAL_原始结果 MATERIAL_原始编码 MATERIAL_修正结果 MATERIAL_修正编码
STANDARD_原始结果 STANDARD_原始编码 STANDARD_修正结果 STANDARD_修正编码

输出表头：
描述 描述(多行) 项目名称 分类 国标美标标记 编码 置信度 来源 修正后编码
原始种类 标准化种类 修正种类
原始尺寸 标准化尺寸 修正尺寸
原始壁厚 标准化壁厚 修正壁厚
原始磅级 标准化磅级 修正磅级
原始材质 标准化材质 修正材质
原始规范 标准化规范


  - --shuffle
      - 按分类尽量均衡打散后再拆分
      - 默认开启
  - --no-shuffle
      - 保持原始顺序拆分，不打乱

  ### 只拆 Excel，不上传

  python apps/trainer/qwen3_fte/src/convert_code_review_excel.py \
    /Users/guoxi/Downloads/encodings_2026-05-26.csv \
    --chunk-size 200 \
    --no-shuffle





  ### 拆 Excel + 暂存到氚云草稿

  python apps/trainer/qwen3_fte/src/convert_code_review_excel.py \
    /path/to/encodings.xlsx \
    --chunk-size 200 \
    --no-shuffle \
    --upload-h3 \
    --h3-engine-code '你的EngineCode' \
    --h3-engine-secret '你的EngineSecret'

python apps/trainer/qwen3_fte/src/convert_code_review_excel.py \
    /Users/guoxi/Downloads/管件处理0629.xlsx \
    --upload-h3 \
    --chunk-size 500 \
    --no-shuffle \
    --h3-upload-mode controller \
    --h3-created-by-name "夏国玺"

  ### 如果要直接提交生效

  python apps/trainer/qwen3_fte/src/convert_code_review_excel.py \
    /path/to/encodings.xlsx \
    --chunk-size 200 \
    --upload-h3 \
    --h3-engine-code '你的EngineCode' \
    --h3-engine-secret '你的EngineSecret' \
    --h3-submit

python apps/trainer/qwen3_fte/src/convert_code_review_excel.py /Users/guoxi/Downloads/encodings_2026-05-18.csv --chunk-size 200 --no-shuffle

"""

from __future__ import annotations

import argparse
import csv
import json
import os
import random
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List

from openpyxl import Workbook, load_workbook
import requests


OUTPUT_HEADERS = [
    "描述",
    "描述(多行)",
    "项目名称",
    "分类",
    "国标美标标记",
    "编码",
    "置信度",
    "来源",
    "修正后编码",
    "原始种类",
    "标准化种类",
    "修正种类",
    "原始尺寸",
    "标准化尺寸",
    "修正尺寸",
    "原始壁厚",
    "标准化壁厚",
    "修正壁厚",
    "原始磅级",
    "标准化磅级",
    "修正磅级",
    "原始材质",
    "标准化材质",
    "修正材质",
    "原始规范",
    "标准化规范",
    "难度",
]

H3_DEFAULT_URL = "https://www.h3yun.com/OpenApi/Invoke"
H3_REVIEW_SCHEMA_CODE = "D148357c862f0c8cdfa41418c55cef288f8d83c"
H3_REVIEW_SUBTABLE_CODE = "D148357F17c2e0548b94497f873300934ea06164"
H3_REVIEW_APP_CODE = "D148357CLDGGL"
H3_REVIEW_CONTROLLER = "ReviewTaskListApiController"
H3_REVIEW_CREATE_DRAFT_ACTION = "createReviewTaskDraft"
H3_REVIEW_DEBUG_PROJECT_ACTION = "debugProjectLookup"
H3_SUBTABLE_FIELD_MAPPING = {
    "描述": "F0000003",
    "描述(多行)": "F0000056",
    "项目名称": "F0000049",
    "分类": "F0000050",
    "国标美标标记": "F0000051",
    "编码": "F0000004",
    "置信度": "F0000005",
    "来源": "F0000007",
    "原始种类": "F0000009",
    "标准化种类": "F0000010",
    "原始尺寸": "F0000012",
    "标准化尺寸": "F0000013",
    "原始壁厚": "F0000015",
    "标准化壁厚": "F0000016",
    "原始磅级": "F0000018",
    "标准化磅级": "F0000019",
    "原始材质": "F0000021",
    "标准化材质": "F0000022",
    "原始规范": "F0000024",
    "标准化规范": "F0000025",
    "难度": "F0000071",
}


def text(v: object) -> str:
    # 数字 0 是有效的难度值，不能按普通 falsy 值转成空字符串。
    return "" if v is None else str(v).strip()


def first_nonempty(src: Dict[str, str], *keys: str) -> str:
    for key in keys:
        value = text(src.get(key))
        if value:
            return value
    return ""


def infer_standard_flag(standard_code: str, standard_raw: str = "") -> str:
    """
    通过规范编码/规范原文推断“国标美标标记”。

    返回值：
    - 国标
    - 美标
    - 欧标
    - 日标
    - 多种标准用 ` + ` 连接
    - 未识别
    """
    src = f"{text(standard_code)} {text(standard_raw)}".upper()

    # 平台里规范多数已经是编码后的结果，例如：
    # - 国标/行标: GBT12459 / SHT3408 / HGT20553IA / NBT47008
    # - 美标: AB165 / AB1611 / AB3610 / AB3619 / ASTM / API / MSS
    # 这里优先按编码前缀判断，同时兼容原始写法。
    has_gb = bool(
        re.search(
            r"GB/?T|SH/?T|HG/?T|NB/?T|JB/?T|SY/?T|DL/?T|YB/?T|"
            r"GBT\d|SHT\d|HGT\d|NBT\d|JBT\d|SYT\d|DLT\d|YBT\d|"
            r"CJT\d|GJB\d|QBT\d|TSG\d",
            src,
        )
    )
    has_us = bool(
        re.search(
            r"ASME|ASTM|API|MSS|ANSI|AWWA|NACE|"
            r"AB\d|AA\d|MS\d",
            src,
        )
    )
    has_eu = bool(re.search(r"\bDIN\b|\bEN\b|\bISO\b|\bBS\b|DIN\d|EN\d|ISO\d|BS\d", src))
    has_jp = bool(re.search(r"\bJIS\b|JIS\d", src))

    parts: List[str] = []
    if has_gb:
        parts.append("国标")
    if has_us:
        parts.append("美标")
    if has_eu:
        parts.append("欧标")
    if has_jp:
        parts.append("日标")
    return " + ".join(parts) if parts else ""


def split_standard_flags(flag: str) -> List[str]:
    return [part.strip() for part in text(flag).split("+") if text(part)]


STANDARD_FLAG_CANDIDATES = ("国标", "美标", "欧标", "日标")


def build_project_standard_bias(rows: List[Dict[str, str]]) -> Dict[str, str]:
    """
    按项目统计主流标准倾向。

    规则：
    - 只统计单一标签行
    - 混合标签不参与计票
    - 支持 国标 / 美标 / 欧标 / 日标
    - 仅当某一类票数严格最多时返回该标签，否则返回空
    """
    counters: Dict[str, Dict[str, int]] = defaultdict(
        lambda: {flag: 0 for flag in STANDARD_FLAG_CANDIDATES}
    )
    for row in rows:
        project_name = first_nonempty(row, "项目名称", "项目", "项目名")
        if not project_name:
            continue
        flag = infer_standard_flag(text(row.get("STANDARD_原始编码")), text(row.get("STANDARD_原始结果")))
        parts = split_standard_flags(flag)
        if len(parts) != 1:
            continue
        if parts[0] in STANDARD_FLAG_CANDIDATES:
            counters[project_name][parts[0]] += 1

    bias: Dict[str, str] = {}
    for project_name, counts in counters.items():
        max_count = max(counts.values(), default=0)
        if max_count <= 0:
            bias[project_name] = ""
            continue
        leaders = [flag for flag, count in counts.items() if count == max_count]
        bias[project_name] = leaders[0] if len(leaders) == 1 else ""
    return bias


def build_output_row(src: Dict[str, str], project_standard_bias: Dict[str, str] | None = None) -> Dict[str, str]:
    standard_code = text(src.get("STANDARD_原始编码"))
    standard_raw = text(src.get("STANDARD_原始结果"))
    project_name = first_nonempty(src, "项目名称", "项目", "项目名")
    standard_flag = infer_standard_flag(standard_code, standard_raw)
    parts = split_standard_flags(standard_flag)
    if len(parts) >= 2:
        project_bias = text((project_standard_bias or {}).get(project_name))
        if project_bias in STANDARD_FLAG_CANDIDATES:
            standard_flag = project_bias
    return {
        "描述": text(src.get("原始描述")),
        "描述(多行)": text(src.get("原始描述")),
        "项目名称": project_name,
        "分类": text(src.get("分类")),
        "国标美标标记": standard_flag,
        "编码": text(src.get("原始总编码")),
        "置信度": text(src.get("总置信度")),
        "来源": "大模型微调",
        "修正后编码": "",
        "原始种类": text(src.get("TYPE_原始结果")),
        "标准化种类": text(src.get("TYPE_原始编码")),
        "修正种类": "",
        "原始尺寸": text(src.get("SIZE_原始结果")),
        "标准化尺寸": text(src.get("SIZE_原始编码")),
        "修正尺寸": "",
        "原始壁厚": text(src.get("THICKNESS_原始结果")),
        "标准化壁厚": text(src.get("THICKNESS_原始编码")),
        "修正壁厚": "",
        "原始磅级": text(src.get("PRESSURE_原始结果")),
        "标准化磅级": text(src.get("PRESSURE_原始编码")),
        "修正磅级": "",
        "原始材质": text(src.get("MATERIAL_原始结果")),
        "标准化材质": text(src.get("MATERIAL_原始编码")),
        "修正材质": "",
        "原始规范": standard_raw,
        "标准化规范": standard_code,
        "难度": first_nonempty(src, "分流最终难度（0=困难，1=中等，2=简单）", "难度"),
    }


def read_excel_rows(input_path: Path) -> List[Dict[str, str]]:
    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [text(x) for x in next(rows)]

    result: List[Dict[str, str]] = []
    for row in rows:
        item = {headers[i]: text(v) for i, v in enumerate(row) if i < len(headers)}
        if not text(item.get("原始描述")):
            continue
        result.append(item)
    return result


def read_csv_rows(input_path: Path) -> List[Dict[str, str]]:
    encodings = ["utf-8-sig", "utf-8", "gb18030", "gbk"]
    last_error: Exception | None = None

    for encoding in encodings:
        try:
            with input_path.open("r", encoding=encoding, newline="") as f:
                reader = csv.DictReader(f)
                result: List[Dict[str, str]] = []
                for row in reader:
                    item = {text(k): text(v) for k, v in (row or {}).items()}
                    if not text(item.get("原始描述")):
                        continue
                    result.append(item)
                return result
        except UnicodeDecodeError as exc:
            last_error = exc

    raise UnicodeDecodeError(
        getattr(last_error, "encoding", "unknown"),
        getattr(last_error, "object", b""),
        getattr(last_error, "start", 0),
        getattr(last_error, "end", 0),
        "无法识别 CSV 编码，请检查文件编码。",
    )


def read_rows(input_path: Path) -> List[Dict[str, str]]:
    suffix = input_path.suffix.lower()
    if suffix == ".csv":
        return read_csv_rows(input_path)
    return read_excel_rows(input_path)


def balanced_shuffle(rows: List[Dict[str, str]], seed: int) -> List[Dict[str, str]]:
    """
    尽量按“分类”均衡打散，避免连续大量同类。
    """
    rng = random.Random(seed)
    buckets: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    for row in rows:
        key = text(row.get("分类")) or "未分类"
        buckets[key].append(row)

    for group in buckets.values():
        rng.shuffle(group)

    ordered_groups = sorted(buckets.items(), key=lambda kv: (-len(kv[1]), kv[0]))
    result: List[Dict[str, str]] = []
    last_cat = None

    while True:
        available = [(cat, items) for cat, items in ordered_groups if items]
        if not available:
            break

        # 优先选非上一个分类，且剩余数量最多的桶
        candidates = [(cat, items) for cat, items in available if cat != last_cat] or available
        candidates.sort(key=lambda kv: (-len(kv[1]), kv[0]))
        cat, items = candidates[0]
        result.append(items.pop())
        last_cat = cat

    return result


def write_chunk(path: Path, rows: Iterable[Dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(OUTPUT_HEADERS)
    for row in rows:
        ws.append([text(row.get(col)) for col in OUTPUT_HEADERS])
    wb.save(path)


def build_h3_subtable_row(row: Dict[str, str]) -> Dict[str, str]:
    return {
        field_code: text(row.get(column))
        for column, field_code in H3_SUBTABLE_FIELD_MAPPING.items()
    }


def build_h3_biz_object(
    chunk_rows_data: List[Dict[str, str]],
    *,
    batch_index: int,
    total_batches: int,
    name_prefix: str,
    created_by_user_id: str = "",
) -> Dict[str, object]:
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    biz_object = {
        "Name": f"{name_prefix}{timestamp} 第{batch_index}/{total_batches}批",
        H3_REVIEW_SUBTABLE_CODE: [build_h3_subtable_row(row) for row in chunk_rows_data],
    }
    creator_id = text(created_by_user_id)
    if creator_id:
        biz_object["CreatedBy"] = creator_id
        biz_object["OwnerId"] = creator_id
    return biz_object


def build_h3_task_name(*, batch_index: int, total_batches: int, name_prefix: str) -> str:
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return f"{name_prefix}{timestamp} 第{batch_index}/{total_batches}批"


def build_h3_controller_item(row: Dict[str, str]) -> Dict[str, str]:
    return {
        "description": text(row.get("描述")),
        "projectName": text(row.get("项目名称")),
        "category": text(row.get("分类")),
        "standardFlag": text(row.get("国标美标标记")),
        "code": text(row.get("编码")),
        "confidence": text(row.get("置信度")),
        "source": text(row.get("来源")),
        "rawType": text(row.get("原始种类")),
        "normalizedType": text(row.get("标准化种类")),
        "rawSize": text(row.get("原始尺寸")),
        "normalizedSize": text(row.get("标准化尺寸")),
        "rawThickness": text(row.get("原始壁厚")),
        "normalizedThickness": text(row.get("标准化壁厚")),
        "rawPressure": text(row.get("原始磅级")),
        "normalizedPressure": text(row.get("标准化磅级")),
        "rawMaterial": text(row.get("原始材质")),
        "normalizedMaterial": text(row.get("标准化材质")),
        "rawStandard": text(row.get("原始规范")),
        "normalizedStandard": text(row.get("标准化规范")),
        "difficulty": text(row.get("难度")),
    }


def h3_request(
    *,
    url: str,
    engine_code: str,
    engine_secret: str,
    payload: Dict[str, object],
    timeout: int,
) -> Dict[str, object]:
    headers = {
        "Content-Type": "application/json",
        "EngineCode": engine_code,
        "EngineSecret": engine_secret,
    }
    response = requests.post(url, headers=headers, json=payload, timeout=timeout)
    response.raise_for_status()
    return response.json()


def upload_chunks_to_h3(
    chunks: List[List[Dict[str, str]]],
    *,
    url: str,
    engine_code: str,
    engine_secret: str,
    schema_code: str,
    is_submit: bool,
    timeout: int,
    name_prefix: str,
    created_by_user_id: str,
) -> Dict[str, object]:
    biz_objects = [
        build_h3_biz_object(
            chunk,
            batch_index=idx,
            total_batches=len(chunks),
            name_prefix=name_prefix,
            created_by_user_id=created_by_user_id,
        )
        for idx, chunk in enumerate(chunks, 1)
        if chunk
    ]
    if not biz_objects:
        return {"Successful": False, "ErrorMessage": "没有可上传的数据"}

    payload = {
        "ActionName": "CreateBizObjects",
        "SchemaCode": schema_code,
        "BizObjectArray": [json.dumps(obj, ensure_ascii=False) for obj in biz_objects],
        "IsSubmit": bool(is_submit),
    }
    result = h3_request(
        url=url,
        engine_code=engine_code,
        engine_secret=engine_secret,
        payload=payload,
        timeout=timeout,
    )
    result["__biz_object_names"] = [text(obj.get("Name")) for obj in biz_objects]
    return result


def upload_chunks_to_h3_controller(
    chunks: List[List[Dict[str, str]]],
    *,
    url: str,
    engine_code: str,
    engine_secret: str,
    app_code: str,
    controller: str,
    action_name: str,
    timeout: int,
    name_prefix: str,
    created_by_name: str,
    created_by_user_id: str,
) -> Dict[str, object]:
    created: List[Dict[str, object]] = []
    names: List[str] = []
    total_batches = len(chunks)
    for idx, chunk in enumerate(chunks, 1):
        if not chunk:
            continue
        task_name = build_h3_task_name(batch_index=idx, total_batches=total_batches, name_prefix=name_prefix)
        items = [build_h3_controller_item(row) for row in chunk]
        payload = {
            "ActionName": action_name,
            "Controller": controller,
            "AppCode": app_code,
            "TaskName": task_name,
            "ItemsJson": json.dumps(items, ensure_ascii=False),
        }
        if text(created_by_name):
            payload["CreatedByName"] = text(created_by_name)
        if text(created_by_user_id):
            payload["CreatedByUserId"] = text(created_by_user_id)
        result = h3_request(
            url=url,
            engine_code=engine_code,
            engine_secret=engine_secret,
            payload=payload,
            timeout=timeout,
        )
        if not result.get("Successful", False):
            return result
        return_data = result.get("ReturnData", {}) or {}
        if not return_data.get("Success", False):
            return {
                "Successful": False,
                "ErrorMessage": text(return_data.get("Message")) or "自定义控制器创建草稿失败",
                "ReturnData": return_data,
            }
        created.append(return_data)
        names.append(task_name)
    return {
        "Successful": True,
        "ReturnData": created,
        "__biz_object_names": names,
        "__mode": "controller",
    }


def debug_h3_project_lookup(
    *,
    url: str,
    engine_code: str,
    engine_secret: str,
    app_code: str,
    controller: str,
    action_name: str,
    project_text: str,
    timeout: int,
) -> Dict[str, object]:
    payload = {
        "ActionName": action_name,
        "Controller": controller,
        "AppCode": app_code,
        "ProjectText": project_text,
    }
    result = h3_request(
        url=url,
        engine_code=engine_code,
        engine_secret=engine_secret,
        payload=payload,
        timeout=timeout,
    )
    if not result.get("Successful", False):
        return result
    return_data = result.get("ReturnData", {}) or {}
    if not return_data.get("Success", False):
        return {
            "Successful": False,
            "ErrorMessage": text(return_data.get("Message")) or "项目调试查询失败",
            "ReturnData": return_data,
        }
    return {"Successful": True, "ReturnData": return_data}


def load_h3_biz_object(
    *,
    url: str,
    engine_code: str,
    engine_secret: str,
    schema_code: str,
    biz_object_id: str,
    timeout: int,
) -> Dict[str, object]:
    payload = {
        "ActionName": "LoadBizObject",
        "SchemaCode": schema_code,
        "BizObjectId": biz_object_id,
    }
    result = h3_request(
        url=url,
        engine_code=engine_code,
        engine_secret=engine_secret,
        payload=payload,
        timeout=timeout,
    )
    if not result.get("Successful", False):
        raise RuntimeError(f"氚云回查失败: {result.get('ErrorMessage', '未知错误')}")
    return (result.get("ReturnData", {}) or {}).get("BizObject", {}) or {}


def verify_h3_object_statuses(
    *,
    url: str,
    engine_code: str,
    engine_secret: str,
    schema_code: str,
    object_ids: List[str],
    timeout: int,
) -> List[Dict[str, object]]:
    statuses: List[Dict[str, object]] = []
    for object_id in object_ids:
        biz_object = load_h3_biz_object(
            url=url,
            engine_code=engine_code,
            engine_secret=engine_secret,
            schema_code=schema_code,
            biz_object_id=object_id,
            timeout=timeout,
        )
        statuses.append(
            {
                "ObjectId": object_id,
                "Name": text(biz_object.get("Name")),
                "Status": biz_object.get("Status"),
            }
        )
    return statuses


def chunk_rows(rows: List[Dict[str, str]], chunk_size: int) -> List[List[Dict[str, str]]]:
    return [rows[i:i + chunk_size] for i in range(0, len(rows), chunk_size)]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="转换编码核对 Excel/CSV 为新表结构，并可选择是否按分类打散")
    parser.add_argument("input_excel", type=Path, help="输入 Excel/CSV 路径")
    parser.add_argument("--chunk-size", type=int, default=200, help="每个输出 Excel 最大条数，默认 200")
    parser.add_argument("--seed", type=int, default=42, help="打散随机种子")
    parser.add_argument("--upload-h3", action="store_true", help="将拆分后的数据直接上传到氚云")
    parser.add_argument("--h3-upload-mode", choices=["controller", "openapi"], default="controller", help="氚云上传模式：controller=自定义控制器草稿接口，openapi=直接调用CreateBizObjects")
    parser.add_argument("--h3-url", default=os.getenv("H3YUN_URL", H3_DEFAULT_URL), help="氚云 OpenApi 地址")
    parser.add_argument("--h3-engine-code", default=os.getenv("H3YUN_ENGINE_CODE", "ety58sf4upb95mibri9qatvi5"), help="氚云 EngineCode")
    parser.add_argument("--h3-engine-secret", default=os.getenv("H3YUN_ENGINE_SECRET", "u3toDgywMDwgbYgnrHNmjzH0g0fzn9mWAj0PY659taS7sxeVPoor5g=="), help="氚云 EngineSecret")
    parser.add_argument("--h3-schema-code", default=H3_REVIEW_SCHEMA_CODE, help="氚云主表 SchemaCode")
    parser.add_argument("--h3-app-code", default=H3_REVIEW_APP_CODE, help="氚云自定义控制器 AppCode")
    parser.add_argument("--h3-controller", default=H3_REVIEW_CONTROLLER, help="氚云自定义控制器名称")
    parser.add_argument("--h3-controller-action", default=H3_REVIEW_CREATE_DRAFT_ACTION, help="氚云自定义控制器创建草稿动作名")
    parser.add_argument("--h3-controller-debug-action", default=H3_REVIEW_DEBUG_PROJECT_ACTION, help="氚云自定义控制器项目调试动作名")
    parser.add_argument("--h3-timeout", type=int, default=120, help="氚云请求超时秒数")
    parser.add_argument("--h3-name-prefix", default="AI编码核对暂存-", help="氚云主表名称前缀")
    parser.add_argument("--h3-created-by-name", default=os.getenv("H3YUN_CREATED_BY_NAME", ""), help="氚云创建人姓名；controller 模式下可按启用且在职员工解析")
    parser.add_argument("--h3-created-by-id", default=os.getenv("H3YUN_CREATED_BY_ID", ""), help="氚云创建人用户ID；openapi 模式仅支持通过该参数指定创建人")
    parser.add_argument("--h3-submit", action="store_true", help="仅在 openapi 模式下写入生效数据；默认暂存/草稿")
    parser.add_argument("--skip-h3-verify", action="store_true", help="仅在 openapi 模式下上传后不回查氚云主表 Status")
    shuffle_group = parser.add_mutually_exclusive_group()
    shuffle_group.add_argument(
        "--shuffle",
        dest="shuffle",
        action="store_true",
        help="按分类尽量均衡打散后再拆分（默认）",
    )
    shuffle_group.add_argument(
        "--no-shuffle",
        dest="shuffle",
        action="store_false",
        help="保持原始顺序拆分，不打乱",
    )
    parser.set_defaults(shuffle=True)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path: Path = args.input_excel
    if not input_path.exists():
        raise FileNotFoundError(f"输入文件不存在: {input_path}")

    raw_rows = read_rows(input_path)
    project_standard_bias = build_project_standard_bias(raw_rows)
    mapped_rows = [build_output_row(row, project_standard_bias) for row in raw_rows]
    output_rows = balanced_shuffle(mapped_rows, seed=args.seed) if args.shuffle else mapped_rows

    base_name = input_path.stem + "_转换后"
    out_dir = input_path.parent
    chunks = chunk_rows(output_rows, args.chunk_size)

    for idx, chunk in enumerate(chunks, 1):
        out_path = out_dir / f"{base_name}_part{idx}.xlsx"
        write_chunk(out_path, chunk)
        print(f"WROTE {out_path} rows={len(chunk)}")

    if args.upload_h3:
        if not args.h3_engine_code or not args.h3_engine_secret:
            raise ValueError("启用 --upload-h3 时，必须提供 --h3-engine-code 和 --h3-engine-secret，或设置环境变量 H3YUN_ENGINE_CODE / H3YUN_ENGINE_SECRET")
        if args.h3_upload_mode == "openapi" and args.h3_created_by_name and not args.h3_created_by_id:
            raise ValueError("openapi 模式不支持按姓名解析创建人，请改用 --h3-created-by-id")
        if args.h3_upload_mode == "controller":
            h3_result = upload_chunks_to_h3_controller(
                chunks,
                url=args.h3_url,
                engine_code=args.h3_engine_code,
                engine_secret=args.h3_engine_secret,
                app_code=args.h3_app_code,
                controller=args.h3_controller,
                action_name=args.h3_controller_action,
                timeout=args.h3_timeout,
                name_prefix=args.h3_name_prefix,
                created_by_name=args.h3_created_by_name,
                created_by_user_id=args.h3_created_by_id,
            )
            if not h3_result.get("Successful", False):
                raise RuntimeError(f"氚云控制器草稿创建失败: {h3_result.get('ErrorMessage', '未知错误')}")
            return_data = h3_result.get("ReturnData", []) or []
            for info in return_data:
                unresolved = info.get("unresolvedProjects") or []
                dedup_unresolved = []
                seen_unresolved = set()
                for item in unresolved:
                    val = text(item)
                    if val and val not in seen_unresolved:
                        seen_unresolved.add(val)
                        dedup_unresolved.append(val)
                print(
                    "H3_CONTROLLER_DRAFT"
                    f" object={text(info.get('objectId'))}"
                    f" name={text(info.get('name'))}"
                    f" status={text(info.get('status')) or '0'}"
                    f" itemCount={text(info.get('itemCount'))}"
                    f" createdBy={text(info.get('createdByUserId')) or '-'}"
                )
                if dedup_unresolved:
                    print("H3_UNRESOLVED_PROJECTS=" + " | ".join(dedup_unresolved))
                    for project_name in dedup_unresolved:
                        debug_result = debug_h3_project_lookup(
                            url=args.h3_url,
                            engine_code=args.h3_engine_code,
                            engine_secret=args.h3_engine_secret,
                            app_code=args.h3_app_code,
                            controller=args.h3_controller,
                            action_name=args.h3_controller_debug_action,
                            project_text=project_name,
                            timeout=args.h3_timeout,
                        )
                        if not debug_result.get("Successful", False):
                            print(
                                f"H3_PROJECT_DEBUG_FAILED project={project_name} "
                                f"reason={text(debug_result.get('ErrorMessage')) or '未知错误'}"
                            )
                            continue
                        debug_data = debug_result.get("ReturnData", {}) or {}
                        print(
                            f"H3_PROJECT_DEBUG project={project_name} "
                            f"resolved={text(debug_data.get('ResolvedObjectId')) or '-'} "
                            f"count={text(debug_data.get('CandidateCount')) or '0'}"
                        )
                        candidates_text = text(debug_data.get("CandidatesText"))
                        if candidates_text:
                            for line in candidates_text.splitlines():
                                if text(line):
                                    print("H3_PROJECT_CANDIDATE " + line.strip())
            print(
                f"H3_UPLOAD=OK mode=controller objects={len(return_data)} draft=ON "
                f"controller={args.h3_controller} action={args.h3_controller_action}"
            )
        else:
            h3_result = upload_chunks_to_h3(
                chunks,
                url=args.h3_url,
                engine_code=args.h3_engine_code,
                engine_secret=args.h3_engine_secret,
                schema_code=args.h3_schema_code,
                is_submit=args.h3_submit,
                timeout=args.h3_timeout,
                name_prefix=args.h3_name_prefix,
                created_by_user_id=args.h3_created_by_id,
            )
            if not h3_result.get("Successful", False):
                raise RuntimeError(f"氚云暂存失败: {h3_result.get('ErrorMessage', '未知错误')}")
            return_data = h3_result.get("ReturnData", []) or []
            if not args.skip_h3_verify and return_data:
                try:
                    statuses = verify_h3_object_statuses(
                        url=args.h3_url,
                        engine_code=args.h3_engine_code,
                        engine_secret=args.h3_engine_secret,
                        schema_code=args.h3_schema_code,
                        object_ids=[text(v) for v in return_data if text(v)],
                        timeout=args.h3_timeout,
                    )
                    for info in statuses:
                        print(f"H3_STATUS object={info['ObjectId']} name={info['Name']} status={info['Status']}")
                    if not args.h3_submit:
                        non_draft = [info for info in statuses if info.get("Status") != 0]
                        if non_draft:
                            raise RuntimeError(
                                "氚云返回的数据不是草稿状态(Status=0)。"
                                f" 实际状态: {[info.get('Status') for info in non_draft]}"
                            )
                except Exception as exc:
                    names = h3_result.get("__biz_object_names", []) or []
                    print(f"H3_VERIFY=SKIPPED reason={exc}")
                    if names:
                        print("H3_CREATED_NAMES=" + " | ".join(str(name) for name in names if text(name)))
            print(f"H3_UPLOAD=OK mode=openapi objects={len(return_data)} draft={'OFF' if args.h3_submit else 'ON'} schema={args.h3_schema_code}")

    print(f"SHUFFLE={'ON' if args.shuffle else 'OFF'}")
    print(f"TOTAL_ROWS={len(output_rows)}")
    print(f"TOTAL_FILES={len(chunks)}")


if __name__ == "__main__":
    main()
