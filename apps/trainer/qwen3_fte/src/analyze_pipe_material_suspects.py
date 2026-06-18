#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Analyze suspicious MATERIAL labels in pipe draft dataset under raw-text-first rules."""

from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[4]
BASE = ROOT / "apps/trainer/qwen3_fte/output/pipe_project_sampling_full"
DATASET = BASE / "直管训练草稿_语义补强版.json"
OUT_XLSX = BASE / "直管材质可疑样本分析_改进版.xlsx"


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def as_list(value: Any) -> list[Any]:
    if value in (None, ""):
        return []
    return value if isinstance(value, list) else [value]


def material_items(item: dict[str, Any]) -> list[dict[str, Any]]:
    mats = item.get("output", {}).get("MATERIAL", []) or []
    out: list[dict[str, Any]] = []
    for m in mats:
        if isinstance(m, dict):
            out.append(m)
        elif text(m):
            out.append({"ROLE": "MAIN", "VALUE": text(m), "SPECIAL_REQ": []})
    return out


def normalize_special_req_from_text(desc: str) -> list[str]:
    desc_u = desc.upper()
    reqs: list[str] = []
    if any(token in desc_u for token in ["NACE", "ANTI-H2S", "H2S"]):
        reqs.append("NACE")
    if any(token in desc_u for token in ["GALV", "GALVANIZED", "镀锌"]):
        reqs.append("GALV")
    seen = set()
    out = []
    for r in reqs:
        if r not in seen:
            seen.add(r)
            out.append(r)
    return out


def pick_first(desc: str, patterns: list[tuple[str, str]]) -> str:
    for _, pattern in patterns:
        m = re.search(pattern, desc, re.I)
        if m:
            return m.group(1).strip()
    return ""


def extract_api5l_material(desc: str) -> str:
    desc_u = desc.upper()
    if "API 5L" not in desc_u and "API5L" not in desc_u:
        return ""
    m = re.search(r"API\s*5L\s*(?:GRADE\b|GR\.?)\s*([A-Z0-9.]+)", desc, re.I)
    if m:
        grade = m.group(1).upper().rstrip(".")
        if grade:
            return f"API 5L Gr.{grade}"
    m = re.search(r"API\s*5L\s*PSL[12]\s*GR\.?\s*([A-Z0-9.]+)", desc, re.I)
    if m:
        grade = m.group(1).upper().rstrip(".")
        if grade:
            return f"API 5L Gr.{grade}"
    m = re.search(r"API\s*5L[^A-Z0-9]+GR\.?\s*([A-Z0-9.]+)", desc, re.I)
    if m:
        grade = m.group(1).upper().rstrip(".")
        if grade:
            return f"API 5L Gr.{grade}"
    m = re.search(r"API\s*5L\s*([A-Z]\d{2,3})(?!\w)", desc, re.I)
    if m:
        return f"API 5L {m.group(1).upper()}"
    m = re.search(r"PSL[12]\s*GR\.?\s*([A-Z0-9.]+)", desc, re.I)
    if m:
        return f"API 5L Gr.{m.group(1).upper().rstrip('.')}"
    m = re.search(r"GR\.?\s*PSL[12]\b", desc, re.I)
    if m:
        m2 = re.search(r"API\s*5L\s*([A-Z]\d{2,3})", desc, re.I)
        if m2:
            return f"API 5L {m2.group(1).upper()}"
    return "API 5L"


def extract_astm_material(desc: str) -> str:
    patterns = [
        ("A312", r"\b(A312(?:M)?\s*(?:GR\.?|GRADE\s*)?TP[0-9A-Z/.-]+)\b"),
        ("A358", r"\b(A358(?:M)?\s*(?:GR\.?|GRADE\s*)?[0-9A-Z./-]+)\b"),
        ("A333", r"\b(A333(?:M)?\s*(?:GR\.?|GRADE\s*)?[0-9A-Z./-]+)\b"),
        ("A335", r"\b(A335(?:M)?\s*(?:GR\.?|GRADE\s*)?P[0-9][0-9A-Z]*)\b"),
        ("A106", r"\b(A106(?:-|\s)*(?:GR\.?|GRADE\s*)?[A-Z0-9.]+)\b"),
        ("A790", r"\b(A790(?:M)?\s*S[0-9A-Z]+)\b"),
        ("A672", r"\b(A672\s*[A-Z0-9.\s]+)\b"),
    ]
    raw = pick_first(desc, patterns)
    if not raw:
        return ""
    raw = re.sub(r"\s+", " ", raw).strip().replace("GRADE", "Gr.").replace("GR.", "Gr.")
    raw = re.sub(r"\bA(\d{3,4})(M?)\b", r"ASTM A\1\2", raw)
    raw = raw.replace("ASTM ASTM", "ASTM")
    raw = raw.replace("Gr..", "Gr.")
    raw = re.sub(r"\bGr\.\s*", "Gr.", raw)
    raw = raw.replace("A106-Gr.", "ASTM A106 Gr.")
    raw = raw.replace("A335 Gr.", "ASTM A335 Gr.")
    raw = raw.replace("A335 P", "ASTM A335 P")
    raw = raw.replace("ASTM ASTM", "ASTM")
    return raw


def extract_raw_material_value(desc: str) -> str:
    api5l = extract_api5l_material(desc)
    if api5l and api5l != "API 5L":
        return api5l

    astm = extract_astm_material(desc)
    if astm:
        return astm

    exact_patterns = [
        r"\b(S30403|S30408|S31603|S32168|S22053|S31008)\b",
        r"\b(06Cr19Ni10|06Cr18Ni11Ti|022Cr17Ni12Mo2|022Cr25Ni7Mo4N|15CrMoG|15CrMo|12Cr1MoV|20G|Q235B|Q345E|Q345R|L245|TA2|TA10|高硅不锈钢)\b",
        r"\b(20#|20)\b",
        r"\b(304|304L|316L|2205|321|310s|310S)\b",
        r"\b(ZECOR-310M)\b",
        r"\b(X6CrNiMoTi17-12-2\(1\.4571\)|X6CrNiTi18-10\(1\.4541\))\b",
    ]
    for pattern in exact_patterns:
        m = re.search(pattern, desc, re.I)
        if m:
            return m.group(1)

    return ""


def analyze_one(idx: int, item: dict[str, Any]) -> list[list[Any]]:
    inp = text(item.get("input"))
    rows: list[list[Any]] = []
    for mat_idx, mat in enumerate(material_items(item), start=1):
        value = text(mat.get("VALUE"))
        current_special = ",".join(text(x) for x in as_list(mat.get("SPECIAL_REQ")) if text(x))
        coating = mat.get("COATING") if isinstance(mat, dict) else {}
        current_inner = [text(x) for x in as_list((coating or {}).get("INNER")) if text(x)]
        current_outer = [text(x) for x in as_list((coating or {}).get("OUTER")) if text(x)]
        suggested_value = ""
        suggested_special = ",".join(normalize_special_req_from_text(inp))
        issue = ""
        note = ""

        if value.startswith("T20553"):
            issue = "HG/T20553误标为材质"
            suggested_value = extract_raw_material_value(inp)
            note = "HG/T20553 是规范，不是材质"
        elif value.startswith("T12771"):
            issue = "GB/T12771片段误标为材质"
            suggested_value = extract_raw_material_value(inp)
            note = "GB/T12771 是规范，不是材质"
        elif value in {"304", "304L", "316L", "321", "2205", "2507", "310M"}:
            if value == "304L" and "S30403" in inp:
                issue = "S30403被归一化成304L"
                suggested_value = "S30403"
            elif value == "304" and "S30408" in inp:
                issue = "S30408被归一化成304"
                suggested_value = "S30408"
            elif value == "316L" and "S31603" in inp:
                issue = "S31603被归一化成316L"
                suggested_value = "S31603"
            elif value == "2205" and "S22053" in inp:
                issue = "S22053被归一化成2205"
                suggested_value = "S22053"
            elif value == "321" and re.search(r"S32168|06Cr18Ni11Ti", inp):
                issue = "321系材质被归一化"
                suggested_value = extract_raw_material_value(inp)
            elif value == "304" and "06Cr19Ni10" in inp:
                issue = "06Cr19Ni10被归一化成304"
                suggested_value = "06Cr19Ni10"
            elif value == "316L" and "022Cr17Ni12Mo2" in inp:
                issue = "022Cr17Ni12Mo2被归一化成316L"
                suggested_value = "022Cr17Ni12Mo2"
            elif value == "2507" and "022Cr25Ni7Mo4N" in inp:
                issue = "022Cr25Ni7Mo4N被归一化成2507"
                suggested_value = "022Cr25Ni7Mo4N"
            elif value == "310M" and "ZECOR-310M" in inp:
                issue = "ZECOR-310M被截断"
                suggested_value = "ZECOR-310M"
        elif value in {"P11", "P22", "P91", "A333", "API5L"} or value.endswith("(NACE)"):
            raw = extract_raw_material_value(inp)
            if raw:
                issue = "材质前缀或等级丢失"
                suggested_value = raw
        elif value in {"20", "304", "Q235B", "PE"} and re.search(
            r"20/PTFE|20\+PTFE|20#\+PTFE|20/PE|20\+EAA|20#\+EAA|304/PTFE|PTFE/20|FRP/PVC|FRP/CPVC|Q235B外加强级PE内EP|GLASS LINED|STEEL REINFORCED POLYETHYLENE",
            inp,
            re.I,
        ):
            composite_ok = False
            inp_u = inp.upper()
            current_value_u = value.upper()
            if re.search(r"304/PTFE|20/PTFE|20\+PTFE|20#\+PTFE|PTFE/20", inp_u):
                composite_ok = "/PTFE" in current_value_u or current_value_u.endswith("PTFE")
            elif re.search(r"20/PE", inp_u):
                composite_ok = "/PE" in current_value_u or current_value_u.endswith("PE")
            elif re.search(r"20\+EAA|20#\+EAA", inp_u):
                composite_ok = "/EAA" in current_value_u or current_value_u.endswith("EAA")
            elif "Q235B外加强级PE内EP".upper() in inp_u:
                composite_ok = (
                    "EP" in [x.upper() for x in current_inner]
                    and any("PE" in x.upper() for x in current_outer)
                )
            elif "GLASS LINED" in inp_u:
                composite_ok = "GLASS LINED" in current_value_u
            elif "STEEL REINFORCED POLYETHYLENE" in inp_u:
                composite_ok = value.upper() in {"PE", "STEEL REINFORCED POLYETHYLENE"}
            elif re.search(r"FRP/PVC|FRP/CPVC", inp_u):
                composite_ok = value.upper() in {"FRP/PVC", "FRP/CPVC"}

            if not composite_ok:
                issue = "复合/衬里材质信息丢失"
                suggested_value = extract_raw_material_value(inp)
                note = "主材之外仍有复合/衬里信息"

        if issue:
            rows.append(
                [
                    idx,
                    mat_idx,
                    issue,
                    value,
                    current_special,
                    suggested_value,
                    suggested_special,
                    inp,
                    note,
                ]
            )
    return rows


def main() -> None:
    with DATASET.open("r", encoding="utf-8") as f:
        data = json.load(f)

    rows: list[list[Any]] = []
    for idx, item in enumerate(data, start=1):
        rows.extend(analyze_one(idx, item))

    wb = Workbook()
    ws = wb.active
    ws.title = "可疑材质"
    headers = [
        "记录序号",
        "材质项序号",
        "问题类型",
        "当前VALUE",
        "当前SPECIAL_REQ",
        "建议VALUE",
        "建议SPECIAL_REQ",
        "原文描述",
        "说明",
    ]
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    for row in rows:
        ws.append(row)
    widths = {
        "A": 10,
        "B": 10,
        "C": 28,
        "D": 20,
        "E": 18,
        "F": 24,
        "G": 18,
        "H": 120,
        "I": 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("汇总")
    ws2.append(["问题类型", "数量"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    for name, count in Counter(row[2] for row in rows).most_common():
        ws2.append([name, count])
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 10

    wb.save(OUT_XLSX)
    print(OUT_XLSX)
    print("rows", len(rows))
    for name, count in Counter(row[2] for row in rows).most_common():
        print(name, count)


if __name__ == "__main__":
    main()
