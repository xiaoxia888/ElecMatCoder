#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
"""
python apps/trainer/qwen3_fte/src/将csv转换为规范材质训练集.py \
    --input /Users/guoxi/Desktop/workspace/NJNCC/python_code/review_platform/materials_export.csv \
    --output /Users/guoxi/Desktop/workspace/NJNCC/python_code/ElecMatCoder/apps/trainer/qwen3_fte/output/按8类拆分数据集/output.json \
    --audit-xlsx /Users/guoxi/Desktop/workspace/NJNCC/python_code/ElecMatCoder/apps/trainer/qwen3_fte/output/按8类拆分数据集/audit.xlsx
"""

PROJECT_ROOT = Path(__file__).resolve().parents[4]
DEFAULT_INPUT = PROJECT_ROOT.parent / "review_platform" / "materials_export.csv"
DEFAULT_OUTPUT = (
    PROJECT_ROOT
    / "apps"
    / "trainer"
    / "qwen3_fte"
    / "output"
    / "按8类拆分数据集"
    / "材质规范_from_materials_export.json"
)
DEFAULT_AUDIT_XLSX = (
    PROJECT_ROOT
    / "apps"
    / "trainer"
    / "qwen3_fte"
    / "output"
    / "按8类拆分数据集"
    / "材质规范_from_materials_export_audit.xlsx"
)

# 输入列配置。后续如果表头变了，优先在这里改，不要到函数里逐个找。
INPUT_DESC_COLUMNS = ("材料描述(多行)", "材料描述")
INPUT_MATERIAL_COLUMNS = ("修正材质", "标准化材质")
INPUT_STANDARD_COLUMNS = ("修正规范", "标准化规范")

# 审计导出里展示“原始命中列”时使用的列顺序。
AUDIT_RAW_MATERIAL_COLUMNS = INPUT_MATERIAL_COLUMNS
AUDIT_RAW_STANDARD_COLUMNS = INPUT_STANDARD_COLUMNS

EXCLUDED_DESC_KEYWORDS = ("阀门", "垫片")
EXCLUDED_MATERIAL_KEYWORD = "TRIM"


STANDARD_FAMILY_PREFIXES = (
    "MSSSP",
    "ASTM",
    "API",
    "ASMC",
    "ANM",
    "GBT",
    "HGT",
    "NBT",
    "SHT",
    "SYT",
    "DLT",
    "ENI",
    "DIN",
    "CJT",
    "TB",
    "MC",
    "AB",
    "MS",
    "EN",
    "GB",
    "NB",
    "SH",
    "HG",
    "GD",
)

STANDARD_FAMILY_PREFIXES_SORTED = tuple(sorted(STANDARD_FAMILY_PREFIXES, key=len, reverse=True))

# 常见规范误写纠错。这里做的是“已知高置信误写”的归一化，不在拆分逻辑里猜测。
STANDARD_CORRECTIONS = {
    "GB13401": "GBT13401",
    "GB12459I": "GBT12459I",
    "GB12459II": "GBT12459II",
    "GB12459": "GBT12459",
    "NB47010": "NBT47010",
    "SH3408": "SHT3408",
    "SH3419": "SHT3419",
    "HG20553": "HGT20553",
    "GBT13401LA":"GBT13401IA",
    "GBT13401I064":"GBT13401",
    "NBT47009GR":"NBT47009",
    "AB36.10":"AB3610"
}


def text(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def first_nonempty(row: dict[str, str], columns: tuple[str, ...]) -> str:
    for column in columns:
        value = text(row.get(column))
        if value:
            return value
    return ""


def uniq_keep_order(values: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        if not value or value in seen:
            continue
        seen.add(value)
        out.append(value)
    return out


def normalize_standard_code(standard_code: str) -> str:
    raw = text(standard_code).replace(" ", "").replace("（", "(").replace("）", ")")
    # 清掉人工备注类尾注，不参与规范解析。
    raw = re.sub(r"\(更改修正编码后取消\)", "", raw, flags=re.IGNORECASE)
    raw = re.sub(r"\(修正编码后取消\)", "", raw, flags=re.IGNORECASE)
    raw = re.sub(r"\(更改后取消\)", "", raw, flags=re.IGNORECASE)
    # 常见录入/OCR 误差归一化。匹配时大小写不敏感，输出使用规范编码。
    raw = re.sub(r"HGT20553LA", "HGT20553IA", raw, flags=re.IGNORECASE)
    raw = re.sub(r"GBT20553LA", "GBT20553IA", raw, flags=re.IGNORECASE)
    for wrong, corrected in STANDARD_CORRECTIONS.items():
        raw = re.sub(re.escape(wrong), corrected, raw, flags=re.IGNORECASE)
    return raw


def build_standard(standard_code: str) -> list[dict[str, str]]:
    raw = normalize_standard_code(standard_code)
    if not raw:
        return []
    scan_raw = raw.upper()
    pos = 0
    tokens: list[str] = []

    while pos < len(scan_raw):
        standalone_i = re.match(r"I\d+(?:\.\d+)?", scan_raw[pos:])
        if standalone_i:
            token = standalone_i.group(0)
            tokens.append(token)
            pos += len(token)
            continue

        family = next((p for p in STANDARD_FAMILY_PREFIXES_SORTED if scan_raw.startswith(p, pos)), None)
        if not family:
            return []
        next_pos = len(scan_raw)
        search_start = pos + len(family)
        for j in range(search_start, len(scan_raw)):
            if any(scan_raw.startswith(p, j) for p in STANDARD_FAMILY_PREFIXES_SORTED):
                next_pos = j
                break
        token = scan_raw[pos:next_pos]
        # 至少要求 family 之后有数字，避免把纯字母噪声当标准。
        if not re.search(r"\d", token[len(family):]):
            return []
        tokens.append(token)
        pos = next_pos

    if "".join(tokens) != scan_raw:
        return []

    return [{"BODY": token} for token in uniq_keep_order(tokens)]


def build_material(material_value: str) -> list[dict[str, Any]]:
    value = text(material_value)
    if not value:
        return []
    return [{"ROLE": "MAIN", "VALUE": value}]


def normalize_input(desc: str) -> str:
    return re.sub(r"\s+", " ", text(desc)).strip()


def choose_description(row: dict[str, str]) -> str:
    return first_nonempty(row, INPUT_DESC_COLUMNS)


def choose_material(row: dict[str, str]) -> str:
    return first_nonempty(row, INPUT_MATERIAL_COLUMNS)


def choose_standard(row: dict[str, str]) -> str:
    return first_nonempty(row, INPUT_STANDARD_COLUMNS)


def detect_exclusion_reason(desc: str, chosen_material: str) -> str:
    desc_text = text(desc)
    for keyword in EXCLUDED_DESC_KEYWORDS:
        if keyword and keyword in desc_text:
            return f"描述包含{keyword}"

    material_text = text(chosen_material).upper()
    if EXCLUDED_MATERIAL_KEYWORD in material_text:
        return f"材质包含{EXCLUDED_MATERIAL_KEYWORD}"

    return ""


def build_output(row: dict[str, str]) -> dict[str, Any]:
    output: dict[str, Any] = {}
    material = build_material(choose_material(row))
    standard = build_standard(choose_standard(row))
    if standard:
        output["STANDARD"] = standard
    if material:
        output["MATERIAL"] = material
    return output


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return [{str(k): text(v) for k, v in row.items()} for row in reader]


def read_excel_rows(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []
    headers = [text(v) for v in rows[0]]
    out: list[dict[str, str]] = []
    for raw_row in rows[1:]:
        row = {
            headers[idx]: text(raw_row[idx]) if idx < len(raw_row) else ""
            for idx in range(len(headers))
            if headers[idx]
        }
        out.append(row)
    wb.close()
    return out


def read_input_rows(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_rows(path)
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return read_excel_rows(path)
    raise ValueError(f"不支持的输入文件类型: {path.suffix}")


def read_existing_json_dataset(path: Path) -> list[dict[str, Any]]:
    raw = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(raw, list):
        raise ValueError("merge-json 输入必须是对象数组")

    out: list[dict[str, Any]] = []
    for item in raw:
        if not isinstance(item, dict):
            continue
        input_text = text(item.get("input"))
        output = item.get("output")
        if not input_text or not isinstance(output, dict):
            continue
        out.append({"input": input_text, "output": output})
    return out


def summarize_standard(standard_items: list[dict[str, str]]) -> str:
    return " | ".join(item.get("BODY", "") for item in standard_items if item.get("BODY"))


def summarize_material(material_items: list[dict[str, Any]]) -> str:
    parts: list[str] = []
    for item in material_items:
        role = text(item.get("ROLE"))
        value = text(item.get("VALUE"))
        if not value:
            continue
        parts.append(f"{role}:{value}" if role else value)
    return " | ".join(parts)


def write_audit_xlsx(path: Path, rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "audit"
    audit_headers = [
        "原始描述",
        "裁剪后输入",
        "原始材质列",
        "选用材质列",
        "解析后材质",
        "原始规范列",
        "选用规范列",
        "标准规范化后",
        "解析后规范",
        "输出JSON",
        "状态",
        "原因",
        "去重键",
    ]
    ws.append(audit_headers)
    for row in rows:
        ws.append([row.get(col, "") for col in audit_headers])

    material_ws = wb.create_sheet("材质汇总")
    material_headers = [
        "选用材质列",
        "解析后材质",
        "总次数",
        "保留次数",
        "跳过次数",
        "常见原因",
    ]
    material_ws.append(material_headers)
    material_summary: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        key = (row.get("选用材质列", ""), row.get("解析后材质", ""))
        bucket = material_summary.setdefault(
            key,
            {"total": 0, "kept": 0, "skipped": 0, "reasons": Counter()},
        )
        bucket["total"] += 1
        if row.get("状态") == "保留":
            bucket["kept"] += 1
        else:
            bucket["skipped"] += 1
            reason = row.get("原因", "")
            if reason:
                bucket["reasons"][reason] += 1
    for (chosen_material, parsed_material), bucket in sorted(
        material_summary.items(),
        key=lambda item: (-item[1]["total"], item[0][0], item[0][1]),
    ):
        top_reasons = " | ".join(
            f"{reason}:{count}" for reason, count in bucket["reasons"].most_common(3)
        )
        material_ws.append([
            chosen_material,
            parsed_material,
            bucket["total"],
            bucket["kept"],
            bucket["skipped"],
            top_reasons,
        ])

    standard_ws = wb.create_sheet("规范汇总")
    standard_headers = [
        "解析后规范",
        "总次数",
        "保留次数",
        "跳过次数",
        "常见原因",
    ]
    standard_ws.append(standard_headers)
    standard_summary: dict[str, dict[str, Any]] = {}
    for row in rows:
        parsed_values = [
            value.strip()
            for value in row.get("解析后规范", "").split("|")
            if value.strip()
        ]
        if not parsed_values:
            parsed_values = [""]
        for parsed_standard in parsed_values:
            bucket = standard_summary.setdefault(
                parsed_standard,
                {"total": 0, "kept": 0, "skipped": 0, "reasons": Counter()},
            )
            bucket["total"] += 1
            if row.get("状态") == "保留":
                bucket["kept"] += 1
            else:
                bucket["skipped"] += 1
                reason = row.get("原因", "")
                if reason:
                    bucket["reasons"][reason] += 1
    for parsed_standard, bucket in sorted(
        standard_summary.items(),
        key=lambda item: (-item[1]["total"], item[0]),
    ):
        top_reasons = " | ".join(
            f"{reason}:{count}" for reason, count in bucket["reasons"].most_common(3)
        )
        standard_ws.append([
            parsed_standard,
            bucket["total"],
            bucket["kept"],
            bucket["skipped"],
            top_reasons,
        ])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="从 materials_export.csv 生成 MATERIAL/STANDARD 训练集")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument(
        "--merge-json",
        type=Path,
        default=None,
        help="额外读取一个已有 JSON 数据集，按 input 合并；同描述时 JSON 优先覆盖表格生成结果",
    )
    parser.add_argument(
        "--audit-xlsx",
        type=Path,
        default=DEFAULT_AUDIT_XLSX,
        help="导出审计Excel，包含原始值/裁剪后输入/解析后规范与材质/跳过原因",
    )
    args = parser.parse_args()

    dataset_map: dict[str, dict[str, Any]] = {}
    stats = Counter()
    audit_rows: list[dict[str, str]] = []

    for row in read_input_rows(args.input):
        stats["rows"] += 1
        desc = choose_description(row)
        chosen_material = choose_material(row)
        chosen_standard = choose_standard(row)
        exclusion_reason = detect_exclusion_reason(desc, chosen_material)
        normalized_standard = normalize_standard_code(chosen_standard)
        material_items = build_material(chosen_material)
        standard_items = build_standard(chosen_standard)
        output = {}
        if material_items:
            output["MATERIAL"] = material_items
        if standard_items:
            output["STANDARD"] = standard_items
        input_text = normalize_input(desc)
        reason = ""
        status = "保留"
        dedup_key = ""

        if not desc:
            stats["skip_no_desc"] += 1
            status = "跳过"
            reason = "无描述"
        elif exclusion_reason:
            stats["skip_excluded"] += 1
            status = "跳过"
            reason = exclusion_reason
        elif not output:
            stats["skip_no_output"] += 1
            status = "跳过"
            reason = "材质和规范都为空或规范解析失败"
        elif not input_text:
            stats["skip_empty_input"] += 1
            status = "跳过"
            reason = "描述裁剪后为空"
        else:
            item = {"input": input_text, "output": output}
            dedup_key = input_text
            if dedup_key in dataset_map:
                stats["dedup"] += 1
                status = "跳过"
                reason = "描述去重命中"
            else:
                dataset_map[dedup_key] = item
                if "MATERIAL" in output:
                    stats["with_material"] += 1
                if "STANDARD" in output:
                    stats["with_standard"] += 1

        audit_rows.append(
            {
                "原始描述": desc,
                "裁剪后输入": input_text,
                "原始材质列": first_nonempty(row, AUDIT_RAW_MATERIAL_COLUMNS),
                "选用材质列": chosen_material,
                "解析后材质": summarize_material(material_items),
                "原始规范列": first_nonempty(row, AUDIT_RAW_STANDARD_COLUMNS),
                "选用规范列": chosen_standard,
                "标准规范化后": normalized_standard,
                "解析后规范": summarize_standard(standard_items),
                "输出JSON": json.dumps(output, ensure_ascii=False),
                "状态": status,
                "原因": reason,
                "去重键": input_text,
            }
        )

    if args.merge_json:
        merge_items = read_existing_json_dataset(args.merge_json)
        stats["merge_json_rows"] = len(merge_items)
        for item in merge_items:
            input_text = text(item.get("input"))
            if not input_text:
                continue
            if input_text in dataset_map:
                stats["json_override"] += 1
            else:
                stats["json_append"] += 1
            dataset_map[input_text] = item

    dataset = list(dataset_map.values())

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(dataset, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_audit_xlsx(args.audit_xlsx, audit_rows)

    print(f"输入行数: {stats['rows']}")
    print(f"输出条数: {len(dataset)}")
    print(f"含 MATERIAL: {stats['with_material']}")
    print(f"含 STANDARD: {stats['with_standard']}")
    print(f"去重跳过: {stats['dedup']}")
    print(f"过滤跳过: {stats['skip_excluded']}")
    print(f"无输出跳过: {stats['skip_no_output']}")
    if args.merge_json:
        print(f"JSON 读取条数: {stats['merge_json_rows']}")
        print(f"JSON 覆盖条数: {stats['json_override']}")
        print(f"JSON 新增条数: {stats['json_append']}")
    print(f"输出文件: {args.output}")
    print(f"审计文件: {args.audit_xlsx}")


if __name__ == "__main__":
    main()
