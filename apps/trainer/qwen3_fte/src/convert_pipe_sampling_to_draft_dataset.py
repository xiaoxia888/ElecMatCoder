#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from collections import OrderedDict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[4]
DEFAULT_INPUT = PROJECT_ROOT / "apps" / "trainer" / "qwen3_fte" / "output" / "pipe_project_sampling" / "直管覆盖抽样.xlsx"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "apps" / "trainer" / "qwen3_fte" / "output" / "pipe_project_sampling"

SHEET_NAME = "模板抽样结果"
INPUT_COLUMNS = [
    "材料描述",
    "项目名称",
    "分类",
    "国标美标标记",
    "编码",
    "材料名称代码",
    "公称直径代码",
    "壁厚等级代码",
    "压力等级代码",
    "材质代码",
    "标准号代码",
    "材质分类",
]


def text(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def uniq_keep_order(values: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        if not value:
            continue
        if value in seen:
            continue
        seen.add(value)
        out.append(value)
    return out


def load_rows(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    rows = ws.iter_rows(values_only=True)
    header = [text(x) for x in next(rows)]
    idx = {name: i for i, name in enumerate(header)}
    missing = [name for name in INPUT_COLUMNS if name not in idx]
    if missing:
        raise KeyError(f"缺少列: {missing}")

    out: list[dict[str, str]] = []
    for row in rows:
        desc = text(row[idx["材料描述"]])
        if not desc:
            continue
        item = {name: text(row[idx[name]]) for name in INPUT_COLUMNS}
        item["Excel行号"] = text(row[idx["Excel行号"]]) if "Excel行号" in idx else ""
        item["抽样原因"] = text(row[idx["抽样原因"]]) if "抽样原因" in idx else ""
        item["类型编码"] = text(row[idx["类型编码"]]) if "类型编码" in idx else item["材料名称代码"]
        out.append(item)
    return out


def infer_body(type_code: str, desc: str) -> tuple[str, list[str]]:
    desc_u = desc.upper()
    notes: list[str] = []

    if "夹套钢管" in desc or "夹套管" in desc or "JACKETED PIPE" in desc_u or type_code == "PJ":
        return "夹套钢管", notes
    if "法兰管" in desc or "FLANGED PIPE" in desc_u or type_code in {"FP", "LFP"}:
        return "法兰管", notes
    return "直管", notes


def infer_manu(type_code: str, desc: str) -> list[str]:
    desc_u = desc.upper()
    values: list[str] = []

    if re.search(r"SEAMLESS\s+OR\s+WELDED|SMLS\s+OR\s+WELDED|无缝\s*或\s*焊接|无缝\s*/\s*焊接", desc_u):
        return ["WELDED"]

    for token in ["SAWL", "SAWH", "ERW", "EFW", "HFW", "LSAW"]:
        if token in desc_u:
            values.append(token)

    if "SMLS" in desc_u or "SEAMLESS" in desc_u or "无缝" in desc:
        values.insert(0, "SMLS")
    has_non_welded = "SMLS" in values
    welded_markers = ["WELDED", "焊接钢管", "焊管", "有缝"]
    if not has_non_welded and any(marker in desc_u or marker in desc for marker in welded_markers):
        values.append("WELDED")

    if type_code in {"PW", "PWM"} and not values:
        values.append("WELDED")

    return uniq_keep_order(values)


def infer_conn(desc: str) -> list[str]:
    desc_u = desc.upper()
    values: list[str] = []
    if re.search(r"(?<![A-Z0-9])SW(?![A-Z0-9])", desc_u):
        values.append("SW")
    if "TH(NPT)" in desc_u or re.search(r"(?<![A-Z0-9])NPT(?![A-Z0-9])", desc_u):
        # FNPT/MNPT/NPTF 属于端部形式，不进入 CONN。
        if not any(token in desc_u for token in ["FNPT", "MNPT", "NPTF"]):
            values.append("NPT")
    if "THD" in desc_u or "THREADED" in desc_u or "螺纹" in desc:
        values.append("THD")
    return uniq_keep_order(values)


def infer_ends(desc: str) -> list[str]:
    desc_u = desc.upper()
    values: list[str] = []
    for token in ["FNPT", "MNPT", "NPTF", "MTE", "TSE", "FTE", "BOE", "TOE"]:
        if token in desc_u:
            values.append(token)
    return uniq_keep_order(values)


def normalize_number_token(value: str) -> str:
    value = value.strip()
    if "." in value:
        value = value.rstrip("0").rstrip(".")
    return value


def build_size(size_code: str, desc: str) -> dict[str, list[str]]:
    dn_values: list[str] = []
    for token in re.findall(r"\d+(?:\.\d+)?", size_code):
        dn_values.append(f"DN{normalize_number_token(token)}")

    od_values: list[str] = []
    # φ108x5.5 / Φ114.3×6.3 这类明确是外径。
    for match in re.finditer(r"[Φφ]\s*(\d+(?:\.\d+)?)\s*[xX×*]\s*\d+(?:\.\d+)?", desc):
        od_values.append(normalize_number_token(match.group(1)))
    # 323.9x9.0 这类无 DN/Φ 前缀且第一段带小数，按 OD 处理。
    for match in re.finditer(r"(?<!DN)(?<!NPS)(?<!NB)\b(\d+\.\d+)\s*[xX×*]\s*\d+(?:\.\d+)?", desc, re.IGNORECASE):
        od_values.append(normalize_number_token(match.group(1)))

    inch_values: list[str] = []
    for match in re.finditer(r'(\d+(?:\s+\d+/\d+)?|\d+/\d+)\s*"{1,2}', desc):
        inch_values.append(match.group(1).replace(" ", "") + '"')

    length_values: list[str] = []
    for match in re.finditer(r"\bL\s*=\s*(\d+(?:\.\d+)?)\s*MM\b", desc, re.IGNORECASE):
        length_values.append(normalize_number_token(match.group(1)))
    for match in re.finditer(r"长度[:：]?\s*(\d+(?:\.\d+)?)(?:\s*(?:~|-|至)\s*(\d+(?:\.\d+)?))?\s*MM", desc, re.IGNORECASE):
        length_values.append(normalize_number_token(match.group(2) or match.group(1)))
    # 法兰管/衬里管描述中常见末尾 1200mm 作为长度，过滤掉小壁厚 mm。
    if any(token in desc for token in ["法兰管", "衬", "搪玻璃", "长度"]) or "LINED" in desc.upper():
        for match in re.finditer(r"(?<![xX×*=])\b(\d+(?:\.\d+)?)\s*MM\b", desc, re.IGNORECASE):
            num = float(match.group(1))
            if num >= 100:
                length_values.append(normalize_number_token(match.group(1)))

    return {
        "DN": uniq_keep_order(dn_values),
        "OD": uniq_keep_order(od_values),
        "INCH": uniq_keep_order(inch_values),
        "LENGTH": uniq_keep_order(length_values),
    }


def build_thickness(thk_code: str, desc: str) -> dict[str, list[str]]:
    mm_values: list[str] = []
    schedule_values: list[str] = []
    series_values: list[str] = []

    code = thk_code.upper().replace(" ", "")
    if code:
        for part in re.split(r"[xX×]", code):
            part = part.strip()
            if not part:
                continue
            if part.endswith("MM"):
                mm_values.append(normalize_number_token(part[:-2]))
                continue
            part = part.replace("SCH", "").lstrip("S-")
            if part in {"XS", "XXS", "STD"}:
                series_values.append(part)
            elif re.fullmatch(r"\d+(?:\.\d+)?S?", part):
                schedule_values.append(f"SCH{part}")

    if not mm_values:
        for match in re.finditer(r"(?:THK\s*=\s*)?(\d+(?:\.\d+)?)\s*MM\b", desc, re.IGNORECASE):
            num = float(match.group(1))
            if num < 100:
                mm_values.append(normalize_number_token(match.group(1)))
    if not schedule_values and not series_values:
        for match in re.finditer(r"\bSCH\.?\s*([0-9]+S?|XXS|XS|STD)\b", desc, re.IGNORECASE):
            token = match.group(1).upper()
            if token in {"XS", "XXS", "STD"}:
                series_values.append(token)
            else:
                schedule_values.append(f"SCH{token}")
        for match in re.finditer(r"\bS-\s*([0-9]+S?)\b", desc, re.IGNORECASE):
            schedule_values.append(f"SCH{match.group(1).upper()}")

    return {
        "MM": uniq_keep_order(mm_values),
        "SCHEDULE": uniq_keep_order(schedule_values),
        "SERIES": uniq_keep_order(series_values),
        "BWG": [],
        "INCH": [],
    }


def build_pressure(pressure_code: str) -> str:
    code = pressure_code.upper().replace(" ", "")
    if not code:
        return ""
    if code.startswith("C") and code[1:].replace(".", "").isdigit():
        return f"CL{code[1:]}"
    return code


STD_PATTERNS: list[tuple[str, str, str | None]] = [
    ("AB3610", "ASME B36.10M", None),
    ("AB3619", "ASME B36.19M", None),
    ("AB318", "ASME B31.8", None),
    ("AB169", "ASME B16.9", None),
    ("AB165", "ASME B16.5", None),
    ("MSSSP-83", "MSS SP-83", None),
    ("MSSSP83", "MSS SP-83", None),
    ("MS97", "MSS SP-97", None),
    ("MS83", "MSS SP-83", None),
    ("GBT14976", "GB/T14976", None),
    ("GBT8163", "GB/T8163", None),
    ("GBT5310", "GB/T5310", None),
    ("GBT9711", "GB/T9711", None),
    ("GBT3091", "GB/T3091", None),
    ("GBT9948", "GB/T9948", None),
    ("GBT6479", "GB/T6479", None),
    ("GBT3087", "GB/T3087", None),
    ("GBT12771", "GB/T12771", None),
    ("GBT17395", "GB/T17395", None),
    ("GBT21833.2", "GB/T21833.2", None),
    ("GBT21832.2", "GB/T21832.2", None),
    ("GBT18984", "GB/T18984", None),
    ("GBT40317", "GB/T40317", None),
    ("GBT187422", "GB/T18742.2", None),
    ("GBT100021", "GB/T10002.1", None),
    ("GBT3621", "GB/T3621", None),
    ("GBT23257", "GB/T23257", None),
    ("GBT50235", "GB/T50235", None),
    ("HGT20553IA", "HG/T20553", "Ia"),
    ("HGT20553II", "HG/T20553", "II"),
    ("HGT20553I", "HG/T20553", "I"),
    ("HGT20553", "HG/T20553", None),
    ("HGT20538", "HG/T20538", None),
    ("HGT2130", "HG/T2130", None),
    ("HGT3731", "HG/T3731", None),
    ("HGT2437", "HG/T2437", None),
    ("HGT21501", "HG/T21501", None),
    ("HGT20592", "HG/T20592", None),
    ("SHT3405", "SH/T3405", None),
    ("SHT3406", "SH/T3406", None),
    ("SYT5037", "SY/T5037", None),
    ("SYT5257", "SY/T5257", None),
    ("EN102165", "EN 10216-5", None),
    ("EN10305", "EN 10305", None),
    ("ENI1127", "EN ISO 1127", None),
    ("I2037", "I2037", None),
]


def build_standard(standard_code: str) -> list[dict[str, str]]:
    raw = standard_code.upper().replace(" ", "")
    if not raw:
        return []

    found: list[tuple[int, int, str, str | None]] = []
    occupied: list[range] = []
    for pattern, body, grade in sorted(STD_PATTERNS, key=lambda x: -len(x[0])):
        for match in re.finditer(re.escape(pattern), raw):
            span = range(match.start(), match.end())
            if any(match.start() in r or match.end() - 1 in r for r in occupied):
                continue
            occupied.append(span)
            found.append((match.start(), match.end(), body, grade))

    found.sort(key=lambda x: (x[0], x[1]))
    out: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for _, _, body, grade in found:
        key = (body, grade or "")
        if key in seen:
            continue
        seen.add(key)
        out.append({"BODY": body, "GRADE": grade or "", "METHOD": "", "APPENDIX": ""})
    return out


def build_material(material_code: str) -> list[dict[str, Any]]:
    value = text(material_code)
    if not value:
        return []
    return [{"ROLE": "MAIN", "VALUE": value, "SPECIAL_REQ": []}]


def build_output(row: dict[str, str]) -> tuple[dict[str, Any], list[str]]:
    desc = row["材料描述"]
    type_code = row["类型编码"] or row["材料名称代码"]
    body, notes = infer_body(type_code, desc)

    output: OrderedDict[str, Any] = OrderedDict()
    output["TYPE"] = {
        "BODY": body,
        "MANU": infer_manu(type_code, desc),
        "CONN": infer_conn(desc),
        "ENDS": infer_ends(desc),
    }
    output["SIZE"] = build_size(row["公称直径代码"], desc)
    output["THICKNESS"] = build_thickness(row["壁厚等级代码"], desc)

    pressure = build_pressure(row["压力等级代码"])
    if pressure:
        output["PRESSURE"] = pressure

    material = build_material(row["材质代码"])
    if material:
        output["MATERIAL"] = material

    standard = build_standard(row["标准号代码"])
    if standard:
        output["STANDARD"] = standard

    if not output["SIZE"]["DN"] and not output["SIZE"]["OD"] and not output["SIZE"]["INCH"]:
        notes.append("SIZE 为空，请检查")
    if not material:
        notes.append("MATERIAL 为空，请检查")
    if row["标准号代码"] and not standard:
        notes.append(f"STANDARD 未能从标准号代码解析: {row['标准号代码']}")

    return output, notes


def write_review(path: Path, review_rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "需审核"
    headers = [
        "Excel行号",
        "材料描述",
        "类型编码",
        "BODY",
        "MANU",
        "CONN",
        "ENDS",
        "SIZE",
        "THICKNESS",
        "MATERIAL",
        "STANDARD",
        "审核原因",
    ]
    ws.append(headers)
    for row in review_rows:
        ws.append([row.get(h, "") for h in headers])
    for col in ws.columns:
        width = min(max(len(text(c.value)) for c in col[:30]) + 2, 100)
        ws.column_dimensions[col[0].column_letter].width = max(12, width)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="将直管覆盖抽样表转换为训练草稿 JSON")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args()

    rows = load_rows(args.input)
    dataset: list[dict[str, Any]] = []
    review_rows: list[dict[str, Any]] = []
    seen_inputs: set[str] = set()
    duplicate_count = 0

    for row in rows:
        desc = row["材料描述"]
        if desc in seen_inputs:
            duplicate_count += 1
            continue
        seen_inputs.add(desc)

        output, notes = build_output(row)
        dataset.append({"input": desc, "output": output})
        if notes:
            review_rows.append(
                {
                    "Excel行号": row.get("Excel行号", ""),
                    "材料描述": desc,
                    "类型编码": row.get("类型编码", ""),
                    "BODY": output["TYPE"]["BODY"],
                    "MANU": json.dumps(output["TYPE"]["MANU"], ensure_ascii=False),
                    "CONN": json.dumps(output["TYPE"]["CONN"], ensure_ascii=False),
                    "ENDS": json.dumps(output["TYPE"]["ENDS"], ensure_ascii=False),
                    "SIZE": json.dumps(output["SIZE"], ensure_ascii=False),
                    "THICKNESS": json.dumps(output["THICKNESS"], ensure_ascii=False),
                    "MATERIAL": json.dumps(output.get("MATERIAL", []), ensure_ascii=False),
                    "STANDARD": json.dumps(output.get("STANDARD", []), ensure_ascii=False),
                    "审核原因": "；".join(notes),
                }
            )

    args.output_dir.mkdir(parents=True, exist_ok=True)
    output_json = args.output_dir / "直管训练草稿.json"
    type_only_json = args.output_dir / "直管TYPE草稿.json"
    output_json.write_text(json.dumps(dataset, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    type_only_dataset = [
        {"input": item["input"], "output": {"TYPE": item["output"]["TYPE"]}}
        for item in dataset
    ]
    type_only_json.write_text(json.dumps(type_only_dataset, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_review(args.output_dir / "直管训练草稿_需审核.xlsx", review_rows)

    print(f"输入抽样行数: {len(rows)}")
    print(f"训练草稿条数: {len(dataset)}")
    print(f"重复描述跳过: {duplicate_count}")
    print(f"需审核条数: {len(review_rows)}")
    print(f"输出: {output_json}")
    print(f"TYPE输出: {type_only_json}")


if __name__ == "__main__":
    main()
