from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[4]
"""
 python apps/trainer/qwen3_fte/src/材料描述分类.py \
    --input "/Users/guoxi/Documents/PCF-IDF材料描述/中简项目.xlsx" \
    --column "材料描述"
"""
import sys

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.config import get_ner_config
from src.llm_ner.router import build_category_router
from src.tokenizer_utils.preprocessor import TextPreprocessor

try:
    from tqdm import tqdm
except Exception:  # pragma: no cover
    tqdm = None


APPEND_HEADERS = [
    "预处理描述",
    "路由类别",
    "进入编码",
    "跳过编码原因",
    "路由置信度",
    "路由来源",
    "路由理由",
    "候选类别",
    "路由错误",
]


def text(value: Any) -> str:
    return str(value or "").strip()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="按当前平台一阶段路由配置对 Excel 描述列进行分类，并导出结果 Excel")
    parser.add_argument("--input", required=True, type=Path, help="输入 Excel 路径")
    parser.add_argument("--column", required=True, help="描述列表头名称")
    parser.add_argument("--output", type=Path, help="输出 Excel 路径；默认在输入文件旁生成“_分类结果.xlsx”")
    parser.add_argument("--sheet", help="工作表名称；不填则使用第一个工作表")
    parser.add_argument("--header-row", type=int, default=1, help="表头行号，默认 1")
    parser.add_argument("--no-preprocess", action="store_true", help="不做文本预处理")
    return parser.parse_args()


def get_router_and_config():
    ner_config = get_ner_config()
    qwen3_config = ner_config.get("qwen3", {}) or {}
    router_cfg = qwen3_config.get("router", {}) or {}
    router = build_category_router(router_cfg, project_root=PROJECT_ROOT)
    encodable_categories = set(router_cfg.get("encodable_categories") or [])
    return router, encodable_categories


def build_output_path(input_path: Path, output: Path | None) -> Path:
    if output is not None:
        return output
    return input_path.with_name(f"{input_path.stem}_分类结果.xlsx")


def load_sheet(input_path: Path, sheet_name: str | None):
    wb = load_workbook(input_path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"工作表不存在: {sheet_name}")
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]
    return wb, ws


def classify_text(router, encodable_categories: set[str], preprocessor: TextPreprocessor | None, raw_text: str) -> dict[str, Any]:
    if not text(raw_text):
        return {
            "processed_text": "",
            "category": "",
            "encoding_enabled": "否",
            "skip_reason": "空描述",
            "confidence": "",
            "source": "",
            "reason": "",
            "candidates": "",
            "error": "",
        }
    processed = preprocessor.process(raw_text) if preprocessor is not None else raw_text
    try:
        route_info = router.route(processed)
        category = text(route_info.get("category"))
        encoding_enabled = category in encodable_categories
        skip_reason = "" if encoding_enabled else (f"类别“{category}”只分类，不参与编码" if category else "")
        return {
            "processed_text": processed,
            "category": category,
            "encoding_enabled": "是" if encoding_enabled else "否",
            "skip_reason": skip_reason,
            "confidence": route_info.get("confidence", ""),
            "source": text(route_info.get("source")),
            "reason": text(route_info.get("reason")),
            "candidates": json.dumps(route_info.get("candidates") or [], ensure_ascii=False),
            "error": text(route_info.get("error")),
        }
    except Exception as exc:
        return {
            "processed_text": processed,
            "category": "",
            "encoding_enabled": "否",
            "skip_reason": "",
            "confidence": "",
            "source": "",
            "reason": "",
            "candidates": "",
            "error": str(exc),
        }


def main():
    args = parse_args()
    input_path = args.input.resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"输入文件不存在: {input_path}")

    router, encodable_categories = get_router_and_config()
    preprocessor = None if args.no_preprocess else TextPreprocessor()

    _, ws = load_sheet(input_path, args.sheet)
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel 为空")

    header_idx = args.header_row - 1
    if header_idx < 0 or header_idx >= len(rows):
        raise ValueError(f"header-row 超出范围: {args.header_row}")

    headers = [text(v) for v in rows[header_idx]]
    if args.column not in headers:
        raise ValueError(f"未找到列名: {args.column}")
    col_idx = headers.index(args.column)

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "分类结果"

    out_headers = headers + APPEND_HEADERS
    out_ws.append(out_headers)

    data_rows = rows[header_idx + 1:]
    iterator = data_rows
    if tqdm is not None:
        iterator = tqdm(data_rows, desc="分类进度", unit="行")
    else:
        print(f"开始分类，共 {len(data_rows)} 行", flush=True)

    for idx, row in enumerate(iterator, start=1):
        row_values = list(row)
        if len(row_values) < len(headers):
            row_values.extend([""] * (len(headers) - len(row_values)))
        raw_text = text(row_values[col_idx])
        result = classify_text(router, encodable_categories, preprocessor, raw_text)
        out_ws.append(
            row_values
            + [
                result["processed_text"],
                result["category"],
                result["encoding_enabled"],
                result["skip_reason"],
                result["confidence"],
                result["source"],
                result["reason"],
                result["candidates"],
                result["error"],
            ]
        )
        if tqdm is None and (idx % 50 == 0 or idx == len(data_rows)):
            print(f"已完成 {idx}/{len(data_rows)} 行", flush=True)

    output_path = build_output_path(input_path, args.output).resolve()
    out_wb.save(output_path)
    print(f"输出完成: {output_path}")


if __name__ == "__main__":
    main()
