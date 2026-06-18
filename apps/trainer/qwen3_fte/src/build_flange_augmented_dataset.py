#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[4]
DEFAULT_BASE = PROJECT_ROOT / "apps/trainer/qwen3_fte/output/flange_project_sampling/法兰训练草稿_with_body_contrast_clean.json"
DEFAULT_EXCEL = Path("/Users/guoxi/Desktop/总数据集.xlsx")
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "apps/trainer/qwen3_fte/output/flange_project_sampling"


TYPE_CODE_TO_BODY = {
    "F": "法兰",
    "FSO": "带颈平焊法兰",
    "FS": "承插焊法兰",
    "FRJ": "带颈对焊法兰",
    "FLM": "带颈对焊法兰",
    "FLF": "带颈对焊法兰",
    "FM": "带颈对焊法兰",
    "FFF": "带颈对焊法兰",
    "BF": "法兰盖",
    "BFRJ": "盲法兰",
    "BFN": "盲法兰",
    "8BF": "8字盲板",
    "8BFFF": "8字盲板",
    "8BFRJ": "8字盲板",
    "8BFW": "8字盲板",
    "8BFWFF": "8字盲板",
    "FPL": "板式平焊法兰",
    "FTH": "螺纹法兰",
    "FFN": "螺纹法兰",
    "FN": "螺纹法兰",
    "FJ": "带颈对焊夹套法兰",
    "FJDL": "带颈对焊夹套法兰",
    "FJSO": "带颈平焊夹套法兰",
    "JSOF": "带颈平焊夹套法兰",
    "RFJSO": "带颈平焊夹套法兰",
    "LF": "松套法兰",
    "LFPL": "松套法兰",
    "LFP": "衬里法兰",
    "FSLM": "承插焊法兰",
    "FSLF": "承插焊法兰",
    "FSRJ": "承插焊法兰",
    "FRJS": "承插焊法兰",
}

BODY_TARGETS = {
    "法兰": 450,
    "带颈平焊法兰": 450,
    "承插焊法兰": 520,
    "螺纹法兰": 160,
    "板式平焊法兰": 120,
    "插板": 80,
    "带颈对焊夹套法兰": 100,
    "带颈平焊夹套法兰": 80,
    "松套法兰": 100,
    "衬里法兰": 80,
    "孔板对焊法兰": 60,
    "对焊法兰": 60,
}


def text(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def clean_num(v: str) -> str:
    v = v.strip()
    if "." in v:
        v = v.rstrip("0").rstrip(".")
    return v


def norm_code(v: str) -> str:
    return re.sub(r"[^A-Za-z0-9.]+", "", v or "").upper()


def row_type_code(row: dict[str, str]) -> str:
    return text(row.get("修正种类")) or text(row.get("标准化种类"))


def infer_body(row: dict[str, str]) -> str:
    code = row_type_code(row)
    if code in TYPE_CODE_TO_BODY:
        return TYPE_CODE_TO_BODY[code]
    desc = text(row.get("材料描述"))
    raw_type = text(row.get("原始种类"))
    s = f"{desc} {raw_type}".upper()
    if any(x in s for x in ["SPECTACLE", "FIGURE8", "FIGURE-8", "8字盲板", "八字盲板"]):
        return "8字盲板"
    if "BLIND" in s or "盲法兰" in s:
        return "盲法兰"
    if "法兰盖" in s:
        return "法兰盖"
    if "SOCKET" in s or re.search(r"(?<![A-Z0-9])SW(?![A-Z0-9])", s):
        return "承插焊法兰"
    if "THREAD" in s or "THRD" in s or "SCRD" in s or "螺纹" in s or "FNPT" in s or "NPT" in s:
        return "螺纹法兰"
    if "SLIP" in s or re.search(r"(?<![A-Z0-9])SO(?![A-Z0-9])", s):
        return "带颈平焊法兰"
    if "WELD NECK" in s or "WELDING NECK" in s or re.search(r"(?<![A-Z0-9])WN(?![A-Z0-9])", s):
        return "带颈对焊法兰"
    if "板式平焊" in s or "PAD FLANGE" in s:
        return "板式平焊法兰"
    if "衬里法兰" in s:
        return "衬里法兰"
    if "松套" in s or "LAP JOINT" in s:
        return "松套法兰"
    return ""


def parse_seal(desc: str, code: str) -> list[str]:
    s = f"{desc} {code}".upper()
    found: list[str] = []
    for token in ["FLRJ", "MFM", "RTJ", "RJ", "LM", "LF", "FF", "RF"]:
        if re.search(rf"(?<![A-Z0-9]){token}(?![A-Z0-9])", s) or token in norm_code(s):
            if token == "RTJ":
                found.append("RTJ")
            elif token == "FLRJ":
                found.append("FLRJ")
            else:
                found.append(token)
            break
    return found


def parse_conn_ends(body: str, desc: str, code: str) -> tuple[list[str], list[str]]:
    s = f"{desc} {code}".upper()
    conn: list[str] = []
    ends: list[str] = []
    if body == "承插焊法兰" or "SOCKET" in s or re.search(r"(?<![A-Z0-9])SW(?![A-Z0-9])", s):
        conn.append("SW")
    if "FNPT" in s:
        ends.append("FNPT")
    elif "NPT" in s:
        conn.append("NPT")
    elif "THD" in s or "THREAD" in s or "THRD" in s or "SCRD" in s or "螺纹" in s:
        conn.append("THD")
    return list(dict.fromkeys(conn)), list(dict.fromkeys(ends))


def build_size(code: str, desc: str) -> dict[str, list[str]]:
    raw = text(code)
    dn: list[str] = []
    od: list[str] = []
    inch: list[str] = []
    length: list[str] = []
    if raw:
        for part in re.split(r"[xX×+/]", raw):
            part = part.strip()
            if not part:
                continue
            if '"' in part:
                inch.append(part)
            elif re.fullmatch(r"\d+(?:\.\d+)?", part):
                dn.append("DN" + clean_num(part))
    for m in re.finditer(r"(\d+(?:\.\d+)?(?:-\d+/\d+)?|\d+/\d+)\s*\"", desc):
        inch.append(m.group(1).replace(" ", "") + '"')
    for m in re.finditer(r"[Φφ]\s*(\d+(?:\.\d+)?)", desc):
        od.append(clean_num(m.group(1)))
    return {"DN": list(dict.fromkeys(dn)), "OD": list(dict.fromkeys(od)), "INCH": list(dict.fromkeys(inch)), "LENGTH": length}


def build_thickness(code: str) -> dict[str, list[str]]:
    raw = text(code).replace(" ", "")
    mm: list[str] = []
    sch: list[str] = []
    series: list[str] = []
    if raw:
        for part in re.split(r"[xX×+/]", raw):
            p = part.strip()
            if not p:
                continue
            up = p.upper()
            if up in {"STD", "XS", "XXS"}:
                series.append(up)
            elif up.endswith("MM"):
                mm.append(clean_num(up[:-2]))
            elif re.fullmatch(r"\d+(?:\.\d+)?", up):
                mm.append(clean_num(up))
            elif up.startswith("SCH"):
                val = up[3:]
                if val in {"STD", "XS", "XXS"}:
                    series.append(val)
                elif val:
                    sch.append("SCH" + val)
            elif up.startswith("S"):
                val = up[1:].lstrip("-")
                if val in {"TD", "STD"}:
                    series.append("STD")
                elif val in {"XS", "XXS"}:
                    series.append(val)
                elif val:
                    sch.append("SCH" + val)
    return {"MM": list(dict.fromkeys(mm)), "SCHEDULE": list(dict.fromkeys(sch)), "SERIES": list(dict.fromkeys(series)), "BWG": [], "INCH": []}


def build_pressure(code: str) -> str:
    c = text(code).replace(" ", "").upper()
    if not c:
        return ""
    if c.startswith("C") and c[1:].replace(".", "").isdigit():
        return "CL" + c[1:]
    if re.fullmatch(r"\d+LB", c):
        return "CL" + c[:-2]
    if re.fullmatch(r"\d+#", c):
        return "CL" + c[:-1]
    return c


STD_MAP = [
    ("ASMEB16.48", "ASME B16.48", ""),
    ("ASMEB16.47", "ASME B16.47", ""),
    ("ASMEB16.36", "ASME B16.36", ""),
    ("ASMEB16.5", "ASME B16.5", ""),
    ("AB1648", "ASME B16.48", ""),
    ("AB1647", "ASME B16.47", ""),
    ("AB1636", "ASME B16.36", ""),
    ("AB165", "ASME B16.5", ""),
    ("HGT20592SERIESA", "HG/T20592", "Series A"),
    ("HGT20592SERIESB", "HG/T20592", "Series B"),
    ("HGT20592A", "HG/T20592", "A"),
    ("HGT20592B", "HG/T20592", "B"),
    ("HGT20592", "HG/T20592", ""),
    ("HGT20615", "HG/T20615", ""),
    ("HGT20623", "HG/T20623", ""),
    ("HGT21547", "HG/T21547", ""),
    ("HG21547", "HG21547", ""),
    ("SHT3406", "SH/T3406", ""),
    ("SHT3425", "SH/T3425", ""),
    ("GBT12228", "GB/T12228", ""),
    ("GBT9124.1", "GB/T9124.1", ""),
    ("GBT9124.2", "GB/T9124.2", ""),
    ("NBT47008", "NB/T47008", ""),
    ("NBT47009", "NB/T47009", ""),
    ("NBT47010", "NB/T47010", ""),
    ("GBT4237", "GB/T4237", ""),
    ("GBT713", "GB/T713", ""),
    ("GBT711", "GB/T711", ""),
]


def build_standard(code: str) -> list[dict[str, str]]:
    raw = norm_code(text(code))
    if not raw:
        return []
    out: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for pat, body, grade in sorted(STD_MAP, key=lambda x: -len(x[0])):
        if pat in raw:
            g = grade
            if body == "ASME B16.47":
                m = re.search(r"(?:ASMEB16.47|AB1647)([AB])", raw)
                if m:
                    g = m.group(1)
            key = (body, g)
            if key not in seen:
                seen.add(key)
                out.append({"BODY": body, "GRADE": g, "METHOD": "", "APPENDIX": ""})
    return out


def material_value(row: dict[str, str]) -> str:
    return text(row.get("原始材质")) or text(row.get("修正材质")) or text(row.get("标准化材质"))


def make_record(row: dict[str, str]) -> dict[str, Any] | None:
    body = infer_body(row)
    if not body:
        return None
    desc = text(row.get("材料描述"))
    code = row_type_code(row)
    conn, ends = parse_conn_ends(body, desc, code)
    seal = parse_seal(desc, code)
    output: dict[str, Any] = {
        "TYPE": {"BODY": body, "CONN": conn, "SEAL": seal, "ENDS": ends},
        "SIZE": build_size(text(row.get("修正尺寸")) or text(row.get("标准化尺寸")), desc),
        "THICKNESS": build_thickness(text(row.get("修正壁厚")) or text(row.get("标准化壁厚"))),
        "PRESSURE": build_pressure(text(row.get("修正磅级")) or text(row.get("磅级编码"))),
        "MATERIAL": [],
        "STANDARD": build_standard(text(row.get("修正规范")) or text(row.get("标准化规范"))),
    }
    mat = material_value(row)
    if mat:
        special = []
        if any(x in desc.upper() for x in ["NACE", "ANTI-H2S", "H2S"]):
            special.append("NACE")
        if any(x in desc.upper() for x in ["GALV", "GALVANIZED", "镀锌", "+ZN"]):
            special.append("GALV")
        output["MATERIAL"] = [{"ROLE": "MAIN", "VALUE": mat, "SPECIAL_REQ": special}]
    return {"input": desc, "output": output}


def record_values(rec: dict[str, Any]) -> dict[str, list[str]]:
    out = rec["output"]
    thk = out.get("THICKNESS", {})
    size = out.get("SIZE", {})
    return {
        "body": [out.get("TYPE", {}).get("BODY", "")],
        "pressure": [out.get("PRESSURE", "")],
        "material": [m.get("VALUE", "") for m in out.get("MATERIAL", [])],
        "standard": [s.get("BODY", "") + (" [" + s.get("GRADE", "") + "]" if s.get("GRADE") else "") for s in out.get("STANDARD", [])],
        "thickness": (thk.get("MM", []) or []) + (thk.get("SCHEDULE", []) or []) + (thk.get("SERIES", []) or []),
        "size": (size.get("DN", []) or []) + (size.get("OD", []) or []) + (size.get("INCH", []) or []),
        "seal": out.get("TYPE", {}).get("SEAL", []),
        "conn": out.get("TYPE", {}).get("CONN", []),
        "ends": out.get("TYPE", {}).get("ENDS", []),
    }


def load_excel_rows(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [text(x) for x in next(rows)]
    idx = {h: i for i, h in enumerate(header)}
    out = []
    for values in rows:
        row = {h: text(values[i]) if i < len(values) else "" for h, i in idx.items()}
        if row.get("分类") == "法兰" and row.get("材料描述"):
            out.append(row)
    return out


def main() -> None:
    parser = argparse.ArgumentParser(description="从总数据集抽取法兰合法增强样本")
    parser.add_argument("--base", type=Path, default=DEFAULT_BASE)
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--limit", type=int, default=800)
    args = parser.parse_args()

    base = json.loads(args.base.read_text(encoding="utf-8"))
    base_inputs = {r["input"] for r in base}
    base_counts = {k: Counter() for k in ["body", "pressure", "material", "standard", "thickness", "size", "seal", "conn", "ends"]}
    for rec in base:
        vals = record_values(rec)
        for k, vs in vals.items():
            for v in vs or ["<EMPTY>"]:
                base_counts[k][v or "<EMPTY>"] += 1

    rows = load_excel_rows(args.excel)
    candidates: list[tuple[float, dict[str, Any], str]] = []
    for row in rows:
        rec = make_record(row)
        if not rec or rec["input"] in base_inputs:
            continue
        vals = record_values(rec)
        body = vals["body"][0]
        if body not in {k for k in TYPE_CODE_TO_BODY.values()}:
            continue
        # Skip obvious non-flange fittings living in 法兰 category.
        desc_u = rec["input"].upper()
        if any(x in desc_u for x in ["ELBOW", "弯头", "TEE", "三通", "VALVE", "阀门", "软管"]):
            continue
        score = 0.0
        reasons = []
        target = BODY_TARGETS.get(body)
        if target and base_counts["body"][body] < target:
            score += (target - base_counts["body"][body]) / max(target, 1) * 50
            reasons.append(f"补BODY={body}")
        for dim, weight in [
            ("pressure", 8),
            ("material", 7),
            ("standard", 8),
            ("thickness", 8),
            ("size", 4),
            ("seal", 6),
            ("conn", 6),
            ("ends", 6),
        ]:
            for v in vals.get(dim, []):
                if not v:
                    continue
                c = base_counts[dim][v]
                if c <= 3:
                    score += weight * 3
                    reasons.append(f"低频{dim}={v}")
                elif c <= 10:
                    score += weight
        if score > 0:
            candidates.append((score, rec, "；".join(reasons[:8])))

    selected: list[dict[str, Any]] = []
    seen = set(base_inputs)
    for score, rec, reason in sorted(candidates, key=lambda x: -x[0]):
        if len(selected) >= args.limit:
            break
        if rec["input"] in seen:
            continue
        seen.add(rec["input"])
        rec["_增强原因"] = reason
        selected.append(rec)
        vals = record_values(rec)
        for k, vs in vals.items():
            for v in vs or ["<EMPTY>"]:
                base_counts[k][v or "<EMPTY>"] += 1

    supplement = [{k: v for k, v in rec.items() if not k.startswith("_")} for rec in selected]
    merged = base + supplement

    args.output_dir.mkdir(parents=True, exist_ok=True)
    supp_path = args.output_dir / "法兰骨架增强补充样本_v2.json"
    merged_path = args.output_dir / "法兰训练草稿_augmented_v2.json"
    report_path = args.output_dir / "法兰骨架增强报告_v2.xlsx"
    supp_path.write_text(json.dumps(supplement, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    merged_path.write_text(json.dumps(merged, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    wb = Workbook()
    ws = wb.active
    ws.title = "增强样本"
    headers = ["增强原因", "BODY", "PRESSURE", "MATERIAL", "STANDARD", "THICKNESS", "SIZE", "input"]
    ws.append(headers)
    for rec in selected:
        vals = record_values(rec)
        ws.append([
            rec.get("_增强原因", ""),
            ",".join(vals["body"]),
            ",".join(vals["pressure"]),
            ",".join(vals["material"]),
            ",".join(vals["standard"]),
            ",".join(vals["thickness"]),
            ",".join(vals["size"]),
            rec["input"],
        ])
    ws2 = wb.create_sheet("分布")
    ws2.append(["维度", "值", "增强后次数"])
    for dim, counter in base_counts.items():
        for value, count in counter.most_common():
            ws2.append([dim, value, count])
    for sheet in wb.worksheets:
        for col in sheet.columns:
            width = min(max(len(text(c.value)) for c in col[:30]) + 2, 100)
            sheet.column_dimensions[col[0].column_letter].width = max(12, width)
    wb.save(report_path)

    print(f"base={len(base)} supplement={len(supplement)} merged={len(merged)}")
    print(supp_path)
    print(merged_path)
    print(report_path)
    print(Counter(rec['output']['TYPE']['BODY'] for rec in supplement))


if __name__ == "__main__":
    main()
