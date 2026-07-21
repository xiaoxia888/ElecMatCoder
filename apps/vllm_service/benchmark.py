# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import csv
import json
import statistics
import subprocess
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from requests.adapters import HTTPAdapter


_HTTP_SESSION = requests.Session()
_HTTP_ADAPTER = HTTPAdapter(
    pool_connections=64,
    pool_maxsize=64,
    max_retries=0,
    pool_block=True,
)
_HTTP_SESSION.mount("http://", _HTTP_ADAPTER)
_HTTP_SESSION.mount("https://", _HTTP_ADAPTER)


DEFAULT_TEXTS = [
    "TEE,RED SMLS BW A234 WPB ASME B16.9 SCH160xSCH160 DN100x80",
    "法兰管，PTFE lined GB/T 8163-20，RF，PN16，HG/T20538，SMLS，DN80，4.0mm",
    "SWAGE NIPPLE,CON Forged PBE NB/T 47008 12Cr5Mo SH/T3419 SCH80xSCH80 DN40x25",
]


def _percentile(values: list[float], percentile: float) -> float:
    if not values:
        return 0.0
    ordered = sorted(values)
    position = (len(ordered) - 1) * percentile
    lower = int(position)
    upper = min(lower + 1, len(ordered) - 1)
    fraction = position - lower
    return ordered[lower] * (1 - fraction) + ordered[upper] * fraction


def _summary(values: list[float]) -> dict[str, float]:
    if not values:
        return {"count": 0, "mean": 0.0, "p50": 0.0, "p95": 0.0, "min": 0.0, "max": 0.0}
    return {
        "count": len(values),
        "mean": round(statistics.mean(values), 4),
        "p50": round(_percentile(values, 0.50), 4),
        "p95": round(_percentile(values, 0.95), 4),
        "min": round(min(values), 4),
        "max": round(max(values), 4),
    }


def _read_tabular_rows(path: Path, text_column: str) -> list[str]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))
        if rows and text_column not in rows[0]:
            raise ValueError(f"CSV不存在列 {text_column}，可用列: {list(rows[0])}")
        return [str(row.get(text_column) or "").strip() for row in rows]
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
        if text_column not in headers:
            raise ValueError(f"Excel不存在列 {text_column}，可用列: {headers}")
        index = headers.index(text_column)
        return [str(row[index] or "").strip() for row in rows if index < len(row)]
    raise ValueError(f"不支持的表格格式: {suffix}")


def read_texts(path: Path | None, text_column: str, inline_texts: list[str], limit: int) -> list[str]:
    texts = [value.strip() for value in inline_texts if value.strip()]
    if path is not None:
        path = path.expanduser()
        suffix = path.suffix.lower()
        if suffix in {".csv", ".xlsx", ".xlsm"}:
            texts.extend(_read_tabular_rows(path, text_column))
        elif suffix == ".jsonl":
            for line in path.read_text(encoding="utf-8").splitlines():
                row = json.loads(line)
                value = row.get(text_column) if isinstance(row, dict) else row
                texts.append(str(value or "").strip())
        elif suffix == ".json":
            data = json.loads(path.read_text(encoding="utf-8"))
            if not isinstance(data, list):
                raise ValueError("JSON输入必须是数组")
            for row in data:
                value = row.get(text_column) if isinstance(row, dict) else row
                texts.append(str(value or "").strip())
        else:
            texts.extend(line.strip() for line in path.read_text(encoding="utf-8").splitlines())
    result = [value for value in texts if value]
    if not result:
        result = list(DEFAULT_TEXTS)
    return result[:limit] if limit > 0 else result


@dataclass
class GpuSample:
    timestamp: float
    devices: list[dict[str, Any]]


class GpuSampler:
    def __init__(self, interval_seconds: float):
        self.interval_seconds = max(0.1, interval_seconds)
        self.samples: list[GpuSample] = []
        self.error = ""
        self._stop = threading.Event()
        self._thread: threading.Thread | None = None

    def start(self) -> None:
        self._sample_once()
        self._thread = threading.Thread(target=self._run, daemon=True)
        self._thread.start()

    def stop(self) -> None:
        self._stop.set()
        if self._thread is not None:
            self._thread.join(timeout=self.interval_seconds * 2 + 1)
        self._sample_once()

    def _run(self) -> None:
        while not self._stop.wait(self.interval_seconds):
            self._sample_once()

    def _sample_once(self) -> None:
        command = [
            "nvidia-smi",
            "--query-gpu=index,name,memory.used,memory.total,utilization.gpu",
            "--format=csv,noheader,nounits",
        ]
        try:
            output = subprocess.check_output(command, text=True, timeout=10)
            devices = []
            for line in output.splitlines():
                parts = [part.strip() for part in line.split(",")]
                if len(parts) != 5:
                    continue
                devices.append(
                    {
                        "index": int(parts[0]),
                        "name": parts[1],
                        "memory_used_mb": float(parts[2]),
                        "memory_total_mb": float(parts[3]),
                        "utilization_percent": float(parts[4]),
                    }
                )
            self.samples.append(GpuSample(timestamp=time.time(), devices=devices))
        except Exception as exc:
            if not self.error:
                self.error = str(exc)

    def summary(self) -> dict[str, Any]:
        by_index: dict[int, dict[str, Any]] = {}
        for sample in self.samples:
            for device in sample.devices:
                index = int(device["index"])
                item = by_index.setdefault(
                    index,
                    {
                        "index": index,
                        "name": device["name"],
                        "memory_total_mb": device["memory_total_mb"],
                        "memory_used": [],
                        "utilization": [],
                    },
                )
                item["memory_used"].append(device["memory_used_mb"])
                item["utilization"].append(device["utilization_percent"])
        devices = []
        for item in by_index.values():
            devices.append(
                {
                    "index": item["index"],
                    "name": item["name"],
                    "memory_total_mb": item["memory_total_mb"],
                    "memory_used_mb": _summary(item["memory_used"]),
                    "utilization_percent": _summary(item["utilization"]),
                }
            )
        return {
            "sample_count": len(self.samples),
            "devices": sorted(devices, key=lambda item: item["index"]),
            "error": self.error,
        }


def call_predict(
    *,
    service_url: str,
    model: str,
    text: str,
    max_new_tokens: int,
    timeout: int,
) -> dict[str, Any]:
    started = time.perf_counter()
    response = _HTTP_SESSION.post(
        f"{service_url.rstrip('/')}/predict",
        headers={"Connection": "keep-alive"},
        json={
            "model": model,
            "text": text,
            "max_new_tokens": max_new_tokens,
            "temperature": 0.0,
            "top_p": 1.0,
        },
        timeout=timeout,
    )
    wall_seconds = time.perf_counter() - started
    response.raise_for_status()
    payload = response.json()
    return {
        "model": model,
        "wall_seconds": round(wall_seconds, 4),
        "service_elapsed_seconds": payload.get("elapsed_seconds"),
        "engine": payload.get("engine"),
        "upstream_model": payload.get("upstream_model"),
        "json_parse_ok": payload.get("json_parse_ok"),
        "usage": payload.get("usage") or {},
        "raw_preview": str(payload.get("raw_response") or "")[:160],
    }


def run_group(
    *,
    service_url: str,
    models: list[str],
    text: str,
    max_new_tokens: int,
    timeout: int,
    parallel: bool,
) -> dict[str, Any]:
    started = time.perf_counter()
    if parallel:
        with ThreadPoolExecutor(max_workers=len(models)) as pool:
            futures = [
                pool.submit(
                    call_predict,
                    service_url=service_url,
                    model=model,
                    text=text,
                    max_new_tokens=max_new_tokens,
                    timeout=timeout,
                )
                for model in models
            ]
            calls = [future.result() for future in futures]
    else:
        calls = [
            call_predict(
                service_url=service_url,
                model=model,
                text=text,
                max_new_tokens=max_new_tokens,
                timeout=timeout,
            )
            for model in models
        ]
    wall_seconds = time.perf_counter() - started
    return {
        "text_preview": text[:160],
        "wall_seconds": round(wall_seconds, 4),
        "calls": calls,
    }


def run_scenario(
    *,
    service_url: str,
    models: list[str],
    texts: list[str],
    rounds: int,
    max_new_tokens: int,
    timeout: int,
    parallel_models: bool,
    group_concurrency: int,
) -> dict[str, Any]:
    jobs = [text for _ in range(rounds) for text in texts]
    started = time.perf_counter()
    groups: list[dict[str, Any]] = []
    with ThreadPoolExecutor(max_workers=max(1, group_concurrency)) as pool:
        futures = [
            pool.submit(
                run_group,
                service_url=service_url,
                models=models,
                text=text,
                max_new_tokens=max_new_tokens,
                timeout=timeout,
                parallel=parallel_models,
            )
            for text in jobs
        ]
        for future in as_completed(futures):
            groups.append(future.result())
    total = time.perf_counter() - started
    group_walls = [float(group["wall_seconds"]) for group in groups]
    model_latencies: dict[str, list[float]] = {model: [] for model in models}
    for group in groups:
        for call in group["calls"]:
            model_latencies.setdefault(call["model"], []).append(float(call["wall_seconds"]))
    return {
        "mode": "parallel_models" if parallel_models else "sequential_models",
        "group_concurrency": group_concurrency,
        "group_count": len(groups),
        "request_count": len(groups) * len(models),
        "total_wall_seconds": round(total, 4),
        "groups_per_second": round(len(groups) / total, 4) if total else 0.0,
        "requests_per_second": round(len(groups) * len(models) / total, 4) if total else 0.0,
        "group_wall_seconds": _summary(group_walls),
        "model_wall_seconds": {
            model: _summary(values) for model, values in model_latencies.items()
        },
        "groups": groups,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="vLLM多LoRA串行/并发及GPU显存批量基准测试")
    parser.add_argument("--service-url", default="http://127.0.0.1:8200")
    parser.add_argument(
        "--models",
        nargs="+",
        default=["type", "size-thick-pressure", "material-standard"],
    )
    parser.add_argument("--input", type=Path, help="txt/json/jsonl/csv/xlsx测试数据")
    parser.add_argument("--text-column", default="材料描述")
    parser.add_argument("--text", action="append", default=[])
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--rounds", type=int, default=1)
    parser.add_argument("--group-concurrency", type=int, default=1, help="同时处理多少条材料描述")
    parser.add_argument("--max-new-tokens", type=int, default=512)
    parser.add_argument("--timeout", type=int, default=600)
    parser.add_argument("--warmup", action="store_true")
    parser.add_argument("--gpu-sample-interval", type=float, default=0.5)
    parser.add_argument("--output", type=Path)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    texts = read_texts(args.input, args.text_column, args.text, args.limit)

    health_response = _HTTP_SESSION.get(
        f"{args.service_url.rstrip('/')}/health", timeout=min(30, args.timeout)
    )
    health_response.raise_for_status()
    health = health_response.json()
    if not health.get("ok"):
        raise RuntimeError(f"服务健康检查未通过: {json.dumps(health, ensure_ascii=False)}")

    if args.warmup:
        for model in args.models:
            call_predict(
                service_url=args.service_url,
                model=model,
                text=texts[0],
                max_new_tokens=args.max_new_tokens,
                timeout=args.timeout,
            )

    sampler = GpuSampler(args.gpu_sample_interval)
    sampler.start()
    try:
        sequential = run_scenario(
            service_url=args.service_url,
            models=args.models,
            texts=texts,
            rounds=args.rounds,
            max_new_tokens=args.max_new_tokens,
            timeout=args.timeout,
            parallel_models=False,
            group_concurrency=args.group_concurrency,
        )
        parallel = run_scenario(
            service_url=args.service_url,
            models=args.models,
            texts=texts,
            rounds=args.rounds,
            max_new_tokens=args.max_new_tokens,
            timeout=args.timeout,
            parallel_models=True,
            group_concurrency=args.group_concurrency,
        )
    finally:
        sampler.stop()

    seq_total = float(sequential["total_wall_seconds"])
    parallel_total = float(parallel["total_wall_seconds"])
    result = {
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "service_url": args.service_url,
        "models": args.models,
        "text_count": len(texts),
        "rounds": args.rounds,
        "group_concurrency": args.group_concurrency,
        "comparison": {
            "sequential_total_seconds": seq_total,
            "parallel_total_seconds": parallel_total,
            "speedup": round(seq_total / parallel_total, 4) if parallel_total else None,
            "saved_seconds": round(seq_total - parallel_total, 4),
            "interpretation": "并发时间接近最慢模型表示有效并发；接近各模型耗时之和表示接近串行。",
        },
        "gpu": sampler.summary(),
        "service_health": health,
        "sequential": sequential,
        "parallel": parallel,
    }

    output_path = args.output
    if output_path is None:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = Path("outputs") / f"vllm_benchmark_{stamp}.json"
    output_path = output_path.expanduser()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    concise = {
        "result_file": str(output_path.resolve()),
        "models": args.models,
        "text_count": len(texts),
        "comparison": result["comparison"],
        "sequential_group_latency": sequential["group_wall_seconds"],
        "parallel_group_latency": parallel["group_wall_seconds"],
        "parallel_model_latency": parallel["model_wall_seconds"],
        "gpu": result["gpu"],
    }
    print(json.dumps(concise, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
