from __future__ import annotations

import argparse
import csv
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


DEFAULT_INPUT = Path(
    "/Users/guoxi/Desktop/workspace/NJNCC/python_code/review_platform/materials_export.csv"
)
DEFAULT_OUTPUT = (
    Path(__file__).resolve().parents[1] / "output" / "material_rule_analysis.xlsx"
)

MIN_COUNT = 3
STRONG_RATIO = 0.98

PURE_NUMERIC_RE = re.compile(r"^\d+(?:\.\d+)?$")
NUMERIC_LIKE_RE = re.compile(r"^\d+(?:[#./+-]\d+)?[#A-Za-z]*$")
SHORT_TOKEN_RE = re.compile(r"^[A-Za-z0-9]{1,6}$")
GRADE_III_RE = re.compile(r"\b(?:GR\.?\s*)?III\b", re.IGNORECASE)
GRADE_II_RE = re.compile(r"\b(?:GR\.?\s*)?II\b", re.IGNORECASE)
ZN_RE = re.compile(r"\bZN\b|GALV|GALVANIZED|镀锌", re.IGNORECASE)
CE_RE = re.compile(r"ANTI-?H2S|NACE|抗硫", re.IGNORECASE)
HAS_CJK_RE = re.compile(r"[\u4e00-\u9fff]")


@dataclass
class MaterialEntry:
    raw: str
    total_count: int
    in_desc_count: int
    dominant_target: str
    consistency: float
    all_targets: dict[str, int]
    examples: list[str]
    category: str
    reason: str
    shape: str
    dominant_base: str
    dominant_suffixes: list[str]
    all_bases: dict[str, int]
    rule_candidates: list[str]


def _shape_of(raw: str) -> str:
    value = raw.strip()
    upper = value.upper()
    if upper.startswith("ASTM ") or upper.startswith("ASME "):
        return "标准牌号"
    if value.startswith("S") and any(ch.isdigit() for ch in value[1:]):
        return "S牌号"
    if value.startswith("SF") and any(ch.isdigit() for ch in value[2:]):
        return "SF牌号"
    if value.startswith("CF") and any(ch.isdigit() for ch in value[2:]):
        return "CF牌号"
    if "PTFE" in upper or "PE" in upper or "衬胶" in value or "GLASS LINED" in upper:
        return "复合/衬里"
    if PURE_NUMERIC_RE.fullmatch(value):
        return "纯数字"
    if NUMERIC_LIKE_RE.fullmatch(value):
        return "短值/数字变体"
    if value.startswith("0") and "Cr" in value:
        return "国标不锈钢牌号"
    if "Cr" in value or "Mo" in value or "Mn" in value:
        return "合金牌号"
    if "/" in value or ";" in value or "+" in value:
        return "组合材质"
    if SHORT_TOKEN_RE.fullmatch(value):
        return "短字面"
    return "其他"


def _strip_target_suffixes(value: str) -> tuple[str, list[str]]:
    base = value.strip()
    suffixes: list[str] = []
    changed = True
    while changed:
        changed = False
        for suffix in ("ZN", "CE", "III", "II"):
            if base.upper().endswith(suffix) and len(base) > len(suffix):
                base = base[: -len(suffix)]
                suffixes.append(suffix)
                changed = True
                break
    return base, suffixes


def _infer_surface_suffixes(text: str) -> list[str]:
    value = text.strip()
    suffixes: list[str] = []
    if ZN_RE.search(value):
        suffixes.append("ZN")
    if CE_RE.search(value):
        suffixes.append("CE")
    if GRADE_III_RE.search(value):
        suffixes.append("III")
    elif GRADE_II_RE.search(value):
        suffixes.append("II")
    return suffixes


def _build_rule_candidates(raw: str) -> list[str]:
    value = raw.strip()
    candidates: list[str] = [value]
    upper = value.upper()

    def add(candidate: str) -> None:
        candidate = candidate.strip()
        if candidate and candidate not in candidates:
            candidates.append(candidate)

    # ASTM/ASME 等标准前缀可省略，保留主体牌号候选
    for prefix in ("ASTM ", "ASME ", "AISI ", "JIS "):
        if upper.startswith(prefix):
            core = value[len(prefix) :].strip()
            add(core)
            add(re.sub(r"\s+", " ", core))
            add(re.sub(r"\s*-\s*", "-", core))
            add(re.sub(r"\s+", "", core))
            break

    # 常见 A/F/WP 体系，允许空格与连接符变体
    if re.match(r"^[AF]\d+", value, re.IGNORECASE) or re.match(r"^[AW]\d+", value, re.IGNORECASE):
        add(re.sub(r"\s+", " ", value))
        add(re.sub(r"\s*-\s*", "-", value))
        add(re.sub(r"\s+", "", value))

    return candidates


def _classify_entry(
    raw: str,
    total_count: int,
    in_desc_count: int,
    mapping: Counter[str],
    examples: list[str],
) -> MaterialEntry:
    dominant_target, dominant_count = mapping.most_common(1)[0]
    consistency = dominant_count / in_desc_count if in_desc_count else 0.0
    shape = _shape_of(raw)
    dominant_base, dominant_target_suffixes = _strip_target_suffixes(dominant_target)
    raw_suffixes = _infer_surface_suffixes(raw)
    base_counter: Counter[str] = Counter()
    for target, count in mapping.items():
        base, _ = _strip_target_suffixes(target)
        base_counter[base] += count
    dominant_base_count = max(base_counter.values()) if base_counter else 0
    dominant_base_ratio = dominant_base_count / in_desc_count if in_desc_count else 0.0
    all_base_same = len(base_counter) == 1
    only_suffix_variant = len(mapping) > 1 and all_base_same
    dominant_suffixes = list(dict.fromkeys(raw_suffixes + dominant_target_suffixes))

    if only_suffix_variant and dominant_base_ratio >= STRONG_RATIO:
        category = "主体可规则_后缀另处理"
        reason = "主体材质稳定，差异主要来自 ZN/CE/等级 后缀"
    elif HAS_CJK_RE.search(raw):
        category = "兜底规则"
        reason = "原始材质含中文，表面写法波动大，不建议直接进入强/弱规则"
    elif len(mapping) > 1 or consistency < STRONG_RATIO:
        category = "不建议规则"
        if len(mapping) > 1:
            if dominant_base_ratio >= STRONG_RATIO:
                reason = "主体材质较稳定，但存在复合/特殊后缀差异，不能直接做最终覆盖"
            else:
                reason = "同一原始材质映射到多个不同主体材质，存在真实歧义"
        else:
            reason = "同一原始材质一致性不足，不适合直接规则覆盖"
    elif in_desc_count < MIN_COUNT:
        category = "样本过少"
        reason = "描述中命中样本过少，先不建议上规则"
    elif (
        PURE_NUMERIC_RE.fullmatch(raw)
        or raw in {"20", "20#", "304", "304L", "316", "316L"}
        or shape in {"纯数字", "短值/数字变体", "短字面"}
    ):
        category = "弱规则"
        reason = "字面值过短或过泛，需要依赖上下文，不能直接全局覆盖"
    else:
        category = "可规则"
        reason = "字面值稳定、描述中可直接命中、最终材质一致性高"

    return MaterialEntry(
        raw=raw,
        total_count=total_count,
        in_desc_count=in_desc_count,
        dominant_target=dominant_target,
        consistency=round(consistency, 4),
        all_targets=dict(mapping),
        examples=examples,
        category=category,
        reason=reason,
        shape=shape,
        dominant_base=dominant_base,
        dominant_suffixes=dominant_suffixes,
        all_bases=dict(base_counter),
        rule_candidates=_build_rule_candidates(raw),
    )


def build_analysis(input_csv: Path) -> tuple[list[MaterialEntry], Counter[str]]:
    raw_counter: Counter[str] = Counter()
    final_counter: Counter[str] = Counter()
    raw_in_desc_counter: Counter[str] = Counter()
    raw_in_desc_to_final: dict[str, Counter[str]] = defaultdict(Counter)
    examples: dict[str, list[str]] = defaultdict(list)

    with input_csv.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            desc = (row.get("材料描述(多行)") or row.get("材料描述") or "").strip()
            raw = (row.get("原始材质") or "").strip()
            std = (row.get("标准化材质") or "").strip()
            fix = (row.get("修正材质") or "").strip()
            final = fix or std

            if raw:
                raw_counter[raw] += 1
            if final:
                final_counter[final] += 1

            if raw and final and raw in desc:
                raw_in_desc_counter[raw] += 1
                raw_in_desc_to_final[raw][final] += 1
                if len(examples[raw]) < 2:
                    examples[raw].append(desc)

    entries: list[MaterialEntry] = []
    for raw, in_desc_count in raw_in_desc_counter.items():
        entries.append(
            _classify_entry(
                raw=raw,
                total_count=raw_counter[raw],
                in_desc_count=in_desc_count,
                mapping=raw_in_desc_to_final[raw],
                examples=examples.get(raw, []),
            )
        )

    entries.sort(key=lambda item: (-item.in_desc_count, item.raw))
    return entries, final_counter


def _write_sheet(ws, headers: list[str], rows: list[list[object]]) -> None:
    ws.append(headers)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", horizontal="center")
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)


def write_excel(entries: list[MaterialEntry], final_counter: Counter[str], output_xlsx: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    headers = [
        "原始材质",
        "总出现次数",
        "描述中命中次数",
        "主映射结果",
        "主体材质",
        "后缀语义",
        "一致性",
        "规则建议",
        "值形态",
        "判断依据",
        "建议字面候选",
        "主体材质分布",
        "全部映射分布",
        "示例1",
        "示例2",
    ]

    def entry_row(entry: MaterialEntry) -> list[object]:
        base_targets = " | ".join(
            f"{target}:{count}"
            for target, count in sorted(
                entry.all_bases.items(), key=lambda item: (-item[1], item[0])
            )
        )
        targets = " | ".join(
            f"{target}:{count}"
            for target, count in sorted(
                entry.all_targets.items(), key=lambda item: (-item[1], item[0])
            )
        )
        ex1 = entry.examples[0] if len(entry.examples) > 0 else ""
        ex2 = entry.examples[1] if len(entry.examples) > 1 else ""
        return [
            entry.raw,
            entry.total_count,
            entry.in_desc_count,
            entry.dominant_target,
            entry.dominant_base,
            "/".join(entry.dominant_suffixes),
            entry.consistency,
            entry.category,
            entry.shape,
            entry.reason,
            " | ".join(entry.rule_candidates),
            base_targets,
            targets,
            ex1,
            ex2,
        ]

    all_rows = [entry_row(entry) for entry in entries]
    _write_sheet(wb.create_sheet("全部材质分析"), headers, all_rows)

    _write_sheet(
        wb.create_sheet("可规则"),
        headers,
        [entry_row(entry) for entry in entries if entry.category == "可规则"],
    )
    _write_sheet(
        wb.create_sheet("弱规则"),
        headers,
        [entry_row(entry) for entry in entries if entry.category == "弱规则"],
    )
    _write_sheet(
        wb.create_sheet("兜底规则"),
        headers,
        [entry_row(entry) for entry in entries if entry.category == "兜底规则"],
    )
    _write_sheet(
        wb.create_sheet("主体可规则_后缀另处理"),
        headers,
        [entry_row(entry) for entry in entries if entry.category == "主体可规则_后缀另处理"],
    )
    _write_sheet(
        wb.create_sheet("不建议规则"),
        headers,
        [entry_row(entry) for entry in entries if entry.category == "不建议规则"],
    )

    top_final_headers = ["最终材质", "次数"]
    top_final_rows = [[value, count] for value, count in final_counter.most_common(200)]
    _write_sheet(wb.create_sheet("最终材质统计"), top_final_headers, top_final_rows)

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


def main() -> None:
    parser = argparse.ArgumentParser(description="从 materials_export.csv 生成材质规则分析 Excel")
    parser.add_argument("--input-csv", default=str(DEFAULT_INPUT))
    parser.add_argument("--output-xlsx", default=str(DEFAULT_OUTPUT))
    args = parser.parse_args()

    input_csv = Path(args.input_csv)
    output_xlsx = Path(args.output_xlsx)

    entries, final_counter = build_analysis(input_csv)
    write_excel(entries, final_counter, output_xlsx)
    print(output_xlsx)


if __name__ == "__main__":
    main()
