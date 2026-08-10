#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""分析材料编码准确率、分流难度与模型置信分的关系。

脚本只读取原始 Excel/CSV，不调用模型，不修改输入文件。

默认字段：
- 正确答案：本项目材料代码
- 模型编码：excel2_原始总编码
- 分流难度：excel2_分流最终难度（0=困难，1=中等，2=简单）
- 分流原因：excel2_分流原因
- 模型置信分：excel2_模型置信分

示例：
python scripts/分析编码准确率难度与置信度.py \
    --input /Users/guoxi/Documents/法兰总数据-匹配后.xlsx \
    --project-col 项目名称 \
    --output /Users/guoxi/Documents/法兰编码分析.xlsx
"""

from __future__ import annotations

import argparse
import math
import re
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import pandas as pd
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DEFAULT_TRUTH_COL = "本项目材料代码"
DEFAULT_PRED_COL = "excel2_原始总编码"
DEFAULT_DIFFICULTY_COL = "excel2_分流最终难度（0=困难，1=中等，2=简单）"
DEFAULT_REASON_COL = "excel2_分流原因"
DEFAULT_CONFIDENCE_COL = "excel2_模型置信分"
PROJECT_COLUMN_CANDIDATES = ("项目名称", "项目", "所属项目", "项目名")

DIFFICULTY_ORDER = ("困难", "中等", "简单", "未知")
DIFFICULTY_MAP = {0: "困难", 1: "中等", 2: "简单"}

CONFIDENCE_BINS = [-1e-12, 0.50, 0.70, 0.80, 0.90, 0.95, 0.98, 0.99, 1.0000001]
CONFIDENCE_LABELS = [
    "0%~50%",
    "50%~70%",
    "70%~80%",
    "80%~90%",
    "90%~95%",
    "95%~98%",
    "98%~99%",
    "99%~100%",
]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"", "nan", "none", "null", "nat"} else text


def normalize_code(value: Any) -> str:
    """编码比较只忽略空白和大小写，不删除斜杠等业务字符。"""
    text = clean_text(value)
    return re.sub(r"\s+", "", text).upper() if text else ""


def parse_confidence(value: Any) -> float:
    text = clean_text(value)
    if not text:
        return math.nan
    is_percent = text.endswith("%")
    text = text.rstrip("%").strip().replace(",", "")
    try:
        number = float(text)
    except ValueError:
        return math.nan
    if is_percent or number > 1:
        number /= 100.0
    if not 0 <= number <= 1:
        return math.nan
    return number


def parse_difficulty(value: Any) -> tuple[float, str]:
    text = clean_text(value)
    if not text:
        return math.nan, "未知"

    normalized = text.lower()
    name_mapping = {
        "困难": 0,
        "hard": 0,
        "中等": 1,
        "中等难度": 1,
        "medium": 1,
        "简单": 2,
        "easy": 2,
    }
    if normalized in name_mapping:
        level = name_mapping[normalized]
        return float(level), DIFFICULTY_MAP[level]

    match = re.match(r"^\s*([012])(?:\s|$|[=（(])", text)
    if match:
        level = int(match.group(1))
        return float(level), DIFFICULTY_MAP[level]
    try:
        level = int(float(text))
    except ValueError:
        return math.nan, "未知"
    if level not in DIFFICULTY_MAP:
        return math.nan, "未知"
    return float(level), DIFFICULTY_MAP[level]


def read_table(path: str, sheet_name: str | None = None) -> pd.DataFrame:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return pd.read_csv(file_path, dtype=str).fillna("")
    if suffix in {".xlsx", ".xls"}:
        kwargs: dict[str, Any] = {"dtype": str}
        if sheet_name:
            kwargs["sheet_name"] = sheet_name
        return pd.read_excel(file_path, **kwargs).fillna("")
    raise ValueError(f"不支持的文件类型: {file_path.suffix}；仅支持 csv/xlsx/xls")


def validate_columns(df: pd.DataFrame, columns: Iterable[str]) -> None:
    missing = [column for column in columns if column not in df.columns]
    if missing:
        raise ValueError(f"输入文件缺少必需列: {missing}\n实际列: {list(df.columns)}")


def resolve_project_column(df: pd.DataFrame, requested: str | None) -> str:
    if requested:
        if requested not in df.columns:
            raise ValueError(f"指定的项目列不存在: {requested}")
        return requested
    for candidate in PROJECT_COLUMN_CANDIDATES:
        if candidate in df.columns:
            return candidate
    raise ValueError(
        "未找到项目列，请使用 --project-col 指定。"
        f"已尝试: {list(PROJECT_COLUMN_CANDIDATES)}"
    )


def prepare_analysis_data(
    df: pd.DataFrame,
    *,
    project_col: str,
    truth_col: str,
    pred_col: str,
    difficulty_col: str,
    reason_col: str,
    confidence_col: str,
) -> pd.DataFrame:
    result = df.copy()
    result["分析_项目"] = result[project_col].map(clean_text).replace("", "未指定项目")
    result["分析_正确答案归一"] = result[truth_col].map(normalize_code)
    result["分析_模型编码归一"] = result[pred_col].map(normalize_code)
    result["分析_正确答案有效"] = result["分析_正确答案归一"].ne("")
    result["分析_模型编码有效"] = result["分析_模型编码归一"].ne("")

    is_correct = (
        result["分析_正确答案有效"]
        & result["分析_模型编码有效"]
        & result["分析_正确答案归一"].eq(result["分析_模型编码归一"])
    )
    result["分析_是否正确"] = np.where(
        ~result["分析_正确答案有效"],
        "无正确答案",
        np.where(is_correct, "正确", "错误"),
    )
    result["分析_正确值"] = np.where(
        result["分析_正确答案有效"], is_correct.astype(int), np.nan
    )

    parsed_difficulty = result[difficulty_col].map(parse_difficulty)
    result["分析_难度值"] = parsed_difficulty.map(lambda item: item[0])
    result["分析_难度"] = parsed_difficulty.map(lambda item: item[1])
    result["分析_分流原因"] = result[reason_col].map(clean_text).replace("", "未提供原因")
    result["分析_模型置信分"] = result[confidence_col].map(parse_confidence)
    result["分析_置信分有效"] = result["分析_模型置信分"].notna()
    result["分析_置信分区间"] = pd.cut(
        result["分析_模型置信分"],
        bins=CONFIDENCE_BINS,
        labels=CONFIDENCE_LABELS,
        include_lowest=True,
        right=True,
    ).astype("object")
    result["分析_置信分区间"] = result["分析_置信分区间"].fillna("置信分缺失")
    return result


def confidence_stats(frame: pd.DataFrame, prefix: str = "") -> dict[str, Any]:
    values = frame["分析_模型置信分"].dropna().astype(float)
    label = f"{prefix}_" if prefix else ""
    if values.empty:
        return {
            f"{label}置信分有效数": 0,
            f"{label}平均置信分": math.nan,
            f"{label}中位置信分": math.nan,
            f"{label}P10置信分": math.nan,
            f"{label}P90置信分": math.nan,
        }
    return {
        f"{label}置信分有效数": int(values.size),
        f"{label}平均置信分": float(values.mean()),
        f"{label}中位置信分": float(values.median()),
        f"{label}P10置信分": float(values.quantile(0.10)),
        f"{label}P90置信分": float(values.quantile(0.90)),
    }


def safe_ratio(numerator: float, denominator: float) -> float:
    return float(numerator / denominator) if denominator else math.nan


def build_difficulty_summary(valid: pd.DataFrame) -> pd.DataFrame:
    total = len(valid)
    total_errors = int(valid["分析_正确值"].eq(0).sum())
    rows: list[dict[str, Any]] = []
    for difficulty in DIFFICULTY_ORDER:
        group = valid[valid["分析_难度"].eq(difficulty)]
        if group.empty and difficulty == "未知":
            continue
        correct = int(group["分析_正确值"].eq(1).sum())
        errors = int(group["分析_正确值"].eq(0).sum())
        row = {
            "难度": difficulty,
            "样本数": len(group),
            "样本占比": safe_ratio(len(group), total),
            "正确数": correct,
            "错误数": errors,
            "准确率": safe_ratio(correct, len(group)),
            "该难度错误率": safe_ratio(errors, len(group)),
            "占全部错误比例": safe_ratio(errors, total_errors),
            **confidence_stats(group),
            **confidence_stats(group[group["分析_正确值"].eq(1)], "正确样本"),
            **confidence_stats(group[group["分析_正确值"].eq(0)], "错误样本"),
        }
        rows.append(row)
    return pd.DataFrame(rows)


def build_correctness_summary(valid: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for label, value in (("正确", 1), ("错误", 0)):
        group = valid[valid["分析_正确值"].eq(value)]
        rows.append(
            {
                "结果": label,
                "样本数": len(group),
                "样本占比": safe_ratio(len(group), len(valid)),
                **confidence_stats(group),
            }
        )
    return pd.DataFrame(rows)


def build_confidence_calibration(valid: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, float]]:
    scored = valid[valid["分析_模型置信分"].notna()].copy()
    rows: list[dict[str, Any]] = []
    for label in CONFIDENCE_LABELS:
        group = scored[scored["分析_置信分区间"].eq(label)]
        if group.empty:
            rows.append(
                {
                    "置信分区间": label,
                    "样本数": 0,
                    "样本占比": 0.0,
                    "平均置信分": math.nan,
                    "实际准确率": math.nan,
                    "校准差值（置信分-准确率）": math.nan,
                    "绝对校准差": math.nan,
                    "错误数": 0,
                }
            )
            continue
        avg_conf = float(group["分析_模型置信分"].mean())
        accuracy = float(group["分析_正确值"].mean())
        rows.append(
            {
                "置信分区间": label,
                "样本数": len(group),
                "样本占比": safe_ratio(len(group), len(scored)),
                "平均置信分": avg_conf,
                "实际准确率": accuracy,
                "校准差值（置信分-准确率）": avg_conf - accuracy,
                "绝对校准差": abs(avg_conf - accuracy),
                "错误数": int(group["分析_正确值"].eq(0).sum()),
            }
        )

    calibration = pd.DataFrame(rows)
    if scored.empty:
        return calibration, {"ece": math.nan, "brier": math.nan, "auc": math.nan}

    ece = float(
        (
            calibration["样本占比"].fillna(0)
            * calibration["绝对校准差"].fillna(0)
        ).sum()
    )
    brier = float(
        np.mean(
            np.square(
                scored["分析_模型置信分"].astype(float)
                - scored["分析_正确值"].astype(float)
            )
        )
    )
    auc = confidence_auc(scored)
    return calibration, {"ece": ece, "brier": brier, "auc": auc}


def confidence_auc(scored: pd.DataFrame) -> float:
    positives = int(scored["分析_正确值"].eq(1).sum())
    negatives = int(scored["分析_正确值"].eq(0).sum())
    if not positives or not negatives:
        return math.nan
    ranks = scored["分析_模型置信分"].rank(method="average")
    positive_rank_sum = float(ranks[scored["分析_正确值"].eq(1)].sum())
    return float((positive_rank_sum - positives * (positives + 1) / 2) / (positives * negatives))


def build_confidence_quantiles(valid: pd.DataFrame) -> pd.DataFrame:
    scored = valid[valid["分析_模型置信分"].notna()].copy()
    if scored.empty:
        return pd.DataFrame(columns=["置信分分位区间", "样本数", "平均置信分", "实际准确率", "错误数"])
    try:
        scored["分位区间"] = pd.qcut(
            scored["分析_模型置信分"], q=10, duplicates="drop"
        ).astype(str)
    except ValueError:
        scored["分位区间"] = "无法分组"
    rows = []
    for label, group in scored.groupby("分位区间", sort=False, dropna=False):
        rows.append(
            {
                "置信分分位区间": str(label),
                "样本数": len(group),
                "平均置信分": float(group["分析_模型置信分"].mean()),
                "实际准确率": float(group["分析_正确值"].mean()),
                "错误数": int(group["分析_正确值"].eq(0).sum()),
            }
        )
    return pd.DataFrame(rows)


def build_threshold_summary(valid: pd.DataFrame, thresholds: Iterable[float]) -> pd.DataFrame:
    total = len(valid)
    total_errors = int(valid["分析_正确值"].eq(0).sum())
    rows = []
    for threshold in thresholds:
        accepted = valid[valid["分析_模型置信分"].ge(threshold)]
        accepted_errors = int(accepted["分析_正确值"].eq(0).sum())
        rows.append(
            {
                "自动通过阈值": threshold,
                "通过样本数": len(accepted),
                "覆盖率": safe_ratio(len(accepted), total),
                "通过样本准确率": (
                    float(accepted["分析_正确值"].mean()) if not accepted.empty else math.nan
                ),
                "高于阈值的错误数": accepted_errors,
                "错误漏过比例": safe_ratio(accepted_errors, total_errors),
                "需人工复核数": total - len(accepted),
            }
        )
    return pd.DataFrame(rows)


def build_project_summary(
    valid: pd.DataFrame,
    *,
    high_confidence_threshold: float,
    minimum_project_samples: int,
) -> pd.DataFrame:
    rows = []
    for project, group in valid.groupby("分析_项目", sort=False, dropna=False):
        correct = int(group["分析_正确值"].eq(1).sum())
        errors = int(group["分析_正确值"].eq(0).sum())
        scored = group[group["分析_模型置信分"].notna()]
        average_confidence = (
            float(scored["分析_模型置信分"].mean()) if not scored.empty else math.nan
        )
        high_conf_wrong = group[
            group["分析_正确值"].eq(0)
            & group["分析_模型置信分"].ge(high_confidence_threshold)
        ]
        row = {
            "项目": project,
            "样本数": len(group),
            "正确数": correct,
            "错误数": errors,
            "准确率": safe_ratio(correct, len(group)),
            "置信分覆盖率": safe_ratio(len(scored), len(group)),
            "平均置信分": average_confidence,
            "校准差值（置信分-准确率）": average_confidence - safe_ratio(correct, len(group)),
            "正确样本平均置信分": group.loc[
                group["分析_正确值"].eq(1), "分析_模型置信分"
            ].mean(),
            "错误样本平均置信分": group.loc[
                group["分析_正确值"].eq(0), "分析_模型置信分"
            ].mean(),
            f"高置信错误数(≥{high_confidence_threshold:.0%})": len(high_conf_wrong),
            "高置信错误占本项目错误比例": safe_ratio(len(high_conf_wrong), errors),
            "困难样本数": int(group["分析_难度"].eq("困难").sum()),
            "中等样本数": int(group["分析_难度"].eq("中等").sum()),
            "简单样本数": int(group["分析_难度"].eq("简单").sum()),
            "困难样本占比": float(group["分析_难度"].eq("困难").mean()),
            "中等样本占比": float(group["分析_难度"].eq("中等").mean()),
            "简单样本占比": float(group["分析_难度"].eq("简单").mean()),
            "小样本提示": "是" if len(group) < minimum_project_samples else "",
        }
        rows.append(row)
    return pd.DataFrame(rows).sort_values(["错误数", "样本数"], ascending=[False, False])


def build_project_difficulty_summary(valid: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for (project, difficulty), group in valid.groupby(["分析_项目", "分析_难度"], sort=False):
        rows.append(
            {
                "项目": project,
                "难度": difficulty,
                "样本数": len(group),
                "正确数": int(group["分析_正确值"].eq(1).sum()),
                "错误数": int(group["分析_正确值"].eq(0).sum()),
                "准确率": float(group["分析_正确值"].mean()),
                "平均置信分": group["分析_模型置信分"].mean(),
                "正确样本平均置信分": group.loc[
                    group["分析_正确值"].eq(1), "分析_模型置信分"
                ].mean(),
                "错误样本平均置信分": group.loc[
                    group["分析_正确值"].eq(0), "分析_模型置信分"
                ].mean(),
            }
        )
    return pd.DataFrame(rows)


def explode_reasons(valid: pd.DataFrame) -> pd.DataFrame:
    exploded_rows = []
    for _, row in valid.iterrows():
        reasons = [
            clean_text(item)
            for item in re.split(r"\s*\|\s*|\s*；\s*|\r?\n+", row["分析_分流原因"])
            if clean_text(item)
        ] or ["未提供原因"]
        for reason in dict.fromkeys(reasons):
            exploded_rows.append(
                {
                    "原因": reason,
                    "正确值": row["分析_正确值"],
                    "难度": row["分析_难度"],
                    "置信分": row["分析_模型置信分"],
                }
            )
    exploded = pd.DataFrame(exploded_rows)
    if exploded.empty:
        return pd.DataFrame(columns=["分流原因", "出现次数", "错误数", "该原因样本错误率", "平均置信分"])

    rows = []
    for reason, group in exploded.groupby("原因", sort=False):
        rows.append(
            {
                "分流原因": reason,
                "出现次数": len(group),
                "错误数": int(group["正确值"].eq(0).sum()),
                "该原因样本错误率": float(group["正确值"].eq(0).mean()),
                "平均置信分": group["置信分"].mean(),
                "困难数": int(group["难度"].eq("困难").sum()),
                "中等数": int(group["难度"].eq("中等").sum()),
                "简单数": int(group["难度"].eq("简单").sum()),
            }
        )
    return pd.DataFrame(rows).sort_values(["错误数", "出现次数"], ascending=[False, False])


def build_overview(
    data: pd.DataFrame,
    valid: pd.DataFrame,
    project_summary: pd.DataFrame,
    calibration_metrics: dict[str, float],
    *,
    high_confidence_threshold: float,
    low_confidence_threshold: float,
) -> pd.DataFrame:
    correct = int(valid["分析_正确值"].eq(1).sum())
    errors = int(valid["分析_正确值"].eq(0).sum())
    scored = valid[valid["分析_模型置信分"].notna()]
    high_conf_wrong = valid[
        valid["分析_正确值"].eq(0)
        & valid["分析_模型置信分"].ge(high_confidence_threshold)
    ]
    low_conf_correct = valid[
        valid["分析_正确值"].eq(1)
        & valid["分析_模型置信分"].le(low_confidence_threshold)
    ]
    macro_accuracy = project_summary["准确率"].mean() if not project_summary.empty else math.nan
    metrics = [
        ("输入总行数", len(data), "包含无正确答案的行"),
        ("有效分析样本数", len(valid), "正确答案非空"),
        ("无正确答案行数", len(data) - len(valid), "不参与准确率计算"),
        ("正确数", correct, ""),
        ("错误数", errors, "包含模型编码为空"),
        ("整体准确率（微平均）", safe_ratio(correct, len(valid)), "所有样本等权"),
        ("项目准确率宏平均", macro_accuracy, "每个项目等权"),
        ("项目数", valid["分析_项目"].nunique(), ""),
        ("置信分有效样本数", len(scored), ""),
        ("置信分覆盖率", safe_ratio(len(scored), len(valid)), ""),
        ("平均模型置信分", scored["分析_模型置信分"].mean(), ""),
        ("正确样本平均置信分", valid.loc[valid["分析_正确值"].eq(1), "分析_模型置信分"].mean(), ""),
        ("错误样本平均置信分", valid.loc[valid["分析_正确值"].eq(0), "分析_模型置信分"].mean(), ""),
        (f"高置信错误数(≥{high_confidence_threshold:.0%})", len(high_conf_wrong), "优先审核"),
        ("高置信错误占全部错误比例", safe_ratio(len(high_conf_wrong), errors), ""),
        (f"低置信正确数(≤{low_confidence_threshold:.0%})", len(low_conf_correct), ""),
        ("ECE校准误差", calibration_metrics.get("ece", math.nan), "越低越好"),
        ("Brier Score", calibration_metrics.get("brier", math.nan), "越低越好"),
        ("置信分区分AUC", calibration_metrics.get("auc", math.nan), "0.5接近无区分能力，越高越好"),
    ]
    return pd.DataFrame(metrics, columns=["指标", "数值", "说明"])


def build_findings(
    valid: pd.DataFrame,
    difficulty_summary: pd.DataFrame,
    calibration_metrics: dict[str, float],
    *,
    high_confidence_threshold: float,
) -> pd.DataFrame:
    findings: list[tuple[str, str]] = []
    accuracy = valid["分析_正确值"].mean()
    findings.append(("整体准确率", f"有效样本 {len(valid):,} 条，准确率 {accuracy:.2%}。"))

    difficulty_accuracy = {
        str(row["难度"]): row["准确率"]
        for _, row in difficulty_summary.iterrows()
        if pd.notna(row["准确率"])
    }
    ordered = [difficulty_accuracy.get(name) for name in ("困难", "中等", "简单")]
    if all(value is not None for value in ordered):
        monotonic = ordered[0] <= ordered[1] <= ordered[2]
        findings.append(
            ("难度区分度", f"困难/中等/简单准确率分别为 {ordered[0]:.2%}、{ordered[1]:.2%}、{ordered[2]:.2%}；"
             f"{'符合' if monotonic else '不符合'}预期的递增关系。")
        )

    correct_conf = valid.loc[valid["分析_正确值"].eq(1), "分析_模型置信分"].mean()
    wrong_conf = valid.loc[valid["分析_正确值"].eq(0), "分析_模型置信分"].mean()
    if pd.notna(correct_conf) and pd.notna(wrong_conf):
        findings.append(
            ("正误置信分差异", f"正确样本平均 {correct_conf:.2%}，错误样本平均 {wrong_conf:.2%}，"
             f"差值 {correct_conf - wrong_conf:+.2%}。")
        )

    high_conf_wrong = valid[
        valid["分析_正确值"].eq(0)
        & valid["分析_模型置信分"].ge(high_confidence_threshold)
    ]
    findings.append(
        ("高置信错误", f"置信分不低于 {high_confidence_threshold:.0%} 但编码错误的样本有 {len(high_conf_wrong):,} 条。")
    )

    ece = calibration_metrics.get("ece", math.nan)
    auc = calibration_metrics.get("auc", math.nan)
    if pd.notna(ece):
        findings.append(("置信度校准", f"ECE 为 {ece:.2%}；该值表示分箱后置信分与实际准确率的加权平均差距。"))
    if pd.notna(auc):
        findings.append(("置信分排序能力", f"正误区分 AUC 为 {auc:.4f}；0.5 表示几乎无法用置信分区分正误。"))
    return pd.DataFrame(findings, columns=["分析项", "结论"])


def build_invalid_data(data: pd.DataFrame) -> pd.DataFrame:
    invalid = data[
        ~data["分析_正确答案有效"]
        | ~data["分析_置信分有效"]
        | data["分析_难度"].eq("未知")
    ].copy()
    if invalid.empty:
        return invalid

    def reason(row: pd.Series) -> str:
        reasons = []
        if not row["分析_正确答案有效"]:
            reasons.append("正确答案为空")
        if not row["分析_置信分有效"]:
            reasons.append("置信分为空或格式无效")
        if row["分析_难度"] == "未知":
            reasons.append("难度为空或格式无效")
        return " | ".join(reasons)

    invalid.insert(0, "分析_异常原因", invalid.apply(reason, axis=1))
    return invalid


def build_display_detail(
    frame: pd.DataFrame,
    original_columns: Iterable[str],
) -> pd.DataFrame:
    """对报告只展示用户需要的分析字段，隐藏内部计算中间列。"""
    leading_columns = [
        "分析_异常原因",
        "分析_项目",
        "分析_是否正确",
        "分析_难度",
        "分析_模型置信分",
        "分析_分流原因",
    ]
    selected = []
    for column in [*leading_columns, *original_columns]:
        if column in frame.columns and column not in selected:
            selected.append(column)
    return frame.loc[:, selected].copy()


def output_path_for(input_path: str, requested: str | None) -> Path:
    if requested:
        path = Path(requested)
        return path if path.suffix.lower() == ".xlsx" else path.with_suffix(".xlsx")
    source = Path(input_path)
    return source.with_name(f"{source.stem}_准确率难度置信度分析.xlsx")


def _is_percent_column(name: str) -> bool:
    if any(keyword in name for keyword in ("数", "数量", "样本")):
        return False
    return any(keyword in name for keyword in ("率", "占比", "置信分", "ECE", "校准差")) and "Brier" not in name


def _is_percent_metric(name: str) -> bool:
    if any(keyword in name for keyword in ("数", "数量", "样本")):
        return False
    return any(keyword in name for keyword in ("率", "占比", "置信分", "ECE", "校准差"))


def format_workbook(
    writer: pd.ExcelWriter,
    section_positions: dict[str, list[dict[str, Any]]],
) -> None:
    workbook = writer.book
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    section_font = Font(color="17365D", bold=True, size=13)
    description_fill = PatternFill("solid", fgColor="EDF4F9")
    stripe_fill = PatternFill("solid", fgColor="F7FAFC")
    error_fill = PatternFill("solid", fgColor="FCE4D6")
    correct_fill = PatternFill("solid", fgColor="E2F0D9")
    thin_side = Side(style="thin", color="B7C9D6")
    table_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for worksheet in workbook.worksheets:
        positions = section_positions.get(worksheet.title, [])
        worksheet.freeze_panes = f"A{positions[0]['header_row'] + 1}" if positions else "A2"
        worksheet.sheet_view.showGridLines = False

        for position in positions:
            title_row = position["title_row"]
            description_row = position["description_row"]
            header_row = position["header_row"]
            end_row = position["end_row"]
            max_col = position["max_col"]

            for column in range(1, max_col + 1):
                title_cell = worksheet.cell(title_row, column)
                title_cell.fill = section_fill
                title_cell.border = table_border
                description_cell = worksheet.cell(description_row, column)
                description_cell.fill = description_fill
                description_cell.border = table_border
            worksheet.cell(title_row, 1).font = section_font
            worksheet.cell(title_row, 1).alignment = Alignment(vertical="center")
            worksheet.cell(description_row, 1).font = Font(color="44546A", italic=True)
            worksheet.cell(description_row, 1).alignment = Alignment(vertical="center", wrap_text=True)
            worksheet.row_dimensions[title_row].height = 25
            worksheet.row_dimensions[description_row].height = 32

            headers: dict[int, str] = {}
            for column in range(1, max_col + 1):
                cell = worksheet.cell(header_row, column)
                headers[column] = clean_text(cell.value)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = table_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            worksheet.row_dimensions[header_row].height = 34

            for row in range(header_row + 1, end_row + 1):
                for column in range(1, max_col + 1):
                    cell = worksheet.cell(row, column)
                    cell.border = table_border
                    cell.alignment = Alignment(vertical="top", wrap_text=False)
                    if (row - header_row) % 2 == 0:
                        cell.fill = stripe_fill
                    if _is_percent_column(headers[column]) and isinstance(cell.value, (int, float)):
                        cell.number_format = "0.00%"

                metric = clean_text(worksheet.cell(row, 1).value)
                if headers.get(1) == "指标" and _is_percent_metric(metric):
                    value_cell = worksheet.cell(row, 2)
                    if isinstance(value_cell.value, (int, float)):
                        value_cell.number_format = "0.00%"

                result_column = next(
                    (column for column, header in headers.items() if header == "分析_是否正确"),
                    None,
                )
                if result_column:
                    result = clean_text(worksheet.cell(row, result_column).value)
                    fill = error_fill if result == "错误" else correct_fill if result == "正确" else None
                    if fill:
                        for column in range(1, max_col + 1):
                            worksheet.cell(row, column).fill = fill

        if len(positions) == 1:
            position = positions[0]
            worksheet.auto_filter.ref = (
                f"A{position['header_row']}:{get_column_letter(position['max_col'])}{position['end_row']}"
            )

        for column_index, cells in enumerate(worksheet.iter_cols(), start=1):
            header = clean_text(cells[0].value)
            sampled_values = [clean_text(cell.value) for cell in cells[: min(len(cells), 300)]]
            max_length = max([len(header), *(len(value) for value in sampled_values)], default=8)
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 10), 48)


def add_charts(
    writer: pd.ExcelWriter,
    section_positions: dict[str, list[dict[str, Any]]],
) -> None:
    workbook = writer.book
    if "02_总体与难度" in workbook.sheetnames:
        sheet = workbook["02_总体与难度"]
        position = next(
            (item for item in section_positions[sheet.title] if item["title"] == "按难度统计"),
            None,
        )
        headers = {
            sheet.cell(position["header_row"], column).value: column
            for column in range(1, position["max_col"] + 1)
        } if position else {}
        accuracy_col = headers.get("准确率")
        if accuracy_col and position:
            chart = BarChart()
            chart.title = "各难度实际准确率"
            chart.y_axis.title = "准确率"
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = 1
            chart.add_data(
                Reference(sheet, min_col=accuracy_col, min_row=position["header_row"], max_row=position["end_row"]),
                titles_from_data=True,
            )
            chart.set_categories(
                Reference(sheet, min_col=1, min_row=position["header_row"] + 1, max_row=position["end_row"])
            )
            chart.height = 7
            chart.width = 12
            sheet.add_chart(chart, f"N{position['title_row']}")

    if "04_置信度分析" in workbook.sheetnames:
        sheet = workbook["04_置信度分析"]
        position = next(
            (item for item in section_positions[sheet.title] if item["title"] == "按置信分区间"),
            None,
        )
        headers = {
            sheet.cell(position["header_row"], column).value: column
            for column in range(1, position["max_col"] + 1)
        } if position else {}
        avg_col = headers.get("平均置信分")
        accuracy_col = headers.get("实际准确率")
        if avg_col and accuracy_col and position:
            chart = LineChart()
            chart.title = "置信分与实际准确率"
            chart.y_axis.title = "比例"
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = 1
            chart.add_data(
                Reference(
                    sheet,
                    min_col=avg_col,
                    max_col=accuracy_col,
                    min_row=position["header_row"],
                    max_row=position["end_row"],
                ),
                titles_from_data=True,
                from_rows=False,
            )
            chart.set_categories(
                Reference(sheet, min_col=1, min_row=position["header_row"] + 1, max_row=position["end_row"])
            )
            chart.height = 7
            chart.width = 14
            sheet.add_chart(chart, f"J{position['title_row']}")


def write_analysis_workbook(
    output_path: Path,
    sheets: dict[str, pd.DataFrame | list[tuple[str, str, pd.DataFrame]]],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        section_positions: dict[str, list[dict[str, Any]]] = {}
        for raw_sheet_name, content in sheets.items():
            sheet_name = raw_sheet_name[:31]
            sections = content if isinstance(content, list) else [(raw_sheet_name, "", content)]
            cursor = 0
            section_positions[sheet_name] = []
            for title, description, frame in sections:
                frame = frame.copy()
                frame.to_excel(writer, sheet_name=sheet_name, startrow=cursor + 2, index=False)
                worksheet = writer.book[sheet_name]
                worksheet.cell(cursor + 1, 1, title)
                worksheet.cell(cursor + 2, 1, description)
                max_col = max(len(frame.columns), 1)
                header_row = cursor + 3
                end_row = header_row + max(len(frame), 1)
                section_positions[sheet_name].append(
                    {
                        "title": title,
                        "title_row": cursor + 1,
                        "description_row": cursor + 2,
                        "header_row": header_row,
                        "end_row": end_row,
                        "max_col": max_col,
                    }
                )
                cursor = end_row + 2
        format_workbook(writer, section_positions)
        add_charts(writer, section_positions)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="分析材料编码准确率、分流难度和模型置信分")
    parser.add_argument("--input", required=True, help="输入 csv/xlsx/xls 路径")
    parser.add_argument("--sheet", help="Excel sheet 名，默认读取第一个 sheet")
    parser.add_argument("--output", help="输出 xlsx 路径")
    parser.add_argument("--project-col", help="项目列名；不传时自动尝试常见列名")
    parser.add_argument("--truth-col", default=DEFAULT_TRUTH_COL, help="正确答案列")
    parser.add_argument("--pred-col", default=DEFAULT_PRED_COL, help="本次模型编码列")
    parser.add_argument("--difficulty-col", default=DEFAULT_DIFFICULTY_COL, help="分流难度列")
    parser.add_argument("--reason-col", default=DEFAULT_REASON_COL, help="分流原因列")
    parser.add_argument("--confidence-col", default=DEFAULT_CONFIDENCE_COL, help="模型置信分列")
    parser.add_argument("--high-confidence-threshold", type=float, default=0.95, help="高置信错误阈值，默认 0.95")
    parser.add_argument("--low-confidence-threshold", type=float, default=0.60, help="低置信正确阈值，默认 0.60")
    parser.add_argument("--minimum-project-samples", type=int, default=30, help="项目小样本提示阈值，默认 30")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    for name, value in (
        ("--high-confidence-threshold", args.high_confidence_threshold),
        ("--low-confidence-threshold", args.low_confidence_threshold),
    ):
        if not 0 <= value <= 1:
            raise ValueError(f"{name} 必须在 0~1 之间")

    source = read_table(args.input, args.sheet)
    project_col = resolve_project_column(source, args.project_col)
    validate_columns(
        source,
        [
            project_col,
            args.truth_col,
            args.pred_col,
            args.difficulty_col,
            args.reason_col,
            args.confidence_col,
        ],
    )
    data = prepare_analysis_data(
        source,
        project_col=project_col,
        truth_col=args.truth_col,
        pred_col=args.pred_col,
        difficulty_col=args.difficulty_col,
        reason_col=args.reason_col,
        confidence_col=args.confidence_col,
    )
    valid = data[data["分析_正确答案有效"]].copy()
    if valid.empty:
        raise ValueError(f"正确答案列 {args.truth_col!r} 中没有可用于分析的有效编码")
    difficulty_summary = build_difficulty_summary(valid)
    correctness_summary = build_correctness_summary(valid)
    calibration, calibration_metrics = build_confidence_calibration(valid)
    threshold_summary = build_threshold_summary(
        valid,
        thresholds=(0.50, 0.60, 0.70, 0.80, 0.90, 0.95, 0.98, 0.99, 0.995),
    )
    project_summary = build_project_summary(
        valid,
        high_confidence_threshold=args.high_confidence_threshold,
        minimum_project_samples=args.minimum_project_samples,
    )
    project_difficulty = build_project_difficulty_summary(valid)
    reason_summary = explode_reasons(valid)
    overview = build_overview(
        data,
        valid,
        project_summary,
        calibration_metrics,
        high_confidence_threshold=args.high_confidence_threshold,
        low_confidence_threshold=args.low_confidence_threshold,
    )
    findings = build_findings(
        valid,
        difficulty_summary,
        calibration_metrics,
        high_confidence_threshold=args.high_confidence_threshold,
    )

    high_conf_wrong = valid[
        valid["分析_正确值"].eq(0)
        & valid["分析_模型置信分"].ge(args.high_confidence_threshold)
    ].sort_values("分析_模型置信分", ascending=False)
    low_conf_correct = valid[
        valid["分析_正确值"].eq(1)
        & valid["分析_模型置信分"].le(args.low_confidence_threshold)
    ].sort_values("分析_模型置信分", ascending=True)
    simple_wrong = valid[
        valid["分析_正确值"].eq(0) & valid["分析_难度"].eq("简单")
    ].sort_values("分析_模型置信分", ascending=False)
    invalid_data = build_invalid_data(data)

    difficulty_report = difficulty_summary.loc[
        :,
        [
            "难度",
            "样本数",
            "样本占比",
            "正确数",
            "错误数",
            "准确率",
            "该难度错误率",
            "占全部错误比例",
            "平均置信分",
            "正确样本_平均置信分",
            "错误样本_平均置信分",
        ],
    ]
    project_report_columns = [
        "项目",
        "样本数",
        "正确数",
        "错误数",
        "准确率",
        "平均置信分",
        "校准差值（置信分-准确率）",
        "正确样本平均置信分",
        "错误样本平均置信分",
        f"高置信错误数(≥{args.high_confidence_threshold:.0%})",
        "困难样本数",
        "中等样本数",
        "简单样本数",
        "小样本提示",
    ]
    project_report = project_summary.loc[:, project_report_columns]

    original_columns = list(source.columns)
    high_conf_wrong_report = build_display_detail(high_conf_wrong, original_columns)
    simple_wrong_report = build_display_detail(simple_wrong, original_columns)
    low_conf_correct_report = build_display_detail(low_conf_correct, original_columns)
    all_detail_report = build_display_detail(data, original_columns)
    invalid_report = build_display_detail(invalid_data, original_columns)

    reading_guide = pd.DataFrame(
        [
            (1, "02_总体与难度", "先看整体准确率，再看困难/中等/简单的准确率是否递增。"),
            (2, "03_项目分析", "找出错误数多、准确率低或高置信错误多的项目。"),
            (3, "04_置信度分析", "判断置信分能否区分正误，并选择自动通过阈值。"),
            (4, "05_重点样本", "优先审核高置信错误和被分为简单但编码错误的样本。"),
            (5, "06_全部明细", "需要回到某条原始数据时使用，可按正误、难度、置信分筛选。"),
        ],
        columns=["阅读顺序", "工作表", "用途"],
    )

    notes = pd.DataFrame(
        [
            ("编码比较", "只忽略空白和大小写；斜杠、连字符等业务字符不会被删除。"),
            ("空正确答案", "不参与准确率、校准度和项目指标计算。"),
            ("空模型编码", "正确答案有值时，计为编码错误。"),
            ("该难度错误率", "该难度内错误样本的比例，用于判断分流结果是否合理。"),
            ("占全部错误比例", "全部错误样本中，该难度所占比例。"),
            ("校准差值", "平均置信分减去实际准确率；为正表示模型过度自信。"),
            ("ECE", "置信分与实际准确率的总体差距，越接近 0 越好。"),
            ("AUC", "置信分区分正确与错误的能力；接近 0.5 表示几乎无区分能力。"),
            ("分流原因", "一条数据含多个原因时会拆分统计，因此原因次数之和可能大于样本数。"),
            ("项目小样本", f"项目样本数小于 {args.minimum_project_samples} 时会标记，不建议直接横向排名。"),
        ],
        columns=["项目", "说明"],
    )

    sheets: dict[str, pd.DataFrame | list[tuple[str, str, pd.DataFrame]]] = {
        "01_阅读说明": [
            ("建议阅读顺序", "按下列顺序查看即可，不需要逐张理解所有中间统计。", reading_guide),
            ("自动分析结论", "对本次数据中最重要的现象进行文字概括。", findings),
            ("指标解释", "仅在不理解某个指标时查看。", notes),
        ],
        "02_总体与难度": [
            ("总体结果", "先看整体准确率、错误数和高置信错误数。", overview),
            ("按难度统计", "重点比较困难/中等/简单的准确率；理想情况下应随难度降低而上升。", difficulty_report),
            ("按分流原因统计", "错误数多或错误率高的原因，是优先调整的分流规则。", reason_summary),
        ],
        "03_项目分析": [
            ("按项目汇总", "已按错误数从高到低排序，优先关注错误多且准确率低的项目。", project_report),
            ("项目与难度交叉", "用于判断某个项目的问题是集中在困难、中等还是简单样本。", project_difficulty),
        ],
        "04_置信度分析": [
            ("正确与错误样本对比", "如果正确样本的平均置信分没有明显高于错误样本，说明置信分区分能力较弱。", correctness_summary),
            ("按置信分区间", "比较每个区间的平均置信分与实际准确率，两者越接近越可信。", calibration),
            ("自动通过阈值模拟", "选定阈值后，覆盖率是可自动通过的比例；同时需关注高于阈值的错误数。", threshold_summary),
        ],
        "05_重点样本": [
            ("高置信错误", f"置信分≥{args.high_confidence_threshold:.0%}但编码错误，这类样本风险最高，建议优先审核。", high_conf_wrong_report),
            ("简单但编码错误", "分流判定为简单但实际编码错误，用于查找分流规则漏洞。", simple_wrong_report),
            ("低置信但编码正确", f"置信分≤{args.low_confidence_threshold:.0%}但编码正确，可用于判断置信分是否过度保守。", low_conf_correct_report),
        ],
        "06_全部明细": [
            ("全部分析明细", "蓝色列为本脚本新增的分析结果；可通过表头筛选正误、难度和置信分。", all_detail_report),
        ],
    }
    if not invalid_report.empty:
        sheets["07_无效数据"] = [
            ("缺失或无效数据", "这些行存在正确答案、难度或置信分缺失，部分指标不会纳入统计。", invalid_report),
        ]
    output_path = output_path_for(args.input, args.output)
    write_analysis_workbook(output_path, sheets)

    print(f"输入总行数: {len(data)}")
    print(f"有效分析样本: {len(valid)}")
    print(f"正确数: {int(valid['分析_正确值'].eq(1).sum())}")
    print(f"错误数: {int(valid['分析_正确值'].eq(0).sum())}")
    print(f"整体准确率: {valid['分析_正确值'].mean():.2%}")
    print(f"高置信错误数: {len(high_conf_wrong)}")
    print(f"输出文件: {output_path}")


if __name__ == "__main__":
    main()
