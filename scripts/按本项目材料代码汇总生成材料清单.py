# -*- coding: utf-8 -*-
"""
按“本项目材料代码”汇总 Excel，并生成材料清单。

输出样式：
- 同一“本项目材料代码”占一个分组
- 每条材料描述单独占一行
- 分组列纵向合并，效果接近用户示例图二

主工作表输出列：
1. 本项目材料代码
2. 数量
3. 编号（P-000001 起）
4. 项目名称（可选输入列，不是必传）
5. 材质分类（可选输入列，不是必传）
6. 材料描述

附加工作表：
- 编码差异匹配
- 读取“本次生成材料代码”列（若输入存在）
- 对比“本项目材料代码”和“本次生成材料代码”
- 若不一致，则尝试用“本次生成材料代码”回查主工作表的材料代码清单
- 若匹配到，则补出对应“材料代码编号”

排序规则（启发式）：
- 先按编码模板分组：把数字替换成占位符后，相同模板视为同类编码
- 同模板内优先按数量从大到小
- 再按字母/数字混合片段做自然排序，使差异较小的编码尽量排在一起

python scripts/按本项目材料代码汇总生成材料清单.py \
    /Users/guoxi/Documents/材料数据.xlsx \
    --output /Users/guoxi/Downloads/材料清单.xlsx

    
 python scripts/按本项目材料代码汇总生成材料清单.py \
    /Users/guoxi/Documents/材料数据.xlsx \
    --sheet Sheet1 \
    --output /Users/guoxi/Downloads/材料清单.xlsx
"""

from __future__ import annotations

import argparse
import re
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


CODE_COL = "本项目材料代码"
DESC_COL = "材料描述"
PROJECT_COL = "项目名称"
CATEGORY_COL = "材质分类"
NEW_CODE_COL = "本次生成材料代码"

OUTPUT_COLUMNS = [CODE_COL, "数量", "编号", PROJECT_COL, CATEGORY_COL, DESC_COL]
DIFF_SHEET_COLUMNS = [
    CODE_COL,
    "正确材料代码编号",
    NEW_CODE_COL,
    "是否一致",
    "匹配到材料清单",
    "匹配材料代码编号",
    PROJECT_COL,
    CATEGORY_COL,
    DESC_COL,
]


@dataclass
class GroupedCode:
    code: str
    descriptions: list[str] = field(default_factory=list)
    project_names: list[str] = field(default_factory=list)
    categories: list[str] = field(default_factory=list)
    row_items: list[tuple[str, str, str]] = field(default_factory=list)


def clean_text(value: Any) -> str:
    text = str(value or "").strip()
    return "" if text.lower() == "nan" else text


def load_excel(path: Path, sheet_name: str | int | None) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet_name if sheet_name is not None else 0, dtype=str).fillna("")
    missing = [col for col in (CODE_COL, DESC_COL) if col not in df.columns]
    if missing:
        raise ValueError(f"缺少必要列: {missing}；实际列为: {list(df.columns)}")
    return df


def split_code_segments(code: str) -> list[str]:
    return re.findall(r"[A-Za-z]+|\d+(?:\.\d+)?|[^A-Za-z\d]+", code)


def build_code_template(code: str) -> str:
    return re.sub(r"\d", "0", code.upper())


def natural_segment_key(segment: str) -> tuple[int, Any]:
    if re.fullmatch(r"\d+(?:\.\d+)?", segment):
        return (0, float(segment))
    return (1, segment.upper())


def code_sort_key(code: str) -> tuple[Any, ...]:
    normalized = clean_text(code).upper()
    segments = split_code_segments(normalized)
    return (
        build_code_template(normalized),
        len(segments),
        tuple(natural_segment_key(segment) for segment in segments),
        normalized,
    )


def group_sort_key(group: GroupedCode) -> tuple[Any, ...]:
    return (
        build_code_template(group.code),
        -len(group.row_items),
        *code_sort_key(group.code)[1:],
    )


def aggregate_rows(df: pd.DataFrame) -> list[GroupedCode]:
    grouped: dict[str, GroupedCode] = {}

    for _, row in df.iterrows():
        code = clean_text(row.get(CODE_COL, ""))
        desc = clean_text(row.get(DESC_COL, ""))
        project = clean_text(row.get(PROJECT_COL, "")) if PROJECT_COL in df.columns else ""
        category = clean_text(row.get(CATEGORY_COL, "")) if CATEGORY_COL in df.columns else ""
        if not code:
            continue

        bucket = grouped.setdefault(code, GroupedCode(code=code))
        bucket.descriptions.append(desc)
        if project:
            bucket.project_names.append(project)
        if category:
            bucket.categories.append(category)
        bucket.row_items.append((project, category, desc))

    groups = list(grouped.values())
    groups.sort(key=group_sort_key)
    return groups


def build_code_number_map(groups: list[GroupedCode]) -> dict[str, str]:
    return {group.code: f"P-{index:06d}" for index, group in enumerate(groups, start=1)}


def build_diff_sheet_rows(df: pd.DataFrame, code_number_map: dict[str, str]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    has_new_code_col = NEW_CODE_COL in df.columns

    for _, row in df.iterrows():
        project_code = clean_text(row.get(CODE_COL, ""))
        generated_code = clean_text(row.get(NEW_CODE_COL, "")) if has_new_code_col else ""
        description = clean_text(row.get(DESC_COL, ""))
        project_name = clean_text(row.get(PROJECT_COL, "")) if PROJECT_COL in df.columns else ""
        category = clean_text(row.get(CATEGORY_COL, "")) if CATEGORY_COL in df.columns else ""

        if not project_code or not generated_code:
            continue
        if project_code == generated_code:
            continue

        rows.append(
            {
                CODE_COL: project_code,
                "正确材料代码编号": code_number_map.get(project_code, ""),
                NEW_CODE_COL: generated_code,
                "是否一致": "否",
                "匹配到材料清单": generated_code if generated_code in code_number_map else "",
                "匹配材料代码编号": code_number_map.get(generated_code, ""),
                PROJECT_COL: project_name,
                CATEGORY_COL: category,
                DESC_COL: description,
            }
        )

    result = pd.DataFrame(rows, columns=DIFF_SHEET_COLUMNS)
    if not result.empty:
        result = result.sort_values(
            by=["匹配材料代码编号", NEW_CODE_COL, CODE_COL],
            key=lambda col: col.map(lambda value: clean_text(value) or "ZZZZZZ"),
            kind="stable",
        ).reset_index(drop=True)
    return result


def auto_output_path(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_材料清单.xlsx")


def apply_sheet_style(ws) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="F3F6FA")
    thin_side = Side(style="thin", color="D9E2F3")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    widths = {
        1: 42,
        2: 10,
        3: 14,
        4: 24,
        5: 18,
        6: 120,
    }
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def write_diff_sheet(ws, diff_df: pd.DataFrame) -> None:
    ws.append(DIFF_SHEET_COLUMNS)
    for row in diff_df.itertuples(index=False):
        ws.append(list(row))
    apply_sheet_style(ws)

    widths = {
        1: 42,
        2: 14,
        3: 42,
        4: 10,
        5: 42,
        6: 14,
        7: 20,
        8: 16,
        9: 120,
    }
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_excel(groups: list[GroupedCode], diff_df: pd.DataFrame, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    ws = workbook.active
    ws.title = "材料清单"
    ws.append(OUTPUT_COLUMNS)

    current_row = 2
    for index, group in enumerate(groups, start=1):
        descriptions = group.descriptions or [""]
        row_count = max(1, len(descriptions))
        start_row = current_row
        end_row = current_row + row_count - 1

        row_items = group.row_items or [("", "", "")]
        for offset, (project, category, desc) in enumerate(row_items):
            row_idx = current_row + offset
            ws.cell(row=row_idx, column=4, value=project or None)
            ws.cell(row=row_idx, column=5, value=category or None)
            ws.cell(row=row_idx, column=6, value=desc)

        ws.cell(row=start_row, column=1, value=group.code)
        ws.cell(row=start_row, column=2, value=len(group.descriptions))
        ws.cell(row=start_row, column=3, value=f"P-{index:06d}")

        if end_row > start_row:
            for col in range(1, 4):
                ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)

        current_row = end_row + 1

    apply_sheet_style(ws)
    diff_ws = workbook.create_sheet("编码差异匹配")
    write_diff_sheet(diff_ws, diff_df)
    workbook.save(output_path)


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="按本项目材料代码汇总生成材料清单")
    parser.add_argument("excel", help="输入 Excel 路径")
    parser.add_argument("--sheet", help="工作表名；默认第一个 sheet", default=None)
    parser.add_argument("--output", help="输出 Excel 路径；默认在原文件旁生成 *_材料清单.xlsx", default=None)
    return parser


def main() -> None:
    parser = build_arg_parser()
    args = parser.parse_args()

    input_path = Path(args.excel).expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"输入文件不存在: {input_path}")

    df = load_excel(input_path, args.sheet)
    groups = aggregate_rows(df)
    code_number_map = build_code_number_map(groups)
    diff_df = build_diff_sheet_rows(df, code_number_map)
    output_path = Path(args.output).expanduser().resolve() if args.output else auto_output_path(input_path)
    write_excel(groups, diff_df, output_path)

    print(f"已生成: {output_path}")
    print(f"共汇总 {len(groups)} 个唯一材料代码")
    print(f"编码不一致条数: {len(diff_df)}")


if __name__ == "__main__":
    main()
